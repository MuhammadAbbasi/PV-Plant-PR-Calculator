"""
Calcolatore Performance Ratio (PR) Fotovoltaico - Mazara 01
GET SRL

===============================================================================
STORICO VERSIONI / CHANGELOG
Aggiungere una voce in cima ad ogni modifica: data (AAAA-MM-GG) e cosa cambia.
===============================================================================

v15.0 - 2026-09-01 (UI/UX)
  * Console Live Log di nuovo visibile: era finita completamente fuori dalla
    finestra (servivano 1147 px in una finestra da 975). Ora la colonna destra
    contiene la tabella PVSyst in alto e il log sotto, che occupa lo spazio
    bianco prima sprecato (circa 420 px) sotto la tabella.
  * Tabella PVSyst: tutti i 12 mesi visibili (Dicembre era tagliato a meta').
    Causa: la card si ridimensionava sulla richiesta del contenuto PRIMA che le
    etichette a capo automatico aggiornassero la propria altezza, restando
    17 px corta. Corretto in RoundedCard, quindi vale per tutte le card.
  * Ambito "Intervallo": i campi Dal/Al giorno erano schiacciati fuori dal bordo
    della card (mezzo spinbox disegnato). Ora sono su una riga dedicata,
    mostrata solo quando serve, con etichette "Dal giorno" / "al giorno".
  * La riga "Giorni VCOM" non compare piu' all'avvio con sorgente SCADA o VCOM:
    all'avvio mancava la chiamata di sincronizzazione del suo stato.
  * Barra di avanzamento durante l'elaborazione batch (visibile solo mentre il
    calcolo e' in corso).
  * Testi di stato allineati a sinistra come tutti gli altri suggerimenti.
  * Verifica automatica: 'test_v15_layout.py'.

v15.0 - 2026-09-01
  * Nuovi pulsanti dedicati: "Sync SCADA PR" e "Sync VCOM PR" per individuare
    automaticamente i file esterni (KPI_Report_Daily.xls* e Performance_ratio*.csv /
    Performance_ratio_vcom.csv) e aggiornare direttamente le colonne PR SCADA e
    PR VCOM nel file Madre ('00 PR_recalculation_{MESE}.xlsx') senza dover
    rieseguire l'intero calcolo giornaliero.
  * Pattern matching flessibile per individuare file VCOM con qualsiasi data nel
    nome (es. 'Performance_ratio_2026_07_31.csv') e file SCADA.
  * Selezione manuale tramite finestra di dialogo se il file SCADA o VCOM
    non viene trovato automaticamente nella cartella mese.
  * Sincronizzazione atomica e protetta via Excel COM con backup automatico,
    gestione blocchi ROT/processi e ricalcolo formule.
  * Versione GUI e log aggiornati a v15.0.

v13.0 - 2026-08-31
  * AMBITO DI CALCOLO configurabile su cartella mensile: "Mese intero",
    "Solo 1 giorno" (usa il giorno del campo Data) oppure "Intervallo" (Da..A).
    In v12 la selezione di una cartella mensile elaborava sempre tutti i giorni:
    non era possibile ricalcolare un singolo giorno.
  * Nuova finestra "Opzioni Avanzate" con parametri prima cablati nel codice:
      - degradazione contrattuale del target (%/anno, anno e mese di avvio);
      - sincronizzazione del file Madre attivabile/disattivabile (è il passo
        piu' lento: disattivarla rende immediato il ricalcolo di pochi giorni);
      - backup automatico del file Madre attivabile/disattivabile;
      - i 12 target PR PVSyst mensili, editabili, con ripristino predefiniti.
  * Impostazioni persistenti in 'PR_Calculator_settings.json' accanto all'app
    (soglie, metodo POA, ambito, ultima cartella): non vanno piu' reimpostate
    ad ogni avvio.
  * Avviso a log se il campo Data non appartiene al mese della cartella scelta.
  * Altezza finestra adeguata alla nuova riga e limitata all'altezza schermo.
  * Verifica automatica: 'test_v13_scope.py'.

v12.0 - 2026-07-24
  * Rilevamento e conversione automatica dei dati VCOM per i giorni con file
    SCADA mancanti, con downloader integrato da meteocontrol.
  * Pulsante "Interrompi" con arresto sicuro fra un giorno e l'altro.
  * Colonna "Target Corretto" (degradazione contrattuale) nella tabella PVSyst.

v11.0 - 2026-06-18
  * Riferimento POA selezionabile (Conditional MAX / Media dei piranometri).
  * Riparazione delle serie contatore e recupero dei file Excel bloccati.
  * Backup automatico del file Madre prima della sincronizzazione.
"""

import os
import glob
import sys
import threading
import datetime
import pandas as pd
# pyrefly: ignore [missing-import]
import numpy as np
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from VCOM_to_SCADA import convert_vcom_to_scada
_excel_app = None

def get_resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

# --- v13 user settings -------------------------------------------------------
# Values that used to be hard-coded in the engine (the contractual degradation,
# the PVSyst monthly targets, the Madre sync/backup behaviour) live here and are
# persisted next to the app so they survive a restart.
DEFAULT_SETTINGS = {
    "pvsyst_monthly": {
        "1": 0.904, "2": 0.896, "3": 0.897, "4": 0.868, "5": 0.832, "6": 0.833,
        "7": 0.820, "8": 0.828, "9": 0.852, "10": 0.876, "11": 0.894, "12": 0.900,
    },
    "deg_rate": 0.004,       # contractual PR decay per year (Allegato 9.1)
    "deg_start_year": 2025,  # plant start: Feb 2025
    "deg_start_month": 2,
    "sync_mother": True,     # run the (slow) Madre sync after the calculation
    "backup_mother": True,   # snapshot the Madre file before modifying it
    "threshold": "50",
    "diff_threshold": "10",
    "poa_method": "average",
    "day_scope": "mese",
    "day_from": 1,
    "day_to": 31,
    "last_folder": "",
    "data_source": "misto",  # "scada", "vcom", or "misto"
    "vcom_days": "",         # comma-separated day numbers, e.g. "3, 7, 9"
}

def get_settings_path():
    """Writable location next to the .exe (frozen) or the .py (source). Never
    _MEIPASS, which is a temp dir wiped on exit."""
    base = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) \
        else os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "PR_Calculator_settings.json")

def load_settings():
    import json
    cfg = dict(DEFAULT_SETTINGS)
    cfg["pvsyst_monthly"] = dict(DEFAULT_SETTINGS["pvsyst_monthly"])
    try:
        with open(get_settings_path(), "r", encoding="utf-8") as fh:
            saved = json.load(fh)
        cfg["pvsyst_monthly"].update(saved.pop("pvsyst_monthly", None) or {})
        cfg.update({k: v for k, v in saved.items() if k in DEFAULT_SETTINGS})
    except FileNotFoundError:
        pass
    except Exception as e:
        print(f"DEBUG: impostazioni non caricate ({e}), uso i valori predefiniti.")
    return cfg

def save_settings(cfg):
    import json
    try:
        with open(get_settings_path(), "w", encoding="utf-8") as fh:
            json.dump(cfg, fh, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"DEBUG: salvataggio impostazioni fallito: {e}")
        return False

def get_excel_app():
    global _excel_app
    if _excel_app is None:
        import win32com.client
        _excel_app = win32com.client.DispatchEx("Excel.Application")
        _excel_app.Visible = False
        _excel_app.DisplayAlerts = False
        try:
            # v11 security: force-disable macros in any workbook opened via automation
            # (msoAutomationSecurityForceDisable = 3). Prevents auto-run macros in a
            # rogue/legacy .xlsm input from executing silently with alerts suppressed.
            _excel_app.AutomationSecurity = 3
        except Exception:
            pass
        try:
            _excel_app.Calculation = -4135  # xlCalculationManual
            _excel_app.CalculateBeforeSave = False
        except Exception:
            pass
    return _excel_app

def quit_excel_app():
    global _excel_app
    if _excel_app is not None:
        try:
            _excel_app.Quit()
        except Exception:
            pass
        _excel_app = None

class RedirectText:
    def __init__(self, root, text_widget, original_stream):
        self.root = root
        self.text_widget = text_widget
        self.original_stream = original_stream

    def write(self, string):
        try:
            self.original_stream.write(string)
        except Exception:
            pass
        try:
            self.root.after(0, self._insert_text, string)
        except Exception:
            pass

    def flush(self):
        try:
            self.original_stream.flush()
        except Exception:
            pass

    def _insert_text(self, string):
        try:
            self.text_widget.config(state="normal")
            self.text_widget.insert("end", string)
            self.text_widget.see("end")
            self.text_widget.config(state="disabled")
        except Exception:
            pass

class RoundedCard(tk.Canvas):
    def __init__(self, parent, bg="#ffffff", border_color="#dadce0", radius=12, padding=16, **kwargs):
        super().__init__(parent, bg="#f8f9fa", highlightthickness=0, bd=0, **kwargs)
        self.bg = bg
        self.border_color = border_color
        self.radius = radius
        self.padding = padding
        
        self.bind("<Configure>", self._draw)
        
        self.content_frame = tk.Frame(self, bg=self.bg, bd=0, highlightthickness=0)
        self.window_id = self.create_window(0, 0, window=self.content_frame, anchor="nw")
        self.content_frame.bind("<Configure>", self._on_content_configure)

    def _on_content_configure(self, event):
        margin = self.padding
        req_w = self.content_frame.winfo_reqwidth() + 2 * margin
        req_h = self.content_frame.winfo_reqheight() + 2 * margin
        self.config(width=req_w, height=req_h)
        # _wrap_label binds to this same <Configure> with add="+", so it runs AFTER this
        # handler and can still grow the content's requested height. Re-fit once the whole
        # event burst has drained, or the card stays short by the rewrapped lines.
        self.after_idle(self._sync_height)

    def _sync_height(self):
        """Re-fit the card to what its content currently asks for.

        _on_content_configure only fires when the content frame's ACTUAL size changes, and
        _draw pins that size here. So when a resize rewraps a label and the content's
        REQUESTED height grows, nothing resized the card and the last rows were clipped
        (the PVSyst table lost Dicembre). Converges: height changes never alter wrapping,
        which is driven by width."""
        try:
            need = self.content_frame.winfo_reqheight() + 2 * self.padding
            if need != int(self.cget("height")):
                self.config(height=need)
        except Exception:
            pass

    def _draw(self, event):
        self.delete("bg")
        w = event.width
        h = event.height
        r = self.radius
        
        # Offsets
        x0, y0, x1, y1 = 1, 1, w - 2, h - 2
        
        self._draw_round_rect(x0, y0, x1, y1, r, fill=self.border_color, outline=self.border_color, tags="bg")
        self._draw_round_rect(x0 + 1, y0 + 1, x1 - 1, y1 - 1, r, fill=self.bg, outline=self.bg, tags="bg")
        
        margin = self.padding
        self.coords(self.window_id, margin, margin)
        self.itemconfigure(self.window_id, width=w - 2 * margin, height=h - 2 * margin)
        # The new width may rewrap labels and grow the content's requested height.
        self.after_idle(self._sync_height)

    def _draw_round_rect(self, x0, y0, x1, y1, r, **kwargs):
        self.create_arc(x0, y0, x0 + 2*r, y0 + 2*r, start=90, extent=90, style="pieslice", **kwargs)
        self.create_arc(x1 - 2*r, y0, x1, y0 + 2*r, start=0, extent=90, style="pieslice", **kwargs)
        self.create_arc(x1 - 2*r, y1 - 2*r, x1, y1, start=270, extent=90, style="pieslice", **kwargs)
        self.create_arc(x0, y1 - 2*r, x0 + 2*r, y1, start=180, extent=90, style="pieslice", **kwargs)
        
        self.create_rectangle(x0 + r, y0, x1 - r, y1, **kwargs)
        self.create_rectangle(x0, y0 + r, x1, y1 - r, **kwargs)

class PRCalculatorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Calcolatore Performance Ratio (PR) Fotovoltaico Mazara 01 - v15.0")
        self.cfg = load_settings()
        
        # Center the window on the screen for a soft, user-oriented launch
        self.root.withdraw()
        self.root.update_idletasks()
        width = 1040
        # 975px is what the layout actually requests with the v13 scope row; clamp it so the
        # window never opens taller than the screen (and behind the taskbar) on 1080p panels.
        height = min(975, self.root.winfo_screenheight() - 90)
        x = (self.root.winfo_screenwidth() - width) // 2
        y = (self.root.winfo_screenheight() - height) // 2
        self.root.geometry(f"{width}x{height}+{x}+{y}")
        self.root.deiconify()
        self.root.configure(bg="#f8f9fa")
        
        # Ensure Windows taskbar and task manager correctly display the custom GET logo icon
        try:
            import ctypes
            myappid = 'get.srl.prcalculator.v11'
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
        except Exception:
            pass
            
        icon_path = get_resource_path(os.path.join("assets", "logo.ico"))
        if os.path.exists(icon_path):
            try:
                self.root.iconbitmap(icon_path)
            except Exception:
                pass
                
        logo_png_path = get_resource_path(os.path.join("assets", "logo.png"))
        if os.path.exists(logo_png_path):
            try:
                img_icon = tk.PhotoImage(file=logo_png_path)
                self.root.iconphoto(True, img_icon)
            except Exception:
                pass
        
        # Apply style
        self.style = ttk.Style()
        self.style.theme_use("clam")
        
        # Define Color Palette (Google Styled Clean Light Theme)
        self.bg_color = "#f8f9fa"      # Clean Light Gray background
        self.card_bg = "#ffffff"       # Pure White Surface
        self.accent_color = "#1a73e8"  # Google Blue Primary Accent
        self.accent_hover = "#1557b0"  # Darker Blue for hovers
        self.text_color = "#202124"    # Dark Charcoal for maximum readability
        self.muted_text = "#5f6368"    # Soft Slate Grey
        self.success_color = "#137333" # Google Green
        self.warn_color = "#b06000"    # Google Amber/Orange for secondary/uncompensated
        self.border_color = "#dadce0"  # Google Light Gray border
        
        # Custom Widget Configurations
        self.style.configure("TFrame", background=self.bg_color)
        self.style.configure("Card.TFrame", background=self.card_bg, relief="flat")
        self.style.configure("TLabel", background=self.bg_color, foreground=self.text_color, font=("Segoe UI", 10))
        self.style.configure("Card.TLabel", background=self.card_bg, foreground=self.text_color, font=("Segoe UI", 10))
        self.style.configure("Title.TLabel", background=self.bg_color, foreground=self.text_color, font=("Segoe UI Semibold", 16, "bold"))
        self.style.configure("Section.TLabel", background=self.card_bg, foreground=self.accent_color, font=("Segoe UI Semibold", 12, "bold"))
        self.style.configure("MetricVal.TLabel", background=self.card_bg, foreground=self.success_color, font=("Segoe UI", 22, "bold"))
        self.style.configure("MetricLbl.TLabel", background=self.card_bg, foreground=self.muted_text, font=("Segoe UI Semibold", 9))
        
        # Entry Field styling
        self.style.configure("TEntry", fieldbackground="#ffffff", background="#ffffff", foreground=self.text_color, borderwidth=1, bordercolor="#dadce0", padding=6)
        self.style.map("TEntry", 
                       bordercolor=[("focus", self.accent_color), ("active", self.accent_color)],
                       lightcolor=[("focus", self.accent_color)],
                       darkcolor=[("focus", self.accent_color)])
        
        # Button styling
        self.style.configure("TButton", background=self.accent_color, foreground="#ffffff", borderwidth=0, font=("Segoe UI Semibold", 10, "bold"), padding=[16, 8])
        self.style.map("TButton", background=[("active", self.accent_hover), ("disabled", "#f1f3f4")], foreground=[("disabled", "#94a3b8")])
        
        self.style.configure("Secondary.TButton", background="#ffffff", foreground=self.accent_color, bordercolor="#dadce0", darkcolor="#dadce0", lightcolor="#dadce0", borderwidth=1, font=("Segoe UI Semibold", 10), padding=[12, 6])
        self.style.map("Secondary.TButton", background=[("active", "#f8f9fa"), ("disabled", "#ffffff")], foreground=[("disabled", "#94a3b8")], bordercolor=[("active", self.accent_color)])
        
        self.style.configure("Action.TButton", background=self.accent_color, foreground="#ffffff", font=("Segoe UI Semibold", 11, "bold"), padding=[20, 10])
        self.style.map("Action.TButton", background=[("active", self.accent_hover)])
        
        # Safe-stop button: outlined Google-red, muted while idle/disabled
        self.style.configure("Stop.TButton", background="#ffffff", foreground="#d93025",
                             bordercolor="#f2b8b5", lightcolor="#f2b8b5", darkcolor="#f2b8b5",
                             borderwidth=1, font=("Segoe UI Semibold", 10), padding=[16, 7])
        self.style.map("Stop.TButton",
                       background=[("active", "#fce8e6"), ("disabled", "#ffffff")],
                       foreground=[("disabled", "#bdc1c6")],
                       bordercolor=[("active", "#d93025"), ("disabled", "#e8eaed")],
                       lightcolor=[("disabled", "#e8eaed")], darkcolor=[("disabled", "#e8eaed")])
        
        # Segmented toggle styling for the POA reference method (themed pill buttons).
        # Unselected: white surface with accent text and a light border. Selected: filled
        # accent background with white text so the active choice is obvious at a glance.
        self.style.configure("POA.Toolbutton", background="#ffffff", foreground=self.accent_color,
                             bordercolor=self.border_color, lightcolor=self.border_color, darkcolor=self.border_color,
                             borderwidth=1, relief="solid", font=("Segoe UI Semibold", 9), padding=[16, 7], anchor="center")
        self.style.map("POA.Toolbutton",
                       background=[("selected", self.accent_color), ("active", "#e8f0fe"), ("!selected", "#ffffff")],
                       foreground=[("selected", "#ffffff"), ("active", self.accent_color)],
                       bordercolor=[("selected", self.accent_color), ("active", self.accent_color)],
                       lightcolor=[("selected", self.accent_color)],
                       darkcolor=[("selected", self.accent_color)],
                       relief=[("selected", "solid")])
        
        # Checkbutton styling
        self.style.configure("TCheckbutton", background=self.card_bg, foreground=self.text_color, font=("Segoe UI", 10))
        self.style.map("TCheckbutton", background=[("active", self.card_bg)], foreground=[("active", self.text_color)])
        
        # Treeview styling (for the inverter list)
        self.style.configure("Treeview", background="#ffffff", foreground=self.text_color, fieldbackground="#ffffff", rowheight=28, borderwidth=0, font=("Segoe UI", 10))
        self.style.configure("Treeview.Heading", background="#f8f9fa", foreground=self.text_color, font=("Segoe UI Semibold", 10), borderwidth=1, bordercolor="#dadce0")
        self.style.map("Treeview", background=[("selected", "#e8f0fe")], foreground=[("selected", self.accent_color)])
        self.style.map("Treeview.Heading", background=[("active", "#e8f0fe")], foreground=[("active", self.accent_color)])
        
        # Notebook styling
        self.style.configure("TNotebook", background=self.bg_color, borderwidth=0)
        self.style.configure("TNotebook.Tab", background="#f1f3f4", foreground=self.muted_text, font=("Segoe UI Semibold", 10), padding=[20, 8, 20, 8], borderwidth=0)
        self.style.map("TNotebook.Tab", background=[("selected", "#ffffff"), ("active", "#e8f0fe")], foreground=[("selected", self.accent_color), ("active", self.accent_color)])
        
        # Inverter Nominal capacities
        self.dc_powers = {
            # TX1
            "TX1-INV-1": 343.75, "TX1-INV-2": 343.75, "TX1-INV-3": 343.75, "TX1-INV-4": 343.75,
            "TX1-INV-5": 343.75, "TX1-INV-6": 343.75, "TX1-INV-7": 343.75, "TX1-INV-8": 359.375,
            "TX1-INV-9": 359.375, "TX1-INV-10": 359.375, "TX1-INV-11": 359.375, "TX1-INV-12": 343.75,
            # TX2
            "TX2-INV-1": 328.125, "TX2-INV-2": 343.75, "TX2-INV-3": 343.75, "TX2-INV-4": 328.125,
            "TX2-INV-5": 328.125, "TX2-INV-6": 359.375, "TX2-INV-7": 343.75, "TX2-INV-8": 359.375,
            "TX2-INV-9": 359.375, "TX2-INV-10": 343.75, "TX2-INV-11": 359.375, "TX2-INV-12": 359.375,
            # TX3
            "TX3-INV-1": 359.375, "TX3-INV-2": 359.375, "TX3-INV-3": 359.375, "TX3-INV-4": 359.375,
            "TX3-INV-5": 359.375, "TX3-INV-6": 359.375, "TX3-INV-7": 359.375, "TX3-INV-8": 359.375,
            "TX3-INV-9": 359.375, "TX3-INV-10": 359.375, "TX3-INV-11": 359.375, "TX3-INV-12": 328.125
        }
        self.ac_power_all = 320.0
        
        # State variables
        self.folder_path_var = tk.StringVar(value=self.cfg.get("last_folder", ""))
        
        # Determine current local date and corresponding PVSyst PR default value
        today = datetime.date.today()
        current_date_str = today.strftime("%Y-%m-%d")
        current_month = today.month
        # Monthly PVSyst baseline targets (Year-1 undegraded), user-configurable since v13.
        # Kept as an instance attribute so the reference table, its degradation-adjusted
        # column and the engine all read the same values.
        self.pvsyst_monthly = {int(k): float(v) for k, v in self.cfg["pvsyst_monthly"].items()}
        pvsyst_defaults = self.pvsyst_monthly
        default_pr = f"{pvsyst_defaults.get(current_month, 0.868):.3f}".replace(".", ",")
        
        self.date_var = tk.StringVar(value=current_date_str)
        self.pvsyst_pr_var = tk.StringVar(value=default_pr)
        self.threshold_var = tk.StringVar(value=str(self.cfg["threshold"]))
        self.diff_threshold_var = tk.StringVar(value=str(self.cfg["diff_threshold"]))
        # POA reference method for PR: "condmax" (Conditional MAX) or "average" (two-sensor mean).
        # Default is Media (Average) — the IEC-standard two-pyranometer arithmetic mean.
        self.poa_method_var = tk.StringVar(value=self.cfg["poa_method"])
        self.reprocess_all_var = tk.BooleanVar(value=False)
        
        # v13 day scope: which days of a month folder the batch actually processes.
        # "mese" = all, "giorno" = only the day in the Data field, "intervallo" = Da..A.
        self.day_scope_var = tk.StringVar(value=self.cfg["day_scope"])
        self.day_from_var = tk.StringVar(value=str(self.cfg["day_from"]))
        self.day_to_var = tk.StringVar(value=str(self.cfg["day_to"]))
        # v15: Data source selection (SCADA, VCOM, or Misto)
        self.data_source_var = tk.StringVar(value=self.cfg.get("data_source", "misto"))
        self.vcom_days_var = tk.StringVar(value=str(self.cfg.get("vcom_days", "")))
        # Advanced options (edited through the Opzioni Avanzate dialog)
        self.sync_mother_var = tk.BooleanVar(value=bool(self.cfg["sync_mother"]))
        self.backup_mother_var = tk.BooleanVar(value=bool(self.cfg["backup_mother"]))
        # Safe-stop flag: set by the Interrompi button, polled by the worker at checkpoints
        # between days so the in-flight day finishes writing before we halt.
        self.stop_requested = threading.Event()
        
        # Register a trace on date_var to auto-update the PVSyst PR default value
        self.date_var.trace_add("write", self.on_date_changed)
        # Register a trace on pvsyst_pr_var to update the metric card in real time
        self.pvsyst_pr_var.trace_add("write", self.on_pvsyst_pr_changed)
        
        # Placeholders for results
        self.calc_results = None
        self.df_result = None
        self.all_days_results = []
        
        # Setup GUI elements
        self.create_layout()
        
        # Sync every conditional widget to the restored settings. Without the data-source
        # call the "Giorni VCOM" row stayed visible on startup for SCADA/VCOM, even though
        # it only applies to "Misto".
        self._on_poa_method_change()
        self._on_day_scope_change()
        self._on_data_source_change()
        
        # Select active month row in PVSyst reference treeview on startup
        try:
            self.pvsyst_tree.selection_set(f"m{current_month}")
            self.pvsyst_tree.focus(f"m{current_month}")
            self.pvsyst_tree.see(f"m{current_month}")
        except Exception:
            pass
        
        # Redirect stdout/stderr to Live Log Widget
        self.stdout_redirector = RedirectText(self.root, self.log_widget, sys.stdout)
        self.stderr_redirector = RedirectText(self.root, self.log_widget, sys.stderr)
        sys.stdout = self.stdout_redirector
        sys.stderr = self.stderr_redirector
        
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        
        print(">>> GET SRL - Motore di calcolo Performance Ratio inizializzato (v15).")
        print(">>> Ambito configurabile: mese intero, singolo giorno o intervallo di giorni.")
        if os.path.exists(get_settings_path()):
            print(f">>> Impostazioni caricate da: {get_settings_path()}")
        
    def create_card(self, parent, padding=16):
        card = RoundedCard(parent, bg="#ffffff", border_color="#dadce0", radius=12, padding=padding)
        return card, card.content_frame

    def _wrap_label(self, label, container):
        """Make a left-aligned tk.Label wrap to the live width of `container` instead of
        overflowing/clipping. Keeps long descriptive/hint text fully readable at any window
        size. Guarded against re-layout feedback loops via a cached last width."""
        # anchor="w" as well as justify: justify only aligns the lines of an already-wrapped
        # block, so a single-line label packed with fill="x" (the status line) rendered
        # centred while every other hint sat left.
        label.configure(justify="left", anchor="w")
        label._wrap_cache = -1
        def _update(event):
            new_wrap = max(140, event.width - 6)
            if label._wrap_cache != new_wrap:
                label._wrap_cache = new_wrap
                try:
                    label.config(wraplength=new_wrap)
                except Exception:
                    pass
        container.bind("<Configure>", _update, add="+")

    def create_layout(self):
        # Top Header Bar spanning full width
        header_bar = tk.Frame(self.root, bg="#ffffff", height=60, bd=0, highlightthickness=0)
        header_bar.pack(side="top", fill="x")
        header_bar.pack_propagate(False)
        
        # Subtle bottom border to the header bar
        bottom_border = tk.Frame(self.root, bg="#dadce0", height=1)
        bottom_border.pack(side="top", fill="x")
        
        # Header contents
        header_inner = tk.Frame(header_bar, bg="#ffffff")
        header_inner.pack(fill="both", expand=True, padx=24)
        
        logo_path = get_resource_path(os.path.join("assets", "logo.png"))
        if os.path.exists(logo_path):
            try:
                from PIL import Image, ImageTk
                img = Image.open(logo_path)
                img = img.resize((120, 38), Image.Resampling.LANCZOS)
                self.logo_photo = ImageTk.PhotoImage(img)
                lbl_logo = tk.Label(header_inner, image=self.logo_photo, bg="#ffffff", bd=0)
                lbl_logo.pack(side="left", pady=10)
            except Exception:
                pass
                
        # Vertical divider line in header
        divider = tk.Frame(header_inner, bg="#dadce0", width=1, height=24)
        divider.pack(side="left", padx=16, pady=18)
        
        title_lbl = tk.Label(header_inner, text="Motore di Calcolo Performance Ratio - Mazara 01", bg="#ffffff", fg="#1f2124", font=("Segoe UI Semibold", 13))
        title_lbl.pack(side="left", pady=10)
        
        # Version badge
        version_badge = tk.Label(header_inner, text="v15.0", bg="#e8f0fe", fg="#1a73e8", font=("Segoe UI Semibold", 9), padx=8, pady=2)
        version_badge.pack(side="left", padx=12)
        
        # Guide button on the top right with a ? logo
        btn_guide = ttk.Button(header_inner, text="? Guida d'Uso", style="Secondary.TButton", command=self.show_guide)
        btn_guide.pack(side="right", pady=10)
        
        btn_options = ttk.Button(header_inner, text="Opzioni Avanzate...", style="Secondary.TButton", command=self.show_options)
        btn_options.pack(side="right", padx=(0, 8), pady=10)
        
        # Main container for body (with nice margins/padding)
        main_frame = tk.Frame(self.root, bg="#f8f9fa")
        main_frame.pack(side="top", fill="both", expand=True, padx=24, pady=20)
        
        # Top Grid: Inputs (Left) and Reference table + Log (Right).
        # The controls column grew with every version; stacking the log console underneath
        # it pushed the log off-screen entirely while the reference table wasted ~420px of
        # empty white. Both columns now share the full height: controls left, table + log
        # right, so the log is visible again and the dead space is gone.
        top_grid = tk.Frame(main_frame, bg=self.bg_color)
        top_grid.pack(fill="both", expand=True)
        # Equal split: the left card must fit the full "Ricalcola forzatamente..." checkbox
        # label (~450px), which the old 4:5 weighting squeezed once the PVSyst table gained
        # its 4th column.
        top_grid.columnconfigure(0, weight=1, minsize=460)
        top_grid.columnconfigure(1, weight=1, minsize=440)
        top_grid.rowconfigure(0, weight=1)
        
        # 1. Inputs Frame (Card style using helper)
        inputs_card_border, inputs_card = self.create_card(top_grid, padding=15)
        inputs_card_border.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        
        lbl_sec_in = tk.Label(inputs_card, text="Impostazioni di Calcolo", bg="#ffffff", fg=self.accent_color, font=("Segoe UI Semibold", 12, "bold"))
        lbl_sec_in.pack(anchor="w", pady=(0, 4))
        
        # Add descriptive help text for inputs card
        lbl_desc_in = tk.Label(inputs_card, text="Seleziona la cartella contenente i dati SCADA e configura i parametri.", bg="#ffffff", fg=self.muted_text, font=("Segoe UI", 10))
        lbl_desc_in.pack(anchor="w", pady=(0, 10))
        self._wrap_label(lbl_desc_in, inputs_card)
        
        # Folder row
        folder_frame = tk.Frame(inputs_card, bg="#ffffff")
        folder_frame.pack(fill="x", pady=5)
        lbl_f = tk.Label(folder_frame, text="Cartella File SCADA (Input):", bg="#ffffff", fg=self.text_color, font=("Segoe UI Semibold", 9))
        lbl_f.pack(anchor="w")
        
        folder_entry_frame = tk.Frame(folder_frame, bg="#ffffff")
        folder_entry_frame.pack(fill="x", pady=2)
        
        self.entry_folder = ttk.Entry(folder_entry_frame, textvariable=self.folder_path_var, font=("Segoe UI", 10))
        self.entry_folder.pack(side="left", fill="x", expand=True, padx=(0, 8))
        
        btn_browse = ttk.Button(folder_entry_frame, text="Sfoglia...", style="Secondary.TButton", command=self.browse_folder)
        btn_browse.pack(side="right")
        
        # Parameters Grid (Date, PVSyst PR, Threshold, Diff Threshold)
        params_grid = tk.Frame(inputs_card, bg="#ffffff")
        params_grid.pack(fill="x", pady=10)
        params_grid.columnconfigure(0, weight=1)
        params_grid.columnconfigure(1, weight=1)
        
        # Date Input (Row 0, Column 0)
        date_frame = tk.Frame(params_grid, bg="#ffffff")
        date_frame.grid(row=0, column=0, padx=(0, 10), pady=(0, 10), sticky="ew")
        tk.Label(date_frame, text="Data (AAAA-MM-GG):", bg="#ffffff", fg=self.text_color, font=("Segoe UI Semibold", 9)).pack(anchor="w")
        self.entry_date = ttk.Entry(date_frame, textvariable=self.date_var, font=("Segoe UI", 10))
        self.entry_date.pack(fill="x", pady=2)
        
        # PVSyst PR Input (Row 0, Column 1)
        pr_frame = tk.Frame(params_grid, bg="#ffffff")
        pr_frame.grid(row=0, column=1, padx=(10, 0), pady=(0, 10), sticky="ew")
        tk.Label(pr_frame, text="PR Mensile PVSyst:", bg="#ffffff", fg=self.text_color, font=("Segoe UI Semibold", 9)).pack(anchor="w")
        self.entry_pr = ttk.Entry(pr_frame, textvariable=self.pvsyst_pr_var, font=("Segoe UI", 10))
        self.entry_pr.pack(fill="x", pady=2)
        
        # Irradiance Threshold (Row 1, Column 0)
        thresh_frame = tk.Frame(params_grid, bg="#ffffff")
        thresh_frame.grid(row=1, column=0, padx=(0, 10), pady=(10, 0), sticky="ew")
        tk.Label(thresh_frame, text="Irraggiamento Min (W/m²):", bg="#ffffff", fg=self.text_color, font=("Segoe UI Semibold", 9)).pack(anchor="w")
        self.entry_thresh = ttk.Entry(thresh_frame, textvariable=self.threshold_var, font=("Segoe UI", 10))
        self.entry_thresh.pack(fill="x", pady=2)
        
        # Irradiance Difference Threshold for Conditional MAX (Row 1, Column 1)
        diff_thresh_frame = tk.Frame(params_grid, bg="#ffffff")
        diff_thresh_frame.grid(row=1, column=1, padx=(10, 0), pady=(10, 0), sticky="ew")
        tk.Label(diff_thresh_frame, text="Tolleranza Diff. Irraggiamento (%):", bg="#ffffff", fg=self.text_color, font=("Segoe UI Semibold", 9)).pack(anchor="w")
        self.entry_diff_thresh = ttk.Entry(diff_thresh_frame, textvariable=self.diff_threshold_var, font=("Segoe UI", 10))
        self.entry_diff_thresh.pack(fill="x", pady=2)

        # POA reference method toggle for PR (Row 2, spanning both columns)
        poa_method_frame = tk.Frame(params_grid, bg="#ffffff")
        poa_method_frame.grid(row=2, column=0, columnspan=2, pady=(12, 0), sticky="ew")
        tk.Label(poa_method_frame, text="Riferimento POA per il calcolo del PR:", bg="#ffffff",
                 fg=self.text_color, font=("Segoe UI Semibold", 9)).pack(anchor="w")
        toggle_row = tk.Frame(poa_method_frame, bg="#ffffff")
        toggle_row.pack(anchor="w", pady=(2, 0))
        ttk.Radiobutton(toggle_row, text="Conditional MAX", value="condmax", width=16,
                        variable=self.poa_method_var, style="POA.Toolbutton",
                        command=self._on_poa_method_change).pack(side="left")
        ttk.Radiobutton(toggle_row, text="Media (Average)", value="average", width=16,
                        variable=self.poa_method_var, style="POA.Toolbutton",
                        command=self._on_poa_method_change).pack(side="left", padx=(6, 0))
        self.lbl_poa_method_hint = tk.Label(
            poa_method_frame, bg="#ffffff", fg="#5f6368", font=("Segoe UI", 8),
            text="Media: media aritmetica dei due piranometri (standard IEC). La tolleranza diff. non è usata.")
        self.lbl_poa_method_hint.pack(anchor="w", pady=(2, 0), fill="x")
        self._wrap_label(self.lbl_poa_method_hint, poa_method_frame)

        # v13: day scope selector (Row 3). Restricts which days of a month folder are
        # processed, so a single day or a short range no longer requires re-running
        # (or manually pruning) the whole month.
        scope_frame = tk.Frame(params_grid, bg="#ffffff")
        scope_frame.grid(row=3, column=0, columnspan=2, pady=(12, 0), sticky="ew")
        tk.Label(scope_frame, text="Ambito di calcolo (cartella mensile):", bg="#ffffff",
                 fg=self.text_color, font=("Segoe UI Semibold", 9)).pack(anchor="w")
        scope_row = tk.Frame(scope_frame, bg="#ffffff")
        scope_row.pack(anchor="w", fill="x", pady=(2, 0))
        for text, val, width in (("Mese intero", "mese", 13),
                                 ("Solo 1 giorno", "giorno", 13),
                                 ("Intervallo", "intervallo", 11)):
            ttk.Radiobutton(scope_row, text=text, value=val, width=width,
                            variable=self.day_scope_var, style="POA.Toolbutton",
                            command=self._on_day_scope_change).pack(side="left", padx=(0, 6))
        
        # Da/A on their own row, shown only for "Intervallo". Packed inline next to the three
        # toggles they were clipped off the edge of the card (the toggles alone use the full
        # width), which rendered as a squashed half-spinbox.
        self.day_range_frame = tk.Frame(scope_frame, bg="#ffffff")
        tk.Label(self.day_range_frame, text="Dal giorno", bg="#ffffff", fg=self.muted_text,
                 font=("Segoe UI", 9)).pack(side="left")
        self.spin_day_from = ttk.Spinbox(self.day_range_frame, from_=1, to=31, width=4,
                                         textvariable=self.day_from_var, font=("Segoe UI", 10))
        self.spin_day_from.pack(side="left", padx=(6, 8))
        tk.Label(self.day_range_frame, text="al giorno", bg="#ffffff", fg=self.muted_text,
                 font=("Segoe UI", 9)).pack(side="left")
        self.spin_day_to = ttk.Spinbox(self.day_range_frame, from_=1, to=31, width=4,
                                       textvariable=self.day_to_var, font=("Segoe UI", 10))
        self.spin_day_to.pack(side="left", padx=(6, 0))
        
        self.lbl_scope_hint = tk.Label(
            scope_frame, bg="#ffffff", fg="#5f6368", font=("Segoe UI", 8), text="")
        self.lbl_scope_hint.pack(anchor="w", pady=(2, 0), fill="x")
        self._wrap_label(self.lbl_scope_hint, scope_frame)

        # v15: Data source selector (SCADA vs VCOM vs Misto)
        source_frame = tk.Frame(params_grid, bg="#ffffff")
        source_frame.grid(row=4, column=0, columnspan=2, pady=(12, 0), sticky="ew")
        tk.Label(source_frame, text="Sorgente Dati (SCADA / VCOM):", bg="#ffffff",
                 fg=self.text_color, font=("Segoe UI Semibold", 9)).pack(anchor="w")
        source_row = tk.Frame(source_frame, bg="#ffffff")
        source_row.pack(anchor="w", fill="x", pady=(2, 0))
        for text, val, width in (("SCADA", "scada", 10),
                                 ("VCOM", "vcom", 10),
                                 ("Misto (SCADA + VCOM)", "misto", 22)):
            ttk.Radiobutton(source_row, text=text, value=val, width=width,
                            variable=self.data_source_var, style="POA.Toolbutton",
                            command=self._on_data_source_change).pack(side="left", padx=(0, 6))
        
        self.vcom_days_frame = tk.Frame(source_frame, bg="#ffffff")
        self.vcom_days_frame.pack(anchor="w", fill="x", pady=(4, 0))
        tk.Label(self.vcom_days_frame, text="Giorni VCOM (es. 3, 7, 9):", bg="#ffffff",
                 fg="#5f6368", font=("Segoe UI", 8)).pack(side="left")
        self.entry_vcom_days = ttk.Entry(self.vcom_days_frame, textvariable=self.vcom_days_var,
                                         width=14, font=("Segoe UI", 9))
        self.entry_vcom_days.pack(side="left", padx=(6, 6))
        self.btn_select_vcom_days = ttk.Button(self.vcom_days_frame, text="Seleziona Giorni...",
                                               style="Secondary.TButton", command=self._open_vcom_days_dialog)
        self.btn_select_vcom_days.pack(side="left")

        # Force reprocess checkbox (Batch mode)
        chk_frame = tk.Frame(inputs_card, bg="#ffffff")
        chk_frame.pack(anchor="w", pady=(8, 8), fill="x")
        self.chk_reprocess = ttk.Checkbutton(
            chk_frame, 
            variable=self.reprocess_all_var,
            text="Ricalcola forzatamente i giorni già elaborati (Modalità Batch)"
        )
        self.chk_reprocess.pack(side="left")
        
        # Action button
        self.btn_calculate = ttk.Button(inputs_card, text="Calcola Performance Ratio", style="Action.TButton", command=self.start_calculation)
        self.btn_calculate.pack(fill="x", pady=(5, 4))
        
        self.btn_stop = ttk.Button(inputs_card, text="Interrompi (arresto sicuro)", style="Stop.TButton",
                                   command=self.request_stop, state="disabled")
        self.btn_stop.pack(fill="x", pady=(0, 5))
        
        # Dedicated vendor PR sync buttons frame (v15)
        sync_frame = tk.Frame(inputs_card, bg="#ffffff")
        sync_frame.pack(fill="x", pady=(2, 5))
        sync_frame.columnconfigure(0, weight=1)
        sync_frame.columnconfigure(1, weight=1)

        self.btn_sync_scada = ttk.Button(
            sync_frame,
            text="Sync SCADA PR",
            style="Secondary.TButton",
            command=self.sync_scada_pr
        )
        self.btn_sync_scada.grid(row=0, column=0, padx=(0, 4), sticky="ew")

        self.btn_sync_vcom = ttk.Button(
            sync_frame,
            text="Sync VCOM PR",
            style="Secondary.TButton",
            command=self.sync_vcom_pr
        )
        self.btn_sync_vcom.grid(row=0, column=1, padx=(4, 0), sticky="ew")

        self.lbl_sync_hint = tk.Label(
            sync_frame,
            bg="#ffffff",
            fg="#5f6368",
            font=("Segoe UI", 8),
            text="Trova e aggiorna direttamente i dati PR da SCADA (KPI) o VCOM nel file Madre."
        )
        self.lbl_sync_hint.grid(row=1, column=0, columnspan=2, pady=(3, 0), sticky="w")
        self._wrap_label(self.lbl_sync_hint, sync_frame)
        
        # Progress/Status. The batch can run for minutes; the bar turns "day 7 di 21" into
        # something readable at a glance. Hidden while idle so the card stays uncluttered.
        self.progress = ttk.Progressbar(inputs_card, mode="determinate", maximum=100)
        
        self.lbl_status = tk.Label(inputs_card, text="Pronto. Seleziona la cartella e clicca su Calcola.", bg="#ffffff", fg=self.muted_text, font=("Segoe UI", 11))
        self.lbl_status.pack(anchor="w", pady=(4, 4), fill="x")
        self._wrap_label(self.lbl_status, inputs_card)
        
        # 2. Right column: PVSyst reference table on top, live log console below it.
        right_col = tk.Frame(top_grid, bg=self.bg_color)
        right_col.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        
        metrics_card_border, self.metrics_card = self.create_card(right_col, padding=15)
        metrics_card_border.pack(fill="x")
        
        # Dummy parent for keeping label references without packing them to avoid breaking update logic
        self.dummy_parent = tk.Frame(self.root)
        self.lbl_avg_pr_val = tk.Label(self.dummy_parent)
        self.lbl_comp_pr_val = tk.Label(self.dummy_parent)
        self.lbl_pvsyst_target_val = tk.Label(self.dummy_parent)
        self.lbl_irrad_summary = tk.Label(self.dummy_parent)
        self.btn_export = ttk.Button(self.dummy_parent)
        self.lbl_export_status = tk.Label(self.dummy_parent)
        
        lbl_sec_me = tk.Label(self.metrics_card, text="Tabella Riferimento Target PVSyst", bg="#ffffff", fg=self.accent_color, font=("Segoe UI Semibold", 12, "bold"))
        lbl_sec_me.pack(anchor="w", pady=(0, 2))
        
        # Add descriptive help text for PVSyst reference table
        lbl_desc_me = tk.Label(self.metrics_card, text="Valori teorici mensili di Target PR. La colonna Target Corretto mostra il valore degradato usato per l'anno selezionato. Il mese attivo è evidenziato in automatico.", bg="#ffffff", fg=self.muted_text, font=("Segoe UI", 10))
        lbl_desc_me.pack(anchor="w", pady=(0, 8))
        self._wrap_label(lbl_desc_me, self.metrics_card)
        
        # Table frame with 1px border. Width-only fill: with expand=True the tree collapses
        # to whatever height the card offers and clips the last month (Dicembre); pinned to
        # its natural height it always shows all 12 rows.
        pvsyst_table_border = tk.Frame(self.metrics_card, bg="#dadce0")
        pvsyst_table_border.pack(fill="x")
        
        pvsyst_table_frame = tk.Frame(pvsyst_table_border, bg="#ffffff")
        pvsyst_table_frame.pack(fill="x", padx=1, pady=1)
        
        cols_pvsyst = ("Mese", "Target PR", "Target PR (%)", "Target Corretto")
        # 12 rows: the card no longer stretches, so show every month without scrolling.
        self.pvsyst_tree = ttk.Treeview(pvsyst_table_frame, columns=cols_pvsyst, show="headings", height=12)
        self.pvsyst_tree.pack(fill="x")
        
        col_widths = {"Mese": 95, "Target PR": 90, "Target PR (%)": 95, "Target Corretto": 110}
        for c in cols_pvsyst:
            self.pvsyst_tree.heading(c, text=c)
            align = "w" if c == "Mese" else "center"
            self.pvsyst_tree.column(c, width=col_widths.get(c, 100), anchor=align)
            
        months_names = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
                        "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]
        for idx, name in enumerate(months_names, start=1):
            base = self.pvsyst_monthly[idx]
            dec = f"{base:.3f}".replace(".", ",")
            pct = f"{base * 100:.1f}%".replace(".", ",")
            self.pvsyst_tree.insert("", "end", iid=f"m{idx}", values=(name, dec, pct, "--"))
        # Populate the degradation-adjusted column for the currently selected year
        try:
            _y = int(self.date_var.get().strip().split("-")[0])
        except Exception:
            _y = datetime.date.today().year
        self._refresh_pvsyst_adjusted(_y)
        
        # Dummy parent for keeping the bottom results tabs/notebook and tables without packing them
        dummy_bottom = tk.Frame(self.root)
        
        self.notebook = ttk.Notebook(dummy_bottom)
        self.notebook.pack(fill="both", expand=True)
        
        # Tab 1: Detailed Inverters
        tab1 = ttk.Frame(self.notebook, style="Card.TFrame")
        self.notebook.add(tab1, text="Dettaglio Inverter (Ultimo Giorno)")
        
        lbl_sec_det = tk.Label(tab1, text="Dettaglio Performance Ratio Compensato (36 Inverter)", bg="#ffffff", fg=self.accent_color, font=("Segoe UI Semibold", 12, "bold"))
        lbl_sec_det.pack(anchor="w", pady=(10, 2), padx=5)
        
        lbl_desc_det = tk.Label(tab1, text="Analisi dettagliata per ciascuno dei 36 inverter attivi nell'impianto per l'ultimo giorno elaborato.", bg="#ffffff", fg=self.muted_text, font=("Segoe UI", 10))
        lbl_desc_det.pack(anchor="w", pady=(0, 8), padx=5)
        
        # Table frame with 1px border
        table_border = tk.Frame(tab1, bg="#dadce0")
        table_border.pack(fill="both", expand=True, padx=5, pady=5)
        
        table_frame = tk.Frame(table_border, bg="#ffffff")
        table_frame.pack(fill="both", expand=True, padx=1, pady=1)
        
        scrollbar = ttk.Scrollbar(table_frame)
        scrollbar.pack(side="right", fill="y")
        
        cols = ("Codice Inverter", "Trasformatore", "Potenza CC Nominale (kW)", "Energia Prodotta (kWh)", "Perdita Stimata (kWh)", "PR Compensato (%)")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.tree.yview)
        
        for c in cols:
            self.tree.heading(c, text=c)
            align = "center" if c != "Codice Inverter" else "w"
            width = 110 if c != "Codice Inverter" else 150
            self.tree.column(c, width=width, anchor=align)
 
        # Tab 2: Daily Summary
        tab2 = ttk.Frame(self.notebook, style="Card.TFrame")
        self.notebook.add(tab2, text="Riepilogo Giorni Elaborati")
        
        lbl_sec_days = tk.Label(tab2, text="Risultati Giornalieri (Modalità Batch)", bg="#ffffff", fg=self.accent_color, font=("Segoe UI Semibold", 12, "bold"))
        lbl_sec_days.pack(anchor="w", pady=(10, 2), padx=5)
        
        lbl_desc_days = tk.Label(tab2, text="Cronologia dei risultati giornalieri calcolati e salvati per il mese selezionato.", bg="#ffffff", fg=self.muted_text, font=("Segoe UI", 10))
        lbl_desc_days.pack(anchor="w", pady=(0, 8), padx=5)
        
        # Table frame with 1px border
        days_border = tk.Frame(tab2, bg="#dadce0")
        days_border.pack(fill="both", expand=True, padx=5, pady=5)
        
        days_frame = tk.Frame(days_border, bg="#ffffff")
        days_frame.pack(fill="both", expand=True, padx=1, pady=1)
        
        days_scrollbar = ttk.Scrollbar(days_frame)
        days_scrollbar.pack(side="right", fill="y")
        
        days_cols = ("Data", "Irradiazione (kWh/m²)", "PR Non Comp. (%)", "PR Grezzo Comp. (%)", "Media PR Inv. (%)")
        self.tree_days = ttk.Treeview(days_frame, columns=days_cols, show="headings", yscrollcommand=days_scrollbar.set)
        self.tree_days.pack(side="left", fill="both", expand=True)
        days_scrollbar.config(command=self.tree_days.yview)
        
        for c in days_cols:
            self.tree_days.heading(c, text=c)
            self.tree_days.column(c, width=150, anchor="center")
            
        # 4. Live Log Console Panel — in the right column under the reference table, where
        # it fills the space the table used to waste instead of being squeezed off-screen.
        log_card_border, log_card = self.create_card(right_col, padding=12)
        log_card_border.pack(fill="both", expand=True, pady=(15, 0))
        
        lbl_sec_log = tk.Label(log_card, text="Console Live Log di Esecuzione", bg="#ffffff", fg=self.accent_color, font=("Segoe UI Semibold", 12, "bold"))
        lbl_sec_log.pack(anchor="w", pady=(0, 2))
        
        lbl_desc_log = tk.Label(log_card, text="Messaggi diagnostici in tempo reale sull'elaborazione dei file excel e calcoli.", bg="#ffffff", fg=self.muted_text, font=("Segoe UI", 10))
        lbl_desc_log.pack(anchor="w", pady=(0, 6))
        self._wrap_label(lbl_desc_log, log_card)
        
        log_border = tk.Frame(log_card, bg="#dadce0")
        log_border.pack(fill="both", expand=True, pady=(5, 0))
        
        log_frame = tk.Frame(log_border, bg="#ffffff")
        log_frame.pack(fill="both", expand=True, padx=1, pady=1)
        
        log_scrollbar = ttk.Scrollbar(log_frame)
        log_scrollbar.pack(side="right", fill="y")
        
        self.log_widget = tk.Text(
            log_frame, 
            height=6, 
            bg="#f8f9fa", 
            fg=self.text_color, 
            insertbackground=self.text_color, 
            relief="flat", 
            bd=0,
            font=("Consolas", 9), 
            yscrollcommand=log_scrollbar.set,
            wrap="word",
            state="disabled"
        )
        self.log_widget.pack(side="left", fill="both", expand=True)
        log_scrollbar.config(command=self.log_widget.yview)
        
    def browse_folder(self):
        selected_folder = filedialog.askdirectory(title="Seleziona Cartella Mese (formato AAAA MM)")
        if selected_folder:
            self.folder_path_var.set(selected_folder)
            self.lbl_status.config(text="Cartella selezionata: " + os.path.basename(selected_folder), foreground=self.accent_color)
            # Try to auto-detect date from folder name or file names
            self.auto_detect_date(selected_folder)
            self.auto_detect_vcom_days(selected_folder)
            
    
    def auto_detect_vcom_days(self, folder):
        """Scan folder and subdirectories for days with VCOM data and auto-populate Giorni VCOM field."""
        try:
            if not folder or not os.path.exists(folder):
                return
            vcom_days = []
            subdirs = [d for d in os.listdir(folder) if os.path.isdir(os.path.join(folder, d)) and d.isdigit()]
            for d in sorted(subdirs, key=lambda x: int(x)):
                d_path = os.path.join(folder, d)
                if self._find_vcom_folder_for_day(d_path):
                    vcom_days.append(str(int(d)))
            if vcom_days:
                vcom_str = ", ".join(vcom_days)
                self.vcom_days_var.set(vcom_str)
                print(f"[Auto-Scan VCOM] Rilevati dati VCOM per i giorni: {vcom_str}")
        except Exception as e:
            print(f"[Auto-Scan VCOM] Error scanning folder: {e}")

    def auto_detect_date(self, folder):
        try:
            if not folder or not os.path.exists(folder):
                return
                
            # 1. Try to read from any Excel file inside the folder or its subdirectories
            excel_files = glob.glob(os.path.join(folder, "*.xlsx"))
            if not excel_files:
                # Check subdirectories (like '01' daily folder)
                for subd in os.listdir(folder):
                    subd_path = os.path.join(folder, subd)
                    if os.path.isdir(subd_path) and subd.isdigit():
                        excel_files = glob.glob(os.path.join(subd_path, "*.xlsx"))
                        if excel_files:
                            break
                            
            if excel_files:
                import openpyxl
                import datetime
                import re
                for file_path in excel_files[:3]:  # check up to 3 files to be fast
                    try:
                        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                        sheet = wb.active
                        found = False
                        for row in sheet.iter_rows(max_row=30, max_col=10, values_only=True):
                            for cell in row:
                                if isinstance(cell, (datetime.datetime, datetime.date)):
                                    self.date_var.set(f"{cell.year}-{cell.month:02d}-{cell.day:02d}")
                                    found = True
                                    break
                                if isinstance(cell, str):
                                    m = re.search(r'\b(\d{4})[-/](0[1-9]|1[0-2])[-/](\d{2,4})\b', cell)
                                    if m:
                                        self.date_var.set(f"{m.group(1)}-{m.group(2)}-{int(m.group(3)):02d}")
                                        found = True
                                        break
                                    m2 = re.search(r'\b(\d{1,2})[-/](0[1-9]|1[0-2])[-/](\d{4})\b', cell)
                                    if m2:
                                        self.date_var.set(f"{m2.group(3)}-{m2.group(2)}-{int(m2.group(1)):02d}")
                                        found = True
                                        break
                            if found:
                                return
                    except Exception:
                        continue

            # 2. Fallback to folder name / path parsing
            basename = os.path.basename(folder.rstrip("\\/"))
            parent = os.path.dirname(folder.rstrip("\\/"))
            parent_name = os.path.basename(parent)
            
            # If folder name is YYYY MM, e.g. "2026 05"
            parts = basename.split()
            if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit() and len(parts[0]) == 4:
                year = parts[0]
                month = parts[1]
                self.date_var.set(f"{year}-{month}-01")
                return
                
            if len(basename) == 2 and basename.isdigit():
                # Looks like a day, check parent
                if len(parent_name) == 7 and parent_name[4] == ' ' and parent_name[:4].isdigit() and parent_name[5:].isdigit():
                    year = parent_name[:4]
                    month = parent_name[5:]
                    day = basename
                    self.date_var.set(f"{year}-{month}-{day}")
                    return
                    
            # Or look at active power regulation file name
            files = os.listdir(folder)
            for f in files:
                if "regolazione_della_potenza_attiva_" in f.lower():
                    # Extract date from name like Regolazione_della_potenza_attiva_2026_04_26.xlsx
                    parts = f.replace(".xlsx", "").split("_")
                    # Look for 4 digit year
                    for idx, part in enumerate(parts):
                        if len(part) == 4 and part.isdigit() and idx <= len(parts) - 3:
                            if parts[idx+1].isdigit() and parts[idx+2].isdigit():
                                self.date_var.set(f"{parts[idx]}-{parts[idx+1]}-{parts[idx+2]}")
                                return
        except Exception:
            pass
            
    def show_guide(self):
        html_path = get_resource_path("Manuale_Utente_PR_Calculator.html")
        if os.path.exists(html_path):
            import webbrowser
            webbrowser.open(f"file:///{os.path.abspath(html_path)}")
        else:
            messagebox.showinfo(
                "Guida Utente",
                "Manuale Utente non trovato in locale.\n\n"
                "Procedura d'uso:\n"
                "1. Seleziona la cartella del mese (es. '2026 05').\n"
                "2. La data e il Target PR PVSyst verranno rilevati in automatico.\n"
                "3. Scegli l'Ambito di calcolo: mese intero, solo il giorno indicato nel\n"
                "   campo Data, oppure un intervallo di giorni.\n"
                "4. Clicca su 'Calcola Performance Ratio' per elaborare i file SCADA.\n"
                "5. Esporta i risultati su Excel tramite il pulsante in basso.\n\n"
                "In 'Opzioni Avanzate' puoi configurare la degradazione contrattuale del\n"
                "target, i target PVSyst mensili e il comportamento del file Madre."
            )
            
    def on_date_changed(self, *args):
        try:
            date_str = self.date_var.get().strip()
            parts = date_str.split("-")
            if len(parts) >= 2:
                month_val = int(parts[1])
                pvsyst_defaults = self.pvsyst_monthly
                if month_val in pvsyst_defaults:
                    self.pvsyst_pr_var.set(f"{pvsyst_defaults[month_val]:.3f}".replace(".", ","))
                    # Highlight matching row in the PVSyst reference treeview
                    try:
                        self.pvsyst_tree.selection_set(f"m{month_val}")
                        self.pvsyst_tree.focus(f"m{month_val}")
                        self.pvsyst_tree.see(f"m{month_val}")
                    except Exception:
                        pass
                # Refresh the degradation-adjusted target column for the selected year
                try:
                    year_val = int(parts[0]) if parts[0].isdigit() and len(parts[0]) == 4 else None
                    self._refresh_pvsyst_adjusted(year_val)
                except Exception:
                    pass
        except Exception:
            pass
            
    def on_pvsyst_pr_changed(self, *args):
        try:
            val = float(self.pvsyst_pr_var.get().replace(",", "."))
            self.lbl_pvsyst_target_val.config(text=f"{val * 100:.3f} %".replace(".", ","))
        except ValueError:
            self.lbl_pvsyst_target_val.config(text="-- %")

    def _refresh_pvsyst_adjusted(self, year=None):
        """Fill the 'Target Corretto' column with the degradation-adjusted monthly target
        actually used by the engine for `year` (the contractual 0,4%/anno decay from the
        Feb-2025 plant start). Falls back to the current calendar year when `year` is None."""
        if year is None:
            year = datetime.date.today().year
        for idx in range(1, 13):
            base = self.pvsyst_monthly[idx]
            try:
                _n, factor = self._pr_degradation_factor(year, idx)
                adjusted = base * factor
                text = f"{adjusted:.3f}".replace(".", ",")
            except Exception:
                text = "--"
            try:
                self.pvsyst_tree.set(f"m{idx}", "Target Corretto", text)
            except Exception:
                pass
            
    def start_calculation(self):
        folder = self.folder_path_var.get()
        if not folder or not os.path.exists(folder):
            messagebox.showerror("Errore", "Per favore seleziona una cartella di input valida!")
            return
            
        try:
            pvsyst_pr = float(self.pvsyst_pr_var.get().replace(",", "."))
            threshold = float(self.threshold_var.get().replace(",", "."))
            diff_thresh_val = float(self.diff_threshold_var.get().replace(",", "."))
            if not (0.0 <= diff_thresh_val <= 100.0):
                raise ValueError("Out of range")
            diff_threshold = diff_thresh_val / 100.0
            poa_method = self.poa_method_var.get()
        except ValueError:
            messagebox.showerror("Errore", "I parametri (PR, Soglia Irraggiamento, Tolleranza Diff) devono essere numeri validi, e la Tolleranza Diff. deve essere compresa tra 0% e 100%!")
            return
            
        date_str = self.date_var.get().strip()
        if len(date_str) != 10 or date_str[4] != '-' or date_str[7] != '-':
            messagebox.showerror("Errore", "La data deve essere nel formato AAAA-MM-GG!")
            return
            
        try:
            day_filter = self._resolve_day_filter()
        except ValueError as scope_err:
            messagebox.showerror("Ambito di calcolo non valido", str(scope_err))
            return
        
        save_settings(self._collect_settings())
        
        self.stop_requested.clear()
        self.btn_calculate.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.btn_sync_scada.config(state="disabled")
        self.btn_sync_vcom.config(state="disabled")
        self.btn_export.config(state="disabled")
        self.lbl_status.config(text="Calcolo del PR in corso... attendere prego...", foreground=self.warn_color)
        
        # Run calculation in a separate thread to keep UI active
        thread = threading.Thread(target=self.run_calculation, args=(folder, date_str, pvsyst_pr, threshold, diff_threshold, poa_method, day_filter))
        thread.start()
        
    def request_stop(self):
        """Safe stop: never kills work mid-write. Sets a flag the worker polls at its
        checkpoints (between days, between VCOM downloads/conversions), so the operation
        already in flight completes and we halt at the next boundary."""
        self.stop_requested.set()
        self.btn_stop.config(state="disabled")
        self.lbl_status.config(
            text="Arresto richiesto: completamento dell'operazione in corso...",
            foreground=self.warn_color)
        print(">>> Arresto richiesto dall'utente. Completamento dell'operazione corrente, poi stop.")

    def _on_day_scope_change(self):
        """Show the Da/A row only for the 'Intervallo' scope; update the hint."""
        scope = self.day_scope_var.get()
        try:
            if scope == "intervallo":
                # before= keeps the row between the toggles and the hint text, however many
                # times it is hidden and shown again.
                self.day_range_frame.pack(anchor="w", pady=(6, 0), before=self.lbl_scope_hint)
            else:
                self.day_range_frame.pack_forget()
            self.lbl_scope_hint.config(text={
                "mese": "Elabora tutti i giorni presenti nella cartella mensile.",
                "giorno": "Elabora solo il giorno indicato nel campo Data (AAAA-MM-GG). "
                          "Molto più rapido per un ricalcolo puntuale.",
                "intervallo": "Elabora solo i giorni compresi fra Da e A (estremi inclusi).",
            }[scope])
        except Exception:
            pass

    def _resolve_day_filter(self):
        """Days the batch is allowed to process, as a set of ints, or None for the whole
        month. Raises ValueError with a user-facing message on invalid input."""
        scope = self.day_scope_var.get()
        if scope == "mese":
            return None
        if scope == "giorno":
            try:
                return {datetime.datetime.strptime(self.date_var.get().strip(), "%Y-%m-%d").day}
            except ValueError:
                raise ValueError("Ambito 'Solo 1 giorno': la data deve essere nel formato AAAA-MM-GG.")
        try:
            d_from = int(self.day_from_var.get())
            d_to = int(self.day_to_var.get())
        except ValueError:
            raise ValueError("Ambito 'Intervallo': i giorni Da e A devono essere numeri interi.")
        if not (1 <= d_from <= 31 and 1 <= d_to <= 31):
            raise ValueError("Ambito 'Intervallo': i giorni devono essere compresi fra 1 e 31.")
        if d_from > d_to:
            raise ValueError(f"Ambito 'Intervallo': il giorno iniziale ({d_from}) è successivo a quello finale ({d_to}).")
        return set(range(d_from, d_to + 1))

    def _collect_settings(self):
        """Snapshot the current UI configuration into the persisted settings dict."""
        self.cfg.update({
            "pvsyst_monthly": {str(k): v for k, v in self.pvsyst_monthly.items()},
            "sync_mother": bool(self.sync_mother_var.get()),
            "backup_mother": bool(self.backup_mother_var.get()),
            "threshold": self.threshold_var.get(),
            "diff_threshold": self.diff_threshold_var.get(),
            "poa_method": self.poa_method_var.get(),
            "day_scope": self.day_scope_var.get(),
            "day_from": self.day_from_var.get(),
            "day_to": self.day_to_var.get(),
            "data_source": self.data_source_var.get(),
            "vcom_days": self.vcom_days_var.get(),
            "last_folder": self.folder_path_var.get(),
        })
        return self.cfg

    def show_options(self):
        """Advanced options dialog: contractual degradation, Madre file behaviour and the
        12 monthly PVSyst targets — all previously hard-coded in the engine."""
        win = tk.Toplevel(self.root)
        win.title("Opzioni Avanzate")
        win.configure(bg=self.bg_color)
        win.transient(self.root)
        win.resizable(False, False)
        
        body = tk.Frame(win, bg=self.bg_color)
        body.pack(fill="both", expand=True, padx=18, pady=16)
        
        # --- Degradation ---
        tk.Label(body, text="Degradazione contrattuale del Target PR (Allegato 9.1)",
                 bg=self.bg_color, fg=self.accent_color,
                 font=("Segoe UI Semibold", 11, "bold")).grid(row=0, column=0, columnspan=4, sticky="w")
        tk.Label(body, text="Il target PVSyst decade di questa percentuale all'anno, con capitalizzazione annua, a partire dal mese/anno di avvio impianto.",
                 bg=self.bg_color, fg=self.muted_text, font=("Segoe UI", 9),
                 wraplength=520, justify="left").grid(row=1, column=0, columnspan=4, sticky="w", pady=(2, 8))
        
        rate_var = tk.StringVar(value=f"{float(self.cfg['deg_rate']) * 100:.3f}".rstrip("0").rstrip(".").replace(".", ","))
        sy_var = tk.StringVar(value=str(self.cfg["deg_start_year"]))
        sm_var = tk.StringVar(value=str(self.cfg["deg_start_month"]))
        for col, (text, var, width) in enumerate((("Decadimento (%/anno):", rate_var, 8),
                                                 ("Anno di avvio:", sy_var, 8),
                                                 ("Mese di avvio:", sm_var, 8))):
            cell = tk.Frame(body, bg=self.bg_color)
            cell.grid(row=2, column=col, sticky="w", padx=(0, 14))
            tk.Label(cell, text=text, bg=self.bg_color, fg=self.text_color,
                     font=("Segoe UI Semibold", 9)).pack(anchor="w")
            ttk.Entry(cell, textvariable=var, width=width, font=("Segoe UI", 10)).pack(anchor="w", pady=2)
        
        # --- Mother file behaviour ---
        tk.Label(body, text="File Madre", bg=self.bg_color, fg=self.accent_color,
                 font=("Segoe UI Semibold", 11, "bold")).grid(row=3, column=0, columnspan=4, sticky="w", pady=(16, 2))
        self.style.configure("Opt.TCheckbutton", background=self.bg_color)
        ttk.Checkbutton(body, variable=self.sync_mother_var, style="Opt.TCheckbutton",
                        text="Sincronizza il file Madre al termine del calcolo (disattiva per ricalcoli rapidi di pochi giorni)"
                        ).grid(row=4, column=0, columnspan=4, sticky="w")
        ttk.Checkbutton(body, variable=self.backup_mother_var, style="Opt.TCheckbutton",
                        text="Crea un backup del file Madre prima di modificarlo (ultimi 5 conservati)"
                        ).grid(row=5, column=0, columnspan=4, sticky="w")
        
        # --- PVSyst monthly targets ---
        tk.Label(body, text="Target PR PVSyst mensili (anno 1, non degradati)",
                 bg=self.bg_color, fg=self.accent_color,
                 font=("Segoe UI Semibold", 11, "bold")).grid(row=6, column=0, columnspan=4, sticky="w", pady=(16, 6))
        months_short = ["Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"]
        pr_vars = {}
        grid_pr = tk.Frame(body, bg=self.bg_color)
        grid_pr.grid(row=7, column=0, columnspan=4, sticky="w")
        for idx, name in enumerate(months_short, start=1):
            v = tk.StringVar(value=f"{self.pvsyst_monthly[idx]:.3f}".replace(".", ","))
            pr_vars[idx] = v
            cell = tk.Frame(grid_pr, bg=self.bg_color)
            cell.grid(row=(idx - 1) // 6, column=(idx - 1) % 6, padx=(0, 10), pady=3, sticky="w")
            tk.Label(cell, text=name, bg=self.bg_color, fg=self.muted_text,
                     font=("Segoe UI Semibold", 9)).pack(anchor="w")
            ttk.Entry(cell, textvariable=v, width=7, font=("Segoe UI", 10)).pack(anchor="w")
        
        status = tk.Label(body, text="", bg=self.bg_color, fg=self.muted_text, font=("Segoe UI", 9))
        status.grid(row=8, column=0, columnspan=4, sticky="w", pady=(12, 0))
        
        def apply_and_save():
            try:
                rate = float(rate_var.get().replace(",", ".")) / 100.0
                if not 0.0 <= rate <= 0.5:
                    raise ValueError("Il decadimento deve essere compreso fra 0% e 50% all'anno.")
                start_year = int(sy_var.get())
                start_month = int(sm_var.get())
                if not 1 <= start_month <= 12:
                    raise ValueError("Il mese di avvio deve essere compreso fra 1 e 12.")
                new_targets = {}
                for m, var in pr_vars.items():
                    val = float(var.get().replace(",", "."))
                    if not 0.0 < val <= 1.0:
                        raise ValueError(f"Il target di {months_short[m - 1]} deve essere compreso fra 0 e 1.")
                    new_targets[m] = val
            except ValueError as err:
                messagebox.showerror("Valore non valido", str(err), parent=win)
                return
            
            self.cfg.update({"deg_rate": rate, "deg_start_year": start_year, "deg_start_month": start_month})
            self.pvsyst_monthly = new_targets
            # Refresh both PVSyst columns and the PR field for the currently selected month
            for m, base in self.pvsyst_monthly.items():
                try:
                    self.pvsyst_tree.set(f"m{m}", "Target PR", f"{base:.3f}".replace(".", ","))
                    self.pvsyst_tree.set(f"m{m}", "Target PR (%)", f"{base * 100:.1f}%".replace(".", ","))
                except Exception:
                    pass
            self.on_date_changed()
            ok = save_settings(self._collect_settings())
            status.config(text="Impostazioni salvate." if ok else "Applicate, ma il salvataggio su disco è fallito.",
                          fg=self.success_color if ok else self.warn_color)
            print(f">>> Opzioni aggiornate: decadimento {rate * 100:.3f}%/anno da {start_month:02d}/{start_year}.")
            win.after(700, win.destroy)
        
        def restore_defaults():
            rate_var.set("0,4")
            sy_var.set(str(DEFAULT_SETTINGS["deg_start_year"]))
            sm_var.set(str(DEFAULT_SETTINGS["deg_start_month"]))
            self.sync_mother_var.set(DEFAULT_SETTINGS["sync_mother"])
            self.backup_mother_var.set(DEFAULT_SETTINGS["backup_mother"])
            for m, var in pr_vars.items():
                var.set(f"{DEFAULT_SETTINGS['pvsyst_monthly'][str(m)]:.3f}".replace(".", ","))
            status.config(text="Valori predefiniti ripristinati (non ancora salvati).", fg=self.muted_text)
        
        btns = tk.Frame(body, bg=self.bg_color)
        btns.grid(row=9, column=0, columnspan=4, sticky="e", pady=(14, 0))
        ttk.Button(btns, text="Ripristina predefiniti", style="Secondary.TButton",
                   command=restore_defaults).pack(side="left", padx=(0, 8))
        ttk.Button(btns, text="Annulla", style="Secondary.TButton", command=win.destroy).pack(side="left", padx=(0, 8))
        ttk.Button(btns, text="Salva", style="TButton", command=apply_and_save).pack(side="left")
        
        win.update_idletasks()
        win.geometry(f"+{self.root.winfo_rootx() + 80}+{self.root.winfo_rooty() + 60}")
        win.grab_set()

    def _on_data_source_change(self):
        """Show/hide or enable the VCOM days configuration depending on selected source."""
        src = self.data_source_var.get()
        if src == "misto":
            self.vcom_days_frame.pack(anchor="w", fill="x", pady=(4, 0))
        else:
            self.vcom_days_frame.pack_forget()

    def _parse_vcom_days_set(self):
        """Parse comma/space/range separated day string into a set of integer day numbers."""
        raw = self.vcom_days_var.get().strip()
        days = set()
        if not raw:
            return days
        raw = raw.replace(';', ',').replace(' ', ',')
        for part in raw.split(','):
            part = part.strip()
            if not part:
                continue
            if '-' in part:
                try:
                    s_str, e_str = part.split('-')
                    for d in range(int(s_str), int(e_str) + 1):
                        if 1 <= d <= 31:
                            days.add(d)
                except Exception:
                    pass
            else:
                try:
                    d = int(part)
                    if 1 <= d <= 31:
                        days.add(d)
                except Exception:
                    pass
        return days

    def _format_days_set(self, days_set):
        """Format a set of day numbers into a clean comma-separated string."""
        return ", ".join(str(d) for d in sorted(days_set))

    def _folder_contains_vcom_csvs(self, folder):
        """Check if a folder directly contains VCOM CSV files."""
        if not os.path.exists(folder) or not os.path.isdir(folder):
            return False
        try:
            files = [f.lower() for f in os.listdir(folder) if f.endswith(".csv")]
            has_ac = any("potenza_ac" in f or "potenza ac" in f for f in files)
            has_prod = any("produzione_energetica" in f or "produzione energetica" in f or "potenza_attiva" in f or "potenza attiva" in f for f in files)
            return has_ac and has_prod
        except Exception:
            return False

    def _open_vcom_days_dialog(self):
        """Open a modal dialog to interactively select which days of the month use VCOM data."""
        dlg = tk.Toplevel(self.root)
        dlg.title("Seleziona Giorni con Dati VCOM")
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.configure(bg="#ffffff")
        dlg.geometry("460x480")
        dlg.resizable(False, False)

        tk.Label(dlg, text="Seleziona i giorni da elaborare con dati VCOM (anziché SCADA):",
                 bg="#ffffff", fg=self.text_color, font=("Segoe UI Semibold", 10), wraplength=420, justify="left").pack(padx=16, pady=(16, 8), anchor="w")

        # Frame with 31 checkboxes
        grid_frame = tk.Frame(dlg, bg="#ffffff")
        grid_frame.pack(padx=16, pady=8, fill="both", expand=True)

        current_vcom = self._parse_vcom_days_set()
        
        # Scan month folder to check which days currently have a vcom/ folder
        folder = self.folder_path_var.get().strip()
        days_with_vcom_folder = set()
        if os.path.exists(folder) and os.path.isdir(folder):
            for d in range(1, 32):
                d_path = os.path.join(folder, f"{d:02d}")
                if os.path.exists(d_path) and self._find_vcom_folder_for_day(d_path):
                    days_with_vcom_folder.add(d)

        day_vars = {}
        cols = 5
        for d in range(1, 32):
            r = (d - 1) // cols
            c = (d - 1) % cols
            # Check if initially checked
            is_init = (d in current_vcom) or (d in days_with_vcom_folder)
            var = tk.BooleanVar(value=is_init)
            day_vars[d] = var
            lbl_txt = f"{d:02d}"
            if d in days_with_vcom_folder:
                lbl_txt += " (VCOM)"
            chk = ttk.Checkbutton(grid_frame, text=lbl_txt, variable=var)
            chk.grid(row=r, column=c, padx=6, pady=4, sticky="w")

        # Actions frame
        act_frame = tk.Frame(dlg, bg="#ffffff")
        act_frame.pack(padx=16, pady=8, fill="x")

        def select_all():
            for v in day_vars.values():
                v.set(True)

        def deselect_all():
            for v in day_vars.values():
                v.set(False)

        def auto_detect():
            for d, v in day_vars.items():
                v.set(d in days_with_vcom_folder)

        ttk.Button(act_frame, text="Tutti", style="Secondary.TButton", command=select_all).pack(side="left", padx=(0, 4))
        ttk.Button(act_frame, text="Nessuno", style="Secondary.TButton", command=deselect_all).pack(side="left", padx=(0, 4))
        ttk.Button(act_frame, text="Rileva cartelle vcom/", style="Secondary.TButton", command=auto_detect).pack(side="left", padx=(0, 4))

        # Bottom buttons
        btn_box = tk.Frame(dlg, bg="#ffffff")
        btn_box.pack(padx=16, pady=(8, 16), fill="x")

        def on_confirm():
            selected = {d for d, v in day_vars.items() if v.get()}
            self.vcom_days_var.set(self._format_days_set(selected))
            dlg.destroy()

        ttk.Button(btn_box, text="Conferma", style="Action.TButton", command=on_confirm).pack(side="right", padx=(4, 0))
        ttk.Button(btn_box, text="Annulla", style="Secondary.TButton", command=dlg.destroy).pack(side="right")

    def _on_poa_method_change(self):
        """Enable the deviation-tolerance field only for Conditional MAX; update the hint."""
        try:
            if self.poa_method_var.get() == "average":
                self.entry_diff_thresh.config(state="disabled")
                self.lbl_poa_method_hint.config(
                    text="Media: media aritmetica dei due piranometri (standard IEC). La tolleranza diff. non è usata.")
            else:
                self.entry_diff_thresh.config(state="normal")
                self.lbl_poa_method_hint.config(
                    text="Conditional MAX: usa il sensore maggiore se i due piranometri divergono oltre la tolleranza (più conservativo).")
        except Exception:
            pass

    def find_file_by_patterns(self, folder, patterns):
        for pattern in patterns:
            # Simple direct match
            matches = glob.glob(os.path.join(folder, pattern))
            if matches:
                return matches[0]
            # Case insensitive match
            for file in os.listdir(folder):
                file_path = os.path.join(folder, file)
                if os.path.isfile(file_path):
                    # Check if pattern without asterisks is part of the filename
                    core_pat = pattern.replace("*", "").lower()
                    if core_pat in file.lower():
                        return file_path
        return None

    def _find_vcom_folder_for_day(self, day_path):
        """Check if day_path or its subfolders 'vcom'/'VCOM' contain VCOM CSV files."""
        if self._folder_contains_vcom_csvs(day_path):
            return day_path
        for name in ["vcom", "VCOM", "vcom_data", "VCOM_DATA"]:
            v_dir = os.path.join(day_path, name)
            if self._folder_contains_vcom_csvs(v_dir):
                return v_dir
        return None

    def _save_as_vcom_csv(self, headers, rows, date_str, output_path):
        import codecs
        from datetime import datetime
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        period_str = f"Periodo: {dt.day}/{dt.month}/{dt.year} 0.00.00 - {dt.day}/{dt.month}/{dt.year} 23.59.59"
        
        if headers and headers[0].lower() in ["ora", "time", "date"]:
            headers[0] = "Data"
            
        with codecs.open(output_path, "w", encoding="utf-16") as f:
            f.write(f'"{period_str}"\r\n\r\n')
            
            hdr_line = "\t".join(f'"{h}"' for h in headers)
            f.write(hdr_line + "\r\n")
            
            for r in rows:
                row_line = "\t".join(f'"{cell}"' for cell in r)
                f.write(row_line + "\r\n")

    def _download_vcom_data(self, date_str, output_vcom_folder):
        """Single-day wrapper — delegates to the batch method so the browser is
        only opened once even when called for a single day."""
        results = self._download_vcom_data_batch([(date_str, output_vcom_folder)])
        return results.get(date_str, False)

    def _download_vcom_data_batch(self, date_folder_pairs):
        """Download VCOM data for *all* requested days in a single browser session.

        Parameters
        ----------
        date_folder_pairs : list of (date_str, output_vcom_folder)
            Each element is a (YYYY-MM-DD, path) pair.

        Returns
        -------
        dict  {date_str: bool}  – True if both files were obtained for that date.
        """
        import json
        import time
        from playwright.sync_api import sync_playwright

        # Frozen (PyInstaller) builds resolve the Playwright browser path to the temporary
        # _MEIxxxx extraction dir, which contains no browsers. Point it at the real per-user
        # install so the .exe finds the same browsers as a plain .py run.
        if not os.environ.get("PLAYWRIGHT_BROWSERS_PATH"):
            _browsers_dir = os.path.join(
                os.environ.get("LOCALAPPDATA", os.path.expanduser(r"~\AppData\Local")),
                "ms-playwright")
            if os.path.isdir(_browsers_dir):
                os.environ["PLAYWRIGHT_BROWSERS_PATH"] = _browsers_dir
            else:
                print(f"[VCOM-Downloader] ATTENZIONE: browser Playwright non trovati in {_browsers_dir}. "
                      "Esegui 'playwright install chromium'.")

        config_path = r"\\s01\get\2025.01 Mazara 01 A2A\03 - REPORT\Report\09 Testing\VCOM Automation\config.json"
        if not os.path.exists(config_path):
            print(f"[VCOM-Downloader] Error: Config not found at {config_path}")
            return {d: False for d, _ in date_folder_pairs}

        try:
            with open(config_path, "r", encoding="utf-8") as f:
                cfg = json.load(f)
        except Exception as e:
            print(f"[VCOM-Downloader] Error reading config: {e}")
            return {d: False for d, _ in date_folder_pairs}

        # Pre-check which dates actually need downloading
        todo = []   # (date_str, output_vcom_folder, prod_path, ac_path)
        results = {}
        for date_str, output_vcom_folder in date_folder_pairs:
            os.makedirs(output_vcom_folder, exist_ok=True)
            prod_path = os.path.join(output_vcom_folder, f"Produzione_energetica_{date_str.replace('-', '_')}.csv")
            ac_path   = os.path.join(output_vcom_folder, f"Potenza_AC_{date_str.replace('-', '_')}.csv")
            already_prod = os.path.exists(prod_path) and os.path.getsize(prod_path) > 1000
            already_ac   = os.path.exists(ac_path)   and os.path.getsize(ac_path)   > 1000
            if already_prod and already_ac:
                results[date_str] = True
            else:
                todo.append((date_str, output_vcom_folder, prod_path, ac_path, already_prod, already_ac))

        if not todo:
            return results   # everything was already cached

        with sync_playwright() as p:
            # Headed: the extraction is slow and DOM-sensitive, so keep it visible to watch.
            browser = p.chromium.launch(headless=False)
            context = browser.new_context(viewport={"width": 1450, "height": 900})
            page = context.new_page()

            try:
                # ---- LOGIN (once) ----
                print(f"[VCOM-Downloader] Logging in to VCOM (session per {len(todo)} giorno/i)...")
                page.goto(cfg["SYSTEM_URL"], timeout=60000)
                page.wait_for_load_state("networkidle")

                try:
                    page.locator('button:has-text("Usa solo i cookie necessari"), button:has-text("Accetta tutti i cookie")').click(timeout=5000)
                except Exception:
                    pass

                if page.locator('input[type="password"]:visible').count() > 0:
                    page.locator('input[type="text"]:visible').first.fill(cfg["USERNAME"])
                    page.locator('input[type="password"]:visible').first.fill(cfg["PASSWORD"])
                    page.locator('button:has-text("Login"), button[type="submit"]').first.click()
                    time.sleep(5)

                for i in range(3):
                    if page.locator('input#username:visible').count() > 0:
                        page.locator('input#username').fill(cfg["USERNAME"])
                        page.press('input#username', "Enter")
                        time.sleep(3)
                    if page.locator('input#password:visible').count() > 0:
                        page.locator('input#password').fill(cfg["PASSWORD"])
                        page.press('input#password', "Enter")
                        break
                    time.sleep(2)

                for i in range(15):
                    if "auth.meteocontrol.com" not in page.url and "realms" not in page.url and "#state" not in page.url:
                        break
                    time.sleep(2)
                time.sleep(3)

                def click_dati_tab():
                    try:
                        page.evaluate("""() => {
                            document.querySelectorAll('.modal-backdrop').forEach(el => el.remove());
                            document.querySelectorAll('.modal.show, .modal.fade.show').forEach(el => {
                                el.classList.remove('show');
                                el.style.display = 'none';
                            });
                        }""")
                    except Exception:
                        pass
                    page.evaluate("window.scrollTo(0, 450)")
                    time.sleep(0.5)
                    tab = page.get_by_text("Dati", exact=True).last
                    tab.wait_for(state="visible", timeout=20000)
                    parent_cls = tab.evaluate("el => el.parentElement ? el.parentElement.className : ''")
                    if "active" not in parent_cls and "selected" not in parent_cls and "ui-tabs-active" not in parent_cls:
                        tab.click()
                        time.sleep(2)

                def extract_table():
                    return page.evaluate("""() => {
                        const table = document.querySelector('#infotab-data table');
                        if (!table) return { headers: [], rows: [] };
                        const thEls = Array.from(table.querySelectorAll('thead tr th'));
                        const headers = thEls.map(th => th.innerText.trim());
                        const trEls = Array.from(table.querySelectorAll('tbody tr'));
                        const rows = trEls.map(tr =>
                            Array.from(tr.querySelectorAll('td')).map(td => td.innerText.trim())
                        );
                        return { headers, rows };
                    }""")

                # ---- LOOP through dates (same session) ----
                for date_str, output_vcom_folder, prod_path, ac_path, has_prod, has_ac in todo:
                    if self.stop_requested.is_set():
                        print("[VCOM-Downloader] Arresto richiesto: download interrotto.")
                        results[date_str] = False
                        continue

                    print(f"[VCOM-Downloader] Download in corso per {date_str}...")
                    day_ok = True

                    # 1. Produzione Energetica
                    if not has_prod:
                        try:
                            prod_url = (f"https://vcom.meteocontrol.com/vcom/evaluation/index/index/"
                                        f"systemId/2144635?key=LXLXE&type=ad&date={date_str}"
                                        f"T00%3A00%3A00%2B02%3A00&endDate={date_str}T23%3A59%3A59%2B02%3A00")
                            page.goto(prod_url, timeout=60000)
                            page.wait_for_load_state("networkidle")
                            time.sleep(3)
                            click_dati_tab()
                            page.locator("#infotab-data table tbody tr").first.wait_for(state="visible", timeout=25000)
                            result = extract_table()
                            headers, rows = result.get("headers", []), result.get("rows", [])
                            if headers and rows:
                                self._save_as_vcom_csv(headers, rows, date_str, prod_path)
                                print(f"[VCOM-Downloader] Saved Produzione Energetica to {prod_path}")
                            else:
                                day_ok = False
                        except Exception as e:
                            print(f"[VCOM-Downloader] Produzione Energetica error for {date_str}: {e}")
                            day_ok = False

                    # 2. Potenza AC
                    if not has_ac:
                        try:
                            ac_url = (f"https://vcom.meteocontrol.com/vcom/evaluation/index/index/"
                                      f"systemId/2144635?key=5EJH8&type=wr&date={date_str}"
                                      f"T00%3A00%3A00%2B02%3A00&endDate={date_str}T23%3A59%3A59%2B02%3A00")
                            page.goto(ac_url, timeout=60000)
                            page.wait_for_load_state("networkidle")
                            time.sleep(3)
                            click_dati_tab()
                            page.locator("#infotab-data table tbody tr").first.wait_for(state="visible", timeout=25000)
                            result = extract_table()
                            headers, rows = result.get("headers", []), result.get("rows", [])
                            if headers and rows:
                                self._save_as_vcom_csv(headers, rows, date_str, ac_path)
                                print(f"[VCOM-Downloader] Saved Potenza AC to {ac_path}")
                            else:
                                day_ok = False
                        except Exception as e:
                            print(f"[VCOM-Downloader] Potenza AC error for {date_str}: {e}")
                            day_ok = False

                    results[date_str] = day_ok

            except Exception as e:
                print(f"[VCOM-Downloader] Playwright exception: {e}")
            finally:
                browser.close()

        # Fill any dates not yet in results (browser crashed before reaching them)
        for date_str, _ in date_folder_pairs:
            results.setdefault(date_str, False)

        return results

    # Required SCADA inputs for one day, as (human label, filename patterns).
    REQUIRED_DAY_FILES = [
        ("Regolazione Potenza Attiva", ["Regolazione_della_potenza_attiva_*.xlsx", "*potenza_attiva*.xlsx"]),
        ("Contatore SATAC",            ["SATAC_Meter_15Min.xlsx", "SATAC_Meter*.xlsx", "*SATAC*.xlsx"]),
        ("Meteo TS1",                  ["TS_01_Weather_15Min.xlsx", "*Weather*01*.xlsx", "*TS_01*Weather*.xlsx"]),
        ("Meteo TS3",                  ["TS_03_Weather_15Min.xlsx", "*Weather*03*.xlsx", "*TS_03*Weather*.xlsx"]),
        ("Inverter TS1",               ["TS_01_Inverter_15Min.xlsx", "*Inverter*01*.xlsx", "*TS_01*Inverter*.xlsx"]),
        ("Inverter TS2",               ["TS_02_Inverter_15Min.xlsx", "*Inverter*02*.xlsx", "*TS_02*Inverter*.xlsx"]),
        ("Inverter TS3",               ["TS_03_Inverter_15Min.xlsx", "*Inverter*03*.xlsx", "*TS_03*Inverter*.xlsx"]),
    ]

    def _missing_files_for_day(self, folder):
        """Return the labels of the required SCADA files that are NOT present in `folder`.
        An empty list means the day has all 7 files and can be processed."""
        if not os.path.isdir(folder):
            return [label for label, _ in self.REQUIRED_DAY_FILES]
        missing = []
        for label, patterns in self.REQUIRED_DAY_FILES:
            if not self.find_file_by_patterns(folder, patterns):
                missing.append(label)
        return missing

    def _ask_yes_no_on_gui(self, title, message):
        """Show a blocking yes/no dialog on the Tk main thread and return the answer.
        Safe to call from the worker thread: it marshals to the GUI thread and waits."""
        import threading as _t
        import queue
        if _t.current_thread() is _t.main_thread():
            try:
                return messagebox.askyesno(title, message, icon="warning")
            except Exception:
                return False
        q = queue.Queue()
        def _ask():
            try:
                q.put(messagebox.askyesno(title, message, icon="warning"))
            except Exception:
                q.put(False)
        self.root.after(0, _ask)
        return q.get()

    def _ask_select_days_on_gui(self, title, message, day_strings):
        """Show a modal dialog with a checkbox for each day (all checked by default).
        Returns the list of day_strings the user left checked, or [] if cancelled.
        Safe to call from the worker thread."""
        import threading as _t
        import queue
        q = queue.Queue()

        def _show():
            try:
                dlg = tk.Toplevel(self.root)
                dlg.title(title)
                dlg.configure(bg="#ffffff")
                dlg.resizable(False, False)
                dlg.grab_set()
                dlg.transient(self.root)

                # Window size and centering
                dlg.update_idletasks()
                pw, ph = self.root.winfo_width(), self.root.winfo_height()
                px, py = self.root.winfo_rootx(), self.root.winfo_rooty()
                dw = 460
                dh = min(620, max(380, 120 + 32 * len(day_strings)))
                dlg.geometry(f"{dw}x{dh}+{px + (pw - dw) // 2}+{py + (ph - dh) // 2}")
                dlg.resizable(True, True)

                # 1. Message (TOP)
                tk.Label(dlg, text=message, bg="#ffffff", fg="#202124",
                         font=("Segoe UI", 10, "bold"), wraplength=420, justify="left"
                         ).pack(side="top", anchor="w", padx=16, pady=(14, 8))

                # 2. Button Frame (BOTTOM - packed first so it's guaranteed visible at the bottom)
                btn_frame = tk.Frame(dlg, bg="#ffffff")
                btn_frame.pack(side="bottom", fill="x", padx=16, pady=(10, 14))

                # 3. Middle Scrollable Frame (CENTER - fills remaining space)
                middle_frame = tk.Frame(dlg, bg="#ffffff")
                middle_frame.pack(side="top", fill="both", expand=True, padx=16, pady=(0, 4))

                canvas = tk.Canvas(middle_frame, bg="#ffffff", highlightthickness=0)
                scrollbar = ttk.Scrollbar(middle_frame, orient="vertical", command=canvas.yview)
                inner = tk.Frame(canvas, bg="#ffffff")
                inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
                canvas.create_window((0, 0), window=inner, anchor="nw")
                canvas.configure(yscrollcommand=scrollbar.set)
                
                scrollbar.pack(side="right", fill="y")
                canvas.pack(side="left", fill="both", expand=True)

                # Mouse-wheel scrolling
                def _on_mousewheel(event):
                    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
                canvas.bind_all("<MouseWheel>", _on_mousewheel)

                vars_map = {}
                for ds in sorted(day_strings, key=lambda x: int(x)):
                    var = tk.BooleanVar(value=True)
                    vars_map[ds] = var
                    cb = tk.Checkbutton(inner, text=f"Giorno {ds}", variable=var,
                                        bg="#ffffff", fg="#202124",
                                        font=("Segoe UI", 10), anchor="w",
                                        activebackground="#f1f3f4")
                    cb.pack(fill="x", padx=4, pady=2)

                # Select all / Deselect all + OK / Cancel buttons
                def _select_all():
                    for v in vars_map.values():
                        v.set(True)
                def _deselect_all():
                    for v in vars_map.values():
                        v.set(False)

                tk.Button(btn_frame, text="Seleziona tutti", command=_select_all,
                          font=("Segoe UI", 9), relief="flat", bg="#f1f3f4", cursor="hand2"
                          ).pack(side="left", padx=(0, 4))
                tk.Button(btn_frame, text="Deseleziona tutti", command=_deselect_all,
                          font=("Segoe UI", 9), relief="flat", bg="#f1f3f4", cursor="hand2"
                          ).pack(side="left", padx=(0, 12))

                def _ok():
                    canvas.unbind_all("<MouseWheel>")
                    selected = [d for d, v in vars_map.items() if v.get()]
                    q.put(selected)
                    dlg.destroy()
                def _cancel():
                    canvas.unbind_all("<MouseWheel>")
                    q.put([])
                    dlg.destroy()

                tk.Button(btn_frame, text="Annulla", command=_cancel,
                          font=("Segoe UI", 10), width=10, relief="flat", bg="#e8eaed", cursor="hand2"
                          ).pack(side="right", padx=(4, 0))
                tk.Button(btn_frame, text="Scarica", command=_ok,
                          font=("Segoe UI Semibold", 10, "bold"), width=10,
                          fg="#ffffff", bg=self.accent_color, activebackground=self.accent_hover,
                          relief="flat", cursor="hand2"
                          ).pack(side="right", padx=(4, 0))

                dlg.protocol("WM_DELETE_WINDOW", _cancel)
                dlg.wait_window()
            except Exception as e:
                import traceback
                print(f"[Dialog-Error] Exception in _ask_select_days_on_gui: {e}")
                traceback.print_exc()
                q.put([])

        if _t.current_thread() is _t.main_thread():
            _show()
        else:
            self.root.after(0, _show)
        return q.get()

    def _force_close_workbook(self, file_path, exclude_hwnd=None):
        """Force-close `file_path` in ANY running Excel instance (e.g. a window the user
        left open) by binding to it through the Running Object Table and calling Close
        without saving. If `exclude_hwnd` is given, the Excel Application with that window
        handle is skipped -- used so we never close our OWN automation copy while evicting
        external ones. Returns True if a matching workbook was closed."""
        import pythoncom
        import win32com.client
        target_base = os.path.basename(file_path).strip().lower()
        try:
            target_full = os.path.normcase(os.path.abspath(file_path))
        except Exception:
            target_full = None
        closed_any = False
        try:
            context = pythoncom.CreateBindCtx(0)
            rot = pythoncom.GetRunningObjectTable()
            for moniker in rot.EnumRunning():
                try:
                    disp = moniker.GetDisplayName(context, None)
                except Exception:
                    continue
                if not disp:
                    continue
                # Match a full path precisely; fall back to bare-name matches only.
                if ('\\' in disp) or ('/' in disp):
                    try:
                        same = (target_full is not None and
                                os.path.normcase(os.path.abspath(disp)) == target_full)
                    except Exception:
                        same = False
                else:
                    same = (disp.strip().lower() == target_base)
                if not same:
                    continue
                try:
                    obj = rot.GetObject(moniker)
                    # GetObject returns a raw IUnknown; QueryInterface to IDispatch first.
                    wb = win32com.client.Dispatch(obj.QueryInterface(pythoncom.IID_IDispatch))
                    if exclude_hwnd is not None:
                        try:
                            if int(wb.Application.Hwnd) == int(exclude_hwnd):
                                continue  # never close our own automation instance
                        except Exception:
                            pass
                    # NB: do NOT probe wb.FullName first -- a stuck instance (e.g. blocked
                    # on a modal) errors on any property read, which would wrongly skip the
                    # Close. Just attempt the Close; the path match above already targeted it.
                    wb.Close(SaveChanges=False)
                    closed_any = True
                    print(f"DEBUG: Workbook '{os.path.basename(disp)}' chiuso in un'altra istanza di Excel.")
                except Exception as ce:
                    print(f"DEBUG: impossibile chiudere '{disp}' via ROT: {ce}")
        except Exception as e:
            print(f"DEBUG: enumerazione ROT fallita: {e}")
        return closed_any

    def _pid_from_excel_app(self, excel_app):
        """Return the OS process id backing an Excel Application COM object (or None)."""
        try:
            import win32process
            hwnd = int(excel_app.Hwnd)
            _thread_id, pid = win32process.GetWindowThreadProcessId(hwnd)
            return int(pid)
        except Exception:
            return None

    def _kill_processes_locking_file(self, file_path, exclude_pids=()):
        """Last resort when a graceful ROT close fails: terminate the EXCEL.EXE process(es)
        that still hold an OS lock on `file_path` (typically a stuck/orphaned automation
        instance frozen on a modal). Scoped tightly -- only Excel processes with THIS exact
        file, or its Office '~$' owner file, open are killed; our own PIDs are excluded so
        we never kill the running automation instance or this Python process."""
        try:
            import psutil
        except Exception as e:
            print(f"DEBUG: psutil non disponibile, impossibile terminare i processi bloccanti: {e}")
            return False
        try:
            target_full = os.path.normcase(os.path.abspath(file_path))
        except Exception:
            target_full = None
        target_base = os.path.basename(file_path).lower()
        owner_base = ("~$" + os.path.basename(file_path)).lower()
        exclude = {os.getpid()} | {int(p) for p in exclude_pids if p}
        killed = []
        for proc in psutil.process_iter(['pid', 'name']):
            try:
                if (proc.info.get('name') or '').lower() != 'excel.exe':
                    continue
                if proc.info['pid'] in exclude:
                    continue
                try:
                    ofiles = proc.open_files()
                except Exception:
                    ofiles = []
                hit = False
                for f in ofiles:
                    fp = f.path
                    try:
                        if target_full is not None and os.path.normcase(os.path.abspath(fp)) == target_full:
                            hit = True
                            break
                    except Exception:
                        pass
                    if os.path.basename(fp).lower() in (target_base, owner_base):
                        hit = True
                        break
                if hit:
                    proc.kill()
                    killed.append(proc.info['pid'])
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                continue
            except Exception:
                continue
        if killed:
            print(f"DEBUG: Terminato/i processo/i Excel bloccante/i '{target_base}': PID {killed}")
        return len(killed) > 0

    def _save_workbook_resilient(self, wb, excel_app, abs_path, date_str, attempts=3):
        """Save `wb` without ever hanging on a file that is also open elsewhere.

        A plain `wb.Save()` to a network share blocks indefinitely when another Excel
        window (or an orphaned EXCEL.EXE) holds the file, because the sharing-violation
        prompt is not covered by DisplayAlerts. This helper hardens the app against modal
        dialogs, force-closes any EXTERNAL copy of the file (never our own instance)
        before each attempt, retries, and finally falls back to an atomic temp-file
        replace. Raises RuntimeError only if every strategy fails."""
        import time as _time
        try:
            own_hwnd = int(excel_app.Hwnd)
        except Exception:
            own_hwnd = None
        own_pid = self._pid_from_excel_app(excel_app)

        # Suppress any interactive/modal prompt so a locked save raises instead of hanging.
        saved_flags = {}
        for prop, val in (("DisplayAlerts", False), ("Interactive", False),
                          ("AskToUpdateLinks", False), ("EnableEvents", False)):
            try:
                saved_flags[prop] = getattr(excel_app, prop)
            except Exception:
                saved_flags[prop] = None
            try:
                setattr(excel_app, prop, val)
            except Exception:
                pass

        last_err = None
        try:
            for i in range(attempts):
                # Evict any external window holding the file before trying to save.
                try:
                    if self._force_close_workbook(abs_path, exclude_hwnd=own_hwnd):
                        print(f"[{date_str}] Copia esterna aperta di "
                              f"'{os.path.basename(abs_path)}' chiusa forzatamente prima del salvataggio.")
                except Exception:
                    pass
                try:
                    wb.Save()
                    return True
                except Exception as e:
                    last_err = e
                    print(f"[{date_str}] DEBUG: Salvataggio bloccato (tentativo {i + 1}/{attempts}): {e}. "
                          "Nuovo tentativo dopo chiusura forzata...")
                    # If a stuck instance won't release via ROT, kill the process holding
                    # the file (never our own automation instance) before the next attempt.
                    self._kill_processes_locking_file(abs_path, exclude_pids=(own_pid,))
                    _time.sleep(1.0)

            # Last resort: save to a temp file next to the target, then atomically replace.
            base, ext = os.path.splitext(abs_path)
            tmp_path = f"{base}.__saving__{ext}"
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception:
                pass
            wb.SaveAs(tmp_path)
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
            # Make sure nothing holds the real target, then swap the temp file in.
            self._force_close_workbook(abs_path, exclude_hwnd=own_hwnd)
            self._kill_processes_locking_file(abs_path, exclude_pids=(own_pid,))
            os.replace(tmp_path, abs_path)
            print(f"[{date_str}] File '{os.path.basename(abs_path)}' salvato tramite "
                  "copia temporanea e sostituzione atomica (il file era aperto altrove).")
            return True
        except Exception as e2:
            raise RuntimeError(
                f"Impossibile salvare '{os.path.basename(abs_path)}' anche dopo la chiusura "
                f"forzata delle copie aperte: {last_err or e2}"
            )
        finally:
            # Restore Interactive so later opens on the shared app are not blocked.
            try:
                excel_app.Interactive = True if saved_flags.get("Interactive") is None else saved_flags["Interactive"]
            except Exception:
                pass

    def _open_workbook_writable(self, excel_app, abs_path, max_prompts=2):
        """Open a workbook for writing. If it is locked (opens read-only) because another
        Excel window -- or a stuck/orphaned automation instance -- holds it, ask the user
        for permission, then escalate: (1) gracefully close the workbook in the other
        instance via the ROT, and if that fails (2) kill the EXCEL.EXE process still
        holding the file. Retries after each step. Raises RuntimeError if the user declines
        or it stays locked."""
        own_hwnd = None
        try:
            own_hwnd = int(excel_app.Hwnd)
        except Exception:
            pass
        own_pid = self._pid_from_excel_app(excel_app)

        def _reopen():
            return excel_app.Workbooks.Open(abs_path, UpdateLinks=0)

        wb = _reopen()
        attempts = 0
        while getattr(wb, "ReadOnly", False) and attempts < max_prompts:
            attempts += 1
            try:
                wb.Close(SaveChanges=False)   # release our own read-only handle first
            except Exception:
                pass
            proceed = self._ask_yes_no_on_gui(
                "File aperto in Excel",
                f"Il file:\n\n{os.path.basename(abs_path)}\n\n"
                "è aperto in un'altra finestra di Excel e impedisce il salvataggio.\n\n"
                "Vuoi chiuderlo ora e continuare?\n\n"
                "ATTENZIONE: eventuali modifiche non salvate in quel file andranno perse."
            )
            if not proceed:
                raise RuntimeError(
                    f"Elaborazione annullata dall'utente: il file '{os.path.basename(abs_path)}' "
                    "è aperto in Excel e non è stato chiuso."
                )
            # (1) graceful: close the workbook in any OTHER Excel instance.
            closed = self._force_close_workbook(abs_path, exclude_hwnd=own_hwnd)
            # (2) aggressive: a stuck instance won't respond to Close -- kill the process
            #     that still holds the file (never our own automation instance / this PID).
            if not closed:
                self._kill_processes_locking_file(abs_path, exclude_pids=(own_pid,))
            wb = _reopen()

        if getattr(wb, "ReadOnly", False):
            # Final escalation before giving up: force-kill the locking process and retry once.
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
            if self._kill_processes_locking_file(abs_path, exclude_pids=(own_pid,)):
                wb = _reopen()

        if getattr(wb, "ReadOnly", False):
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
            raise RuntimeError(
                f"Impossibile aprire '{os.path.basename(abs_path)}' in scrittura: "
                "risulta ancora bloccato da un'altra applicazione."
            )
        return wb

    @staticmethod
    def _repair_meter_series(raw, left_anchor=None):
        """Repair the cumulative SATAC meter reading series for one day.

        `raw` is a list of 96 cumulative readings, with None for any interval whose
        row was missing from the file. A reading is treated as BAD when it is missing
        (None), zero, or it decreases versus the last good value (the production meter
        is monotonic, so a drop means a glitch / negative energy delta). Each run of
        consecutive bad readings is replaced by a straight-line interpolation between
        the nearest good readings on each side (which equals the average of the two
        neighbours for a single isolated gap). `left_anchor` is yesterday's last
        reading, used when the day starts with a bad value.

        Returns (repaired_values, bad_flags)."""
        n = len(raw)
        vals = list(raw)
        bad = [False] * n
        last_good = left_anchor
        for i in range(n):
            v = vals[i]
            is_bad = (v is None) or (v == 0)
            if (not is_bad) and (last_good is not None) and (v < last_good):
                is_bad = True  # meter went backwards -> negative delta
            if is_bad:
                bad[i] = True
            else:
                last_good = v
        i = 0
        while i < n:
            if not bad[i]:
                i += 1
                continue
            a = i
            while i < n and bad[i]:
                i += 1
            b = i - 1
            left_val = vals[a - 1] if a - 1 >= 0 else left_anchor
            right_val = vals[b + 1] if b + 1 < n else None
            if left_val is None and right_val is None:
                for k in range(a, b + 1):
                    vals[k] = 0.0
            elif left_val is None:          # bad run at start of day
                for k in range(a, b + 1):
                    vals[k] = right_val
            elif right_val is None:         # bad run at end of day
                for k in range(a, b + 1):
                    vals[k] = left_val
            else:                           # interpolate across the gap
                span = (b + 1) - (a - 1)
                for k in range(a, b + 1):
                    vals[k] = left_val + (right_val - left_val) * (k - (a - 1)) / span
        return vals, bad

    def calculate_single_day(self, folder, date_str, pvsyst_pr, threshold, diff_threshold=0.10, calcolo_folder=None, skip_mother_update=False, poa_method="condmax"):
        import shutil
        import openpyxl
        import datetime
        
        # Format date components
        date_replaced = date_str.replace("-", "_") # e.g. 2026_04_26

        # Apply the contractual PR-target degradation to the PVSyst target for THIS month.
        # Per Allegato 9.1 the guaranteed PR degrades 0.4%/year (compounding annually) from
        # the Feb-2025 plant start. The year/month are taken from date_str (which the batch
        # derives from the /YYYY MM/DD folder layout). Year 1 (Feb 2025 - Jan 2026) is
        # undegraded (n=0); the incoming pvsyst_pr is treated as the Year-1 monthly baseline.
        try:
            _dt = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            deg_n, deg_factor = self._pr_degradation_factor(_dt.year, _dt.month)
            _pvsyst_base = float(pvsyst_pr)
            pvsyst_pr = _pvsyst_base * deg_factor
            if deg_n > 0:
                _rate_txt = f"{float(self.cfg['deg_rate']) * 100:.3f}".rstrip("0").rstrip(".").replace(".", ",")
                print(f"[{date_str}] Target PVSyst degradato: {_pvsyst_base*100:.3f}% "
                      f"x (1-{_rate_txt}%)^{deg_n} = {pvsyst_pr*100:.3f}% (anno contrattuale {deg_n+1})")
        except Exception as _deg_err:
            print(f"[{date_str}] DEBUG: degradazione target PVSyst non applicata: {_deg_err}")

        # Find required files
        reg_patterns = [f"Regolazione_della_potenza_attiva_{date_replaced}.xlsx", f"Regolazione_potenza_attiva_{date_replaced}.xlsx", "*potenza_attiva*.xlsx"]
        satac_patterns = ["SATAC_Meter_15Min.xlsx", "SATAC_Meter*.xlsx", "*SATAC*.xlsx"]
        ts1_w_patterns = ["TS_01_Weather_15Min.xlsx", "*Weather*01*.xlsx", "*TS_01*Weather*.xlsx"]
        ts3_w_patterns = ["TS_03_Weather_15Min.xlsx", "*Weather*03*.xlsx", "*TS_03*Weather*.xlsx"]
        ts1_i_patterns = ["TS_01_Inverter_15Min.xlsx", "*Inverter*01*.xlsx", "*TS_01*Inverter*.xlsx"]
        ts2_i_patterns = ["TS_02_Inverter_15Min.xlsx", "*Inverter*02*.xlsx", "*TS_02*Inverter*.xlsx"]
        ts3_i_patterns = ["TS_03_Inverter_15Min.xlsx", "*Inverter*03*.xlsx", "*TS_03*Inverter*.xlsx"]
        
        reg_file = self.find_file_by_patterns(folder, reg_patterns)
        satac_file = self.find_file_by_patterns(folder, satac_patterns)
        ts1_w_file = self.find_file_by_patterns(folder, ts1_w_patterns)
        ts3_w_file = self.find_file_by_patterns(folder, ts3_w_patterns)
        ts1_i_file = self.find_file_by_patterns(folder, ts1_i_patterns)
        ts2_i_file = self.find_file_by_patterns(folder, ts2_i_patterns)
        ts3_i_file = self.find_file_by_patterns(folder, ts3_i_patterns)
        
        missing_files = []
        if not reg_file: missing_files.append("Active Power Regulation")
        if not satac_file: missing_files.append("SATAC Meter Reading")
        if not ts1_w_file: missing_files.append("TS1 Weather Station")
        if not ts3_w_file: missing_files.append("TS3 Weather Station")
        if not ts1_i_file: missing_files.append("TS1 Inverters")
        if not ts2_i_file: missing_files.append("TS2 Inverters")
        if not ts3_i_file: missing_files.append("TS3 Inverters")
        
        if missing_files:
            err_msg = f"I seguenti file richiesti sono assenti nella cartella '{os.path.basename(folder)}':\n"
            for mf in missing_files:
                err_msg += f"- {mf}\n"
            raise FileNotFoundError(err_msg)
            
        # 1. Load active power regulation
        df_reg = None
        try:
            df_reg = pd.read_excel(reg_file)
        except Exception:
            for enc in ['utf-16', 'utf-16-le', 'utf-16-be', 'utf-8-sig', 'utf-8', 'latin-1']:
                try:
                    df_reg = pd.read_csv(reg_file, sep='\t', encoding=enc)
                    if len(df_reg.columns) > 1:
                        break
                except Exception:
                    continue
                    
        if df_reg is None or len(df_reg.columns) <= 1:
            raise ValueError(f"Could not read regulation file with any known encoding or format: {reg_file}")
            
        val_col = None
        for col in df_reg.columns:
            if 'potenza attiva' in col.lower() or 'valore nominale' in col.lower():
                val_col = col
                break
        if val_col is None:
            raise ValueError(f"Could not find 'potenza attiva' column in regulation file. Columns: {list(df_reg.columns)}")
        df_reg['limit_ratio'] = df_reg[val_col].astype(str).str.replace(',', '.').astype(float) / 100.0
        
        def normalize_columns(df):
            if df is None or len(df) == 0:
                return pd.DataFrame(columns=[f"Colonna{i+1}" for i in range(10)])
            if 'Colonna2' not in df.columns:
                row_data = pd.DataFrame([df.columns.values], columns=[f"Colonna{i+1}" for i in range(len(df.columns))])
                df.columns = [f"Colonna{i+1}" for i in range(len(df.columns))]
                df = pd.concat([row_data, df], ignore_index=True)
            return df

        # 2. Load weather data (TX1 and TX3)
        df_w1 = normalize_columns(pd.read_excel(ts1_w_file))
        df_w3 = normalize_columns(pd.read_excel(ts3_w_file))
        
        df_poa1 = df_w1[df_w1['Colonna2'].astype(str).str.strip() == "POA"].copy() if (len(df_w1) > 0 and 'Colonna2' in df_w1.columns) else pd.DataFrame()
        df_poa3 = df_w3[df_w3['Colonna2'].astype(str).str.strip() == "POA"].copy() if (len(df_w3) > 0 and 'Colonna2' in df_w3.columns) else pd.DataFrame()
        
        # 3. Load meter reading data
        df_m = normalize_columns(pd.read_excel(satac_file))
        df_meter = df_m[df_m['Colonna2'].astype(str).str.strip().str.startswith("Energia attiva prod")].copy()
        
        # Load previous day's meter reading if available
        df_meter_prev = None
        try:
            dt_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            prev_dt_obj = dt_obj - datetime.timedelta(days=1)
            
            # Find project root and construct previous day's folder
            month_folder = os.path.dirname(folder)
            project_root = os.path.dirname(month_folder)
            
            prev_month_folder = os.path.join(project_root, f"{prev_dt_obj.year} {prev_dt_obj.month:02d}")
            prev_day_folder = os.path.join(prev_month_folder, f"{prev_dt_obj.day:02d}")
            
            if os.path.exists(prev_day_folder):
                prev_satac_file = self.find_file_by_patterns(prev_day_folder, satac_patterns)
                if prev_satac_file:
                    df_m_prev = normalize_columns(pd.read_excel(prev_satac_file))
                    df_meter_prev = df_m_prev[df_m_prev['Colonna2'].astype(str).str.strip().str.startswith("Energia attiva prod")].copy()
                    print(f"[{date_str}] Letture contatore SATAC del giorno precedente caricate da '{os.path.basename(prev_satac_file)}'.")
        except Exception as ex:
            print(f"[{date_str}] Avvertenza: Impossibile caricare le letture SATAC del giorno precedente: {ex}")
        
        # 4. Load inverter data (TX1, TX2, TX3)
        df_i1 = normalize_columns(pd.read_excel(ts1_i_file))
        df_i2 = normalize_columns(pd.read_excel(ts2_i_file))
        df_i3 = normalize_columns(pd.read_excel(ts3_i_file))
        
        df_pa1 = df_i1[df_i1['Colonna2'].astype(str).str.strip() == "Potenza attiva"].copy()
        df_pa2 = df_i2[df_i2['Colonna2'].astype(str).str.strip() == "Potenza attiva"].copy()
        df_pa3 = df_i3[df_i3['Colonna2'].astype(str).str.strip() == "Potenza attiva"].copy()
        
        times = pd.date_range("00:00:00", "23:45:00", freq="15min").time
        time_strs = [t.strftime("%H:%M:%S") for t in times]
        
        def clean_float(val):
            if val is None or pd.isna(val): return 0.0
            if isinstance(val, (int, float)): return float(val)
            s = str(val).strip().replace(',', '.').rstrip('.')
            try: return float(s)
            except Exception: return 0.0

        # --- Pre-pass: build and repair the day's cumulative meter reading series ---
        # Missing rows -> None; the repair fills missing / zero / decreasing (negative
        # delta) readings by interpolating between the nearest good neighbours, and
        # flags those intervals so the daily file can highlight them.
        raw_meter = []
        for t_str in time_strs:
            mr = df_meter[df_meter['Colonna6'].astype(str).str[:8] == t_str]
            raw_meter.append(clean_float(mr['Colonna3'].values[0]) if len(mr) > 0 else None)
        meter_left_anchor = None
        if df_meter_prev is not None:
            last_row = df_meter_prev[df_meter_prev['Colonna6'].astype(str).str[:8] == "23:45:00"]
            if len(last_row) > 0:
                meter_left_anchor = clean_float(last_row['Colonna3'].values[0])
            elif len(df_meter_prev) > 0:
                meter_left_anchor = clean_float(df_meter_prev['Colonna3'].values[-1])
        meter_series, meter_bad = self._repair_meter_series(raw_meter, meter_left_anchor)
        n_meter_repaired = sum(1 for b in meter_bad if b)
        if n_meter_repaired:
            repaired_times = [time_strs[i] for i in range(len(time_strs)) if meter_bad[i]]
            print(f"[{date_str}] ATTENZIONE: {n_meter_repaired} letture contatore mancanti/anomale "
                  f"interpolate: {', '.join(repaired_times)}")

        calc_rows = []
        for idx, t_str in enumerate(time_strs):
            p1_row = df_poa1[df_poa1['Colonna6'].astype(str).str[:8] == t_str] if (len(df_poa1) > 0 and 'Colonna6' in df_poa1.columns) else pd.DataFrame()
            p3_row = df_poa3[df_poa3['Colonna6'].astype(str).str[:8] == t_str] if (len(df_poa3) > 0 and 'Colonna6' in df_poa3.columns) else pd.DataFrame()
            
            poa1 = clean_float(p1_row['Colonna3'].values[0]) if len(p1_row) > 0 else 0.0
            poa3 = clean_float(p3_row['Colonna3'].values[0]) if len(p3_row) > 0 else 0.0
            
            if poa1 <= 0 and poa3 > 0:
                poa1 = poa3
            elif poa3 <= 0 and poa1 > 0:
                poa3 = poa1
            
            poa1_kwh = poa1 / 4000.0
            poa3_kwh = poa3 / 4000.0
            
            poa_avg_kwh = (poa1_kwh + poa3_kwh) / 2.0
            poa_avg_w = (poa1 + poa3) / 2.0
            # Column H reference irradiance (instantaneous average POA, W/m²). As of the
            # v10 methodology update, PR is referenced to the Conditional MAX POA
            # (Column I / poa_cond_max_kwh, see below); `h` now feeds ONLY the energy-loss
            # estimates and is kept as the plain average to reproduce v10's losses exactly.
            h = poa_avg_w if poa_avg_w > threshold else 0.0
            
            diff_pct = 0.0
            if poa1_kwh > 0 and poa3_kwh > 0:
                avg_val = (poa1_kwh + poa3_kwh) / 2.0
                if avg_val != 0:
                    diff_pct = abs(poa1_kwh - poa3_kwh) / avg_val
            
            if poa1_kwh == 0 and poa3_kwh == 0:
                poa_cond_max_kwh = 0.0
            elif poa1_kwh == 0 or poa3_kwh == 0:
                # One pyranometer dead: use the working sensor regardless of method
                # (averaging in a 0 would halve the POA). This is a data-quality fallback.
                poa_cond_max_kwh = max(poa1_kwh, poa3_kwh)
            elif poa_method == "average":
                # Two-sensor arithmetic mean (standard IEC); deviation tolerance not applied.
                poa_cond_max_kwh = poa_avg_kwh
            elif diff_pct > diff_threshold:
                # Conditional MAX: sensors disagree beyond tolerance -> trust the higher one.
                poa_cond_max_kwh = max(poa1_kwh, poa3_kwh)
            else:
                poa_cond_max_kwh = poa_avg_kwh
            # PR reference requires POA >= threshold (e.g. 50 W/m²). poa_cond_max_kwh is
            # in kWh/m²; convert back to W/m² (x4000) and zero the interval if below.
            # Mirrors the Excel Column I threshold gate (selected POA*4000 >= $BA$7).
            if poa_cond_max_kwh * 4000.0 < threshold:
                poa_cond_max_kwh = 0.0

            reg_row = df_reg[df_reg['Unnamed: 1'].astype(str).str[:8] == t_str]
            limit_ratio = reg_row['limit_ratio'].values[0] if len(reg_row) > 0 else 0.876
            
            # Meter readings come from the repaired cumulative series built above.
            m_val = meter_series[idx]
            if idx == 0:
                m_val_prev = meter_left_anchor if meter_left_anchor is not None else m_val
            else:
                m_val_prev = meter_series[idx - 1]
            meter_missing = bool(meter_bad[idx])
            
            pa1_t = df_pa1[df_pa1['Colonna6'].astype(str).str[:8] == t_str]
            pa2_t = df_pa2[df_pa2['Colonna6'].astype(str).str[:8] == t_str]
            pa3_t = df_pa3[df_pa3['Colonna6'].astype(str).str[:8] == t_str]
            
            inv_powers = {}
            for i in range(1, 13):
                inv_name = f"MW(EA,MW(17,Data_Mod_TS_01_Inverter_{i:02d}.I01))"
                inv_row = pa1_t[pa1_t['Colonna1'] == inv_name]
                inv_powers[f"TX1-INV-{i}"] = clean_float(inv_row['Colonna3'].values[0]) if len(inv_row) > 0 else 0.0
                
            for i in range(1, 13):
                inv_name = f"MW(EG,MW(18,Data_Mod_TS_02_Inverter_{i:02d}.I01))"
                inv_row = pa2_t[pa2_t['Colonna1'] == inv_name]
                inv_powers[f"TX2-INV-{i}"] = clean_float(inv_row['Colonna3'].values[0]) if len(inv_row) > 0 else 0.0
                
            for i in range(1, 13):
                inv_name = f"MW(EM,MW(19,Data_Mod_TS_03_Inverter_{i:02d}.I01))"
                inv_row = pa3_t[pa3_t['Colonna1'] == inv_name]
                inv_powers[f"TX3-INV-{i}"] = clean_float(inv_row['Colonna3'].values[0]) if len(inv_row) > 0 else 0.0
                
            row_data = {
                "time": t_str,
                "poa1": poa1,
                "poa3": poa3,
                "poa1_kwh": poa1_kwh,
                "poa3_kwh": poa3_kwh,
                "poa_avg_kwh": poa_avg_kwh,
                "poa_avg_w": poa_avg_w,
                "h": h,
                "diff_pct": diff_pct,
                "poa_cond_max_kwh": poa_cond_max_kwh,
                "limit_ratio": limit_ratio,
                "meter_reading": m_val,
                "meter_prev_reading": m_val_prev,
                "meter_missing": meter_missing,
                **inv_powers
            }
            calc_rows.append(row_data)
            
        df_result = pd.DataFrame(calc_rows)
        df_result['active_energy_prod'] = (df_result['meter_reading'] - df_result['meter_prev_reading']) * 1000.0
        
        for tx in ["TX1", "TX2", "TX3"]:
            cols = [f"{tx}-INV-{i}" for i in range(1, 13)]
            df_result[f"{tx}_Average_Power"] = df_result[cols].apply(lambda r: r[r > 1.0].mean() if len(r[r > 1.0]) > 0 else 0.0, axis=1)
            
        # Variant B: outage-recovery (ramp) loss. Flag intervals where the WHOLE plant is
        # offline (all 36 inverters < 1.0 kW). The intervals immediately before/after such a
        # trip are the ramp-down/ramp-up shoulders, where an inverter can read >1 kW yet still
        # be far below its POA-expected output -- energy the zero-output downtime test misses.
        # Scoped to trip-adjacent intervals so normal daily under-production
        # (soiling/temperature/low-sun ramp) is NOT counted.
        all_inv_cols = [f"{tx}-INV-{i}" for tx in ["TX1", "TX2", "TX3"] for i in range(1, 13)]
        plant_off = (df_result[all_inv_cols] < 1.0).all(axis=1)
        adj_outage = (plant_off | plant_off.shift(1, fill_value=False) | plant_off.shift(-1, fill_value=False)).values

        new_cols = {}
        for inv_id in self.dc_powers:
            dc = self.dc_powers[inv_id]
            tx_name = inv_id.split("-")[0]

            # Same ramp-loss rule for every inverter: booked only on trip-adjacent, sun-up,
            # producing, non-curtailed intervals; equals the POA-expected shortfall.
            # Mutually exclusive with dt_loss (needs inv<1 kW) and curt_loss (needs limit<0.875).
            ramp_loss_s = np.where(
                adj_outage
                & (df_result['h'] > threshold).values
                & (df_result[inv_id] >= 1.0).values
                & (df_result['limit_ratio'] >= 0.875).values,
                np.maximum(0.0, np.minimum((df_result['h'] / 1000.0) * dc * pvsyst_pr, self.ac_power_all * 0.876) - df_result[inv_id]) * 0.25,
                0.0
            )

            if tx_name in ["TX1", "TX3"]:
                dt_loss_s = np.where(
                    (df_result['h'] > threshold) & (df_result[inv_id] < 1.0),
                    np.where(
                        df_result[f"{tx_name}_Average_Power"] > 1.0,
                        df_result[f"{tx_name}_Average_Power"] * 0.25,
                        (df_result['h'] / 1000.0) * dc * pvsyst_pr * 0.25
                    ),
                    0.0
                )
                # Curtailment only applies to a PRODUCING inverter (>=1 kW). During a full
                # outage the active-power regulation signal also reads ~0 (limit_ratio ~0.001),
                # which would otherwise book a spurious curtailment loss ON TOP of the downtime
                # loss for the same dead interval -- double counting that pushed the compensated
                # PR above 100% on heavy-outage days. Gating on inv>=1 keeps downtime and
                # curtailment mutually exclusive.
                curt_loss_s = np.where(
                    (df_result['limit_ratio'] < 0.875) & (df_result[inv_id] >= 1.0),
                    np.maximum(0.0, np.minimum((df_result['h'] / 1000.0) * dc * pvsyst_pr, self.ac_power_all * 0.876) - self.ac_power_all * df_result['limit_ratio']) * 0.25,
                    0.0
                )
                new_cols[f"{inv_id}_dt_loss"] = dt_loss_s
                new_cols[f"{inv_id}_curt_loss"] = curt_loss_s
                new_cols[f"{inv_id}_ramp_loss"] = ramp_loss_s
                new_cols[f"{inv_id}_loss"] = dt_loss_s + curt_loss_s + ramp_loss_s
            else:
                # Nested logic for TX2:
                # If downtime condition AND average power <= 1.0:
                # loss = min(irrad_expected_power, ac_power_all * 0.876) * 0.25 (no curtailment added)
                # Else:
                # loss = (average_power * 0.25 if downtime and active < 1.0 else 0) + curtailment_loss
                dt_loss_avg_zero = np.minimum((df_result['h'] / 1000.0) * dc * pvsyst_pr, self.ac_power_all * 0.876) * 0.25
                dt_loss_avg_pos = np.where(
                    (df_result['h'] > threshold) & (df_result[inv_id] < 1.0),
                    df_result[f"{tx_name}_Average_Power"] * 0.25,
                    0.0
                )
                # Curtailment only applies to a PRODUCING inverter (>=1 kW). During a full
                # outage the active-power regulation signal also reads ~0 (limit_ratio ~0.001),
                # which would otherwise book a spurious curtailment loss ON TOP of the downtime
                # loss for the same dead interval -- double counting that pushed the compensated
                # PR above 100% on heavy-outage days. Gating on inv>=1 keeps downtime and
                # curtailment mutually exclusive.
                curt_loss_s = np.where(
                    (df_result['limit_ratio'] < 0.875) & (df_result[inv_id] >= 1.0),
                    np.maximum(0.0, np.minimum((df_result['h'] / 1000.0) * dc * pvsyst_pr, self.ac_power_all * 0.876) - self.ac_power_all * df_result['limit_ratio']) * 0.25,
                    0.0
                )
                dt_loss_s = np.where(
                    (df_result['h'] > threshold) & (df_result[inv_id] < 1.0) & (df_result[f"{tx_name}_Average_Power"] <= 1.0),
                    dt_loss_avg_zero,
                    dt_loss_avg_pos
                )
                loss_s = np.where(
                    (df_result['h'] > threshold) & (df_result[inv_id] < 1.0) & (df_result[f"{tx_name}_Average_Power"] <= 1.0),
                    dt_loss_avg_zero,
                    dt_loss_avg_pos + curt_loss_s
                )
                new_cols[f"{inv_id}_dt_loss"] = dt_loss_s
                new_cols[f"{inv_id}_curt_loss"] = curt_loss_s
                new_cols[f"{inv_id}_ramp_loss"] = ramp_loss_s
                new_cols[f"{inv_id}_loss"] = loss_s + ramp_loss_s
            
        for tx in ["TX1", "TX2", "TX3"]:
            loss_cols = [f"{tx}-INV-{i}_loss" for i in range(1, 13)]
            new_cols[f"{tx}_Total_Loss"] = sum(new_cols[col] for col in loss_cols)
            
        df_result = pd.concat([df_result, pd.DataFrame(new_cols, index=df_result.index)], axis=1)
            
        # v10 methodology: the reference irradiance for ALL PR calculations is the
        # threshold-gated Conditional MAX POA sum (Column I). Each poa_cond_max_kwh is
        # already POA/4000 (kWh/m² per 15-min interval: /4 for the quarter-hour, /1000 for
        # W->kW) and already zeroed below the >=50 W/m² threshold. Summing these per-interval
        # kWh values equals (sum of selected POA in W/m²)/4000 -- identical unit treatment to
        # the old h_sum/4000, only the per-interval selection rule differs. Mirrors Excel
        # SUM($I$15:$I$110).
        h_sum_kwh = df_result['poa_cond_max_kwh'].sum()
        
        inv_prs = {}
        inverter_table_data = []
        for inv_id in self.dc_powers:
            dc = self.dc_powers[inv_id]
            energy_gen = df_result[inv_id].sum() * 0.25
            total_loss = df_result[f"{inv_id}_loss"].sum()
            numerator = energy_gen + total_loss
            denominator = dc * h_sum_kwh
            pr_val = (numerator / denominator * 100.0) if denominator > 0 else 0.0
            inv_prs[inv_id] = pr_val
            tx_name = inv_id.split("-")[0]
            inverter_table_data.append((inv_id, tx_name, dc, f"{energy_gen:.2f}", f"{total_loss:.2f}", f"{pr_val:.3f}"))
            
        avg_inv_pr = np.mean(list(inv_prs.values()))
        total_energy = sum(df_result[inv_id].sum() * 0.25 for inv_id in self.dc_powers)
        uncomp_pr = total_energy / (12625.0 * h_sum_kwh) * 100.0
        total_losses_all = sum(df_result[f"{tx}_Total_Loss"].sum() for tx in ["TX1", "TX2", "TX3"])
        comp_raw_pr = (total_energy + total_losses_all) / (12625.0 * h_sum_kwh) * 100.0
        
        calc_results = {
            "avg_inv_pr": avg_inv_pr,
            "comp_raw_pr": comp_raw_pr,
            "uncomp_pr": uncomp_pr,
            "h_sum_kwh": h_sum_kwh,
            "inverter_table_data": inverter_table_data,
            "date_str": date_str,
            "dc_powers": self.dc_powers
        }
        
        # Determine output folder
        if not calcolo_folder:
            parent_folder = os.path.dirname(folder)
            calcolo_folder = os.path.join(parent_folder, "PR CALCOLO FILE")
            
        os.makedirs(calcolo_folder, exist_ok=True)
        
        month_abbrs = {
            1: "gen", 2: "feb", 3: "mar", 4: "apr", 5: "mag", 6: "giu",
            7: "lug", 8: "ago", 9: "set", 10: "ott", 11: "nov", 12: "dic"
        }
        italian_months_4 = {
            1: "GENN", 2: "FEBB", 3: "MARZ", 4: "APRL", 5: "MAGG", 6: "GIUG",
            7: "LUGL", 8: "AGOS", 9: "SETT", 10: "OTTO", 11: "NOVE", 12: "DICE"
        }
        
        date_parts = date_str.split("-")
        year_val = int(date_parts[0])
        month_val = int(date_parts[1])
        day_val = int(date_parts[2])
        month_name = month_abbrs[month_val]
        
        daily_filename = f"PR_recalculation_{day_val:02d}_{month_name}.xlsx"
        daily_file_path = os.path.join(calcolo_folder, daily_filename)
        
        # Identify pristine template daily file from original_format to avoid openpyxl cumulative corruption
        original_format_dir = get_resource_path("original_format")
        orig_templates = glob.glob(os.path.join(original_format_dir, "PR_recalculation_*.xlsx"))
        template_file = None
        for tf in orig_templates:
            bname = os.path.basename(tf)
            if "00 PR_recalculation" not in bname and "_test_" not in bname:
                template_file = tf
                break
        
        # If the file does not exist, copy from pristine template
        if not os.path.exists(daily_file_path):
            if template_file:
                print(f"[{date_str}] Copia del template Excel giornaliero pulito: '{os.path.basename(template_file)}' -> '{daily_filename}'")
                shutil.copy(template_file, daily_file_path)
            else:
                raise FileNotFoundError("Template Excel giornaliero originale non trovato nella cartella original_format!")
        else:
            print(f"[{date_str}] Il file giornaliero di destinazione '{daily_filename}' esiste già; aggiornamento in corso.")
            
        # Open and write values to daily workbook natively via Excel COM to prevent openpyxl table corruption!
        excel_daily = None
        try:
            import win32com.client
            excel_daily = get_excel_app()
            abs_daily_path = os.path.abspath(daily_file_path).replace('/', '\\')
            print(f"[{date_str}] DEBUG: Apertura cartella di lavoro giornaliera: {abs_daily_path}")
            wb_daily = self._open_workbook_writable(excel_daily, abs_daily_path)
            print(f"[{date_str}] DEBUG: Cartella di lavoro aperta con successo!")
            
            try:
                excel_daily.Calculation = -4135  # xlCalculationManual
                excel_daily.CalculateBeforeSave = False
            except Exception:
                pass
                
            ws_calc = wb_daily.Sheets('PR_Calc')
            ws_inv = wb_daily.Sheets('Inverter_data')
            
            dt_obj = datetime.datetime(year_val, month_val, day_val, 0, 0)
            
            # Prepare fast list assignments for 96 rows (Row 15 to 110)
            # Find matching df_result rows for each slot
            time_slots = []
            for r in range(15, 111):
                time_val = ws_calc.Cells(r, 2).Value
                time_str_prefix = str(time_val)[:8] if time_val else ""
                time_slots.append(time_str_prefix)
                
            print(f"[{date_str}] DEBUG: Lettura Colonna B completata (96 righe).")
                
            calc_rows_ordered = []
            for t_prefix in time_slots:
                match = df_result[df_result['time'].str.startswith(t_prefix)]
                if len(match) > 0:
                    calc_rows_ordered.append(match.iloc[0].to_dict())
                else:
                    # fallback dummy row
                    calc_rows_ordered.append({
                        'poa1': 0.0, 'poa1_kwh': 0.0, 'poa3': 0.0, 'poa3_kwh': 0.0,
                        'meter_prev_reading': 0.0, 'meter_reading': 0.0, 'meter_missing': False, 'limit_ratio': 0.876,
                        **{f"TX1-INV-{i}": 0.0 for i in range(1, 13)},
                        **{f"TX2-INV-{i}": 0.0 for i in range(1, 13)},
                        **{f"TX3-INV-{i}": 0.0 for i in range(1, 13)},
                        **{f"TX1-INV-{i}_loss": 0.0 for i in range(1, 13)},
                        **{f"TX2-INV-{i}_loss": 0.0 for i in range(1, 13)},
                        **{f"TX3-INV-{i}_loss": 0.0 for i in range(1, 13)},
                    })
                    
            print(f"[{date_str}] DEBUG: Creazione calc_rows_ordered completata.")
            
            # Set today's date in Column A as a fast date string (avoids sluggish COM datetime serialization)
            date_col_vals = [[f"{year_val:04d}-{month_val:02d}-{day_val:02d}"] for _ in range(96)]
            print(f"[{date_str}] DEBUG: Scrittura date nella Colonna A di ws_calc...")
            ws_calc.Range("A15:A110").Value = date_col_vals
            print(f"[{date_str}] DEBUG: Scrittura date nella Colonna A di ws_inv...")
            ws_inv.Range("A15:A110").Value = date_col_vals
            print(f"[{date_str}] DEBUG: Scrittura date completata. Scrittura colonne dati PR_Calc...")
            
            # Write WS_Calc columns
            ws_calc.Range("C15:C110").Value = [[float(r['poa1'])] for r in calc_rows_ordered]
            ws_calc.Range("D15:D110").Value = [[float(r['poa1_kwh'])] for r in calc_rows_ordered]
            print(f"[{date_str}] DEBUG: Scritte colonne C e D di WS_Calc.")
            ws_calc.Range("E15:E110").Value = [[float(r['poa3'])] for r in calc_rows_ordered]
            ws_calc.Range("F15:F110").Value = [[float(r['poa3_kwh'])] for r in calc_rows_ordered]
            print(f"[{date_str}] DEBUG: Scritte colonne E e F di WS_Calc.")
            ws_calc.Range("K15:K110").Value = [[float(r['meter_prev_reading'])] for r in calc_rows_ordered]
            ws_calc.Range("L15:L110").Value = [[float(r['meter_reading'])] for r in calc_rows_ordered]
            ws_calc.Range("N15:N110").Value = [[float(r['limit_ratio'])] for r in calc_rows_ordered]

            # Highlight any meter reading that was missing/anomalous and got interpolated:
            # orange fill + a cell note on the K/L/M cells of that interval (rows 15..110).
            ORANGE = 39423  # RGB(255,153,0) as Excel BGR-packed long
            note = ("Dato contatore mancante o anomalo (lettura nulla o decrescente): "
                    "valore interpolato come media delle letture adiacenti valide.")
            for ridx, r in enumerate(calc_rows_ordered):
                if not r.get('meter_missing'):
                    continue
                xl_row = 15 + ridx
                for col in (11, 12, 13):  # K = prec., L = lettura, M = energia
                    try:
                        cell = ws_calc.Cells(xl_row, col)
                        cell.Interior.Color = ORANGE
                    except Exception:
                        pass
                try:
                    note_cell = ws_calc.Cells(xl_row, 12)  # put the note on column L
                    if note_cell.Comment is not None:
                        note_cell.Comment.Delete()
                    note_cell.AddComment(note)
                    note_cell.Comment.Visible = False
                except Exception as note_err:
                    print(f"[{date_str}] DEBUG: impossibile aggiungere nota cella meter: {note_err}")
            
            # Permanently eliminate #DIV/0! errors from daily workbook formula columns.
            # v11 efficiency: write each column in a single COM call (one 96x1 array)
            # instead of 96 per-cell round-trips per column. Cuts ~480 COM calls/day
            # down to 5. Formula strings are literal per-row, so Excel stores them as-is.
            _rows = range(15, 111)
            ws_calc.Range("G15:G110").Formula = [[f"=IFERROR((D{r}+F{r})/2, 0)"] for r in _rows]
            ws_calc.Range("H15:H110").Formula = [[f"=IFERROR(IF(OR(C{r}=0,E{r}=0),IF(MAX(C{r},E{r})>$BA$7,MAX(C{r},E{r}),0),IF(AVERAGE(C{r},E{r})>$BA$7,AVERAGE(C{r},E{r}),0)), 0)"] for r in _rows]
            # Column I = the PR reference POA (kWh/m²), gated by the PR threshold: the
            # selected POA (x4000 -> W/m²) must be >= $BA$7, otherwise the interval is 0.
            # The selection rule depends on the chosen method:
            #   - "average": two-sensor mean when both read (MAX only if one sensor is 0)
            #   - "condmax": as average, but use MAX when the two diverge beyond $BA$6
            if poa_method == "average":
                _sel = lambda r: f"IF(AND(D{r}=0,F{r}=0),0,IF(OR(D{r}=0,F{r}=0),MAX(D{r},F{r}),G{r}))"
            else:
                _sel = lambda r: f"IF(AND(D{r}=0,F{r}=0),0,IF(OR(D{r}=0,F{r}=0),MAX(D{r},F{r}),IF(J{r}>$BA$6,MAX(D{r},F{r}),G{r})))"
            ws_calc.Range("I15:I110").Formula = [[f"=IFERROR(IF(({_sel(r)})*4000>=$BA$7,{_sel(r)},0), 0)"] for r in _rows]
            ws_calc.Range("J15:J110").Formula = [[f"=IFERROR(IF(AND(D{r}>0,F{r}>0),ABS(D{r}-F{r})/AVERAGE(D{r},F{r}),0), 0)"] for r in _rows]
            ws_calc.Range("M15:M110").Formula = [[f"=IFERROR((L{r}-K{r})*1000, 0)"] for r in _rows]
                
            # Programmatically enforce correct inverter PR formulas in row 111 (without losses)
            import openpyxl.utils
            for i in range(1, 13):
                # TX1 (Columns O to Z -> 15 to 26)
                col_calc = 14 + i
                col_inv = 2 + i
                col_calc_letter = openpyxl.utils.get_column_letter(col_calc)
                inv_col_letter = openpyxl.utils.get_column_letter(col_inv)
                ws_calc.Cells(111, col_calc).Formula2 = f"=SUM((Inverter_data!{inv_col_letter}15:{inv_col_letter}110)*0.25)/({col_calc_letter}$10*SUM(PR_Calc!$I$15:$I$110))"
                
                # TX2 (Columns AB to AM -> 28 to 39)
                col_calc = 27 + i
                col_inv = 17 + i
                col_calc_letter = openpyxl.utils.get_column_letter(col_calc)
                inv_col_letter = openpyxl.utils.get_column_letter(col_inv)
                ws_calc.Cells(111, col_calc).Formula2 = f"=SUM((Inverter_data!{inv_col_letter}15:{inv_col_letter}110)*0.25)/({col_calc_letter}$10*SUM(PR_Calc!$I$15:$I$110))"
                
                # TX3 (Columns AO to AZ -> 41 to 52)
                col_calc = 40 + i
                col_inv = 32 + i
                col_calc_letter = openpyxl.utils.get_column_letter(col_calc)
                inv_col_letter = openpyxl.utils.get_column_letter(col_inv)
                ws_calc.Cells(111, col_calc).Formula2 = f"=SUM((Inverter_data!{inv_col_letter}15:{inv_col_letter}110)*0.25)/({col_calc_letter}$10*SUM(PR_Calc!$I$15:$I$110))"
                
            print(f"[{date_str}] DEBUG: Scrittura dati PR_Calc e formule riga 111 completata.")
            
            # Write Inverter_data columns
            tx1_vals = [[float(r[f"TX1-INV-{i}"]) for i in range(1, 13)] for r in calc_rows_ordered]
            ws_inv.Range("C15:N110").Value = tx1_vals
            
            tx2_vals = [[float(r[f"TX2-INV-{i}"]) for i in range(1, 13)] for r in calc_rows_ordered]
            ws_inv.Range("R15:AC110").Value = tx2_vals
            
            tx3_vals = [[float(r[f"TX3-INV-{i}"]) for i in range(1, 13)] for r in calc_rows_ordered]
            ws_inv.Range("AG15:AR110").Value = tx3_vals
            print(f"[{date_str}] DEBUG: Scrittura potenze Inverter_data completata.")

            # Write Python-computed per-inverter energy losses directly to PR_Calc (cols O-Z, AB-AM, AO-AZ)
            # This overrides the Excel array formulas and ensures losses are exactly what Python calculated,
            # independent of BA4 or any other Excel formula parameter.
            tx1_loss_vals = [[float(r.get(f"TX1-INV-{i}_loss", 0.0)) for i in range(1, 13)] for r in calc_rows_ordered]
            ws_calc.Range("O15:Z110").Value = tx1_loss_vals
            tx2_loss_vals = [[float(r.get(f"TX2-INV-{i}_loss", 0.0)) for i in range(1, 13)] for r in calc_rows_ordered]
            ws_calc.Range("AB15:AM110").Value = tx2_loss_vals
            tx3_loss_vals = [[float(r.get(f"TX3-INV-{i}_loss", 0.0)) for i in range(1, 13)] for r in calc_rows_ordered]
            ws_calc.Range("AO15:AZ110").Value = tx3_loss_vals
            print(f"[{date_str}] DEBUG: Scrittura perdite energia per inverter completata (O-Z, AB-AM, AO-AZ).")

            # Update titles and side tables
            italian_months_full_upper = {
                1: "GENNAIO", 2: "FEBBRAIO", 3: "MARZO", 4: "APRILE", 5: "MAGGIO", 6: "GIUGNO",
                7: "LUGLIO", 8: "AGOSTO", 9: "SETTEMBRE", 10: "OTTOBRE", 11: "NOVEMBRE", 12: "DICEMBRE"
            }
            english_months_full = {
                1: "January", 2: "February", 3: "March", 4: "April", 5: "May", 6: "June",
                7: "July", 8: "August", 9: "September", 10: "October", 11: "November", 12: "December"
            }
            
            ws_calc.Cells(1, 1).Value = "PR CALCULATION"
            ws_calc.Cells(2, 1).Value = f"{day_val:02d} {italian_months_full_upper[month_val]} {year_val}"
            ws_calc.Cells(8, 1).Value = "NOMINAL VALUES"
            
            # Write nominal parameters in Column BA (53)
            ws_calc.Cells(4, 53).Value = float(pvsyst_pr)           # BA4 = decimal PR (e.g. 0.897)
            # BA5 = uncompensated (RAW) PR. Use DIRECT ranges (identical to BH11's
            # numerator) instead of a structured table reference: the Inverter_data
            # table is named differently across files (e.g. Tabella01MarzoInverter vs
            # Tabella24MarzoInverter), and a missing table name makes Excel reject this
            # formula with 0x800A0BEC. Direct ranges have no table-name dependency.
            ws_calc.Cells(5, 53).Formula = "=((SUM(Inverter_data!C15:N110, Inverter_data!R15:AC110, Inverter_data!AG15:AR110) * 0.25)) / (12625 * SUM($I$15:$I$110))"
            ws_calc.Cells(6, 53).Value = float(diff_threshold)                     # BA6 is irradiance acceptance limit ratio
            ws_calc.Cells(7, 53).Value = float(threshold)          # BA7 is irradiance minimum value (e.g. 50)
            
            # Write English PR calculation header in Column BD (56) Row 2
            ws_calc.Cells(2, 56).Value = f"{day_val} {english_months_full[month_val]} {year_val} PR Calculation"
            
            # Write PR from SCADA in Column BH (60) Row 8 (BH8 expects percentage e.g. 81.743)
            ws_calc.Cells(8, 60).Value = float(uncomp_pr)
            
            # Write PR Compensated in Column BD (56) Row 11 and Column BH (60) Row 11 (percentage e.g. 81.743)
            ws_calc.Cells(11, 56).Value = "PR Compensated [%]"
            ws_calc.Cells(11, 60).Formula = "=((SUM(Inverter_data!C15:N110, Inverter_data!R15:AC110, Inverter_data!AG15:AR110)*0.25 + AA111 + AN111 + BA111) / (12625 * SUM($I$15:$I$110))) * 100"
            
            print(f"[{date_str}] DEBUG: Aggiornamento tabelle laterali completato. Salvataggio cartella di lavoro in corso...")
            
            try:
                excel_daily.Calculation = -4105  # xlCalculationAutomatic
            except Exception:
                pass
                
            self._save_workbook_resilient(wb_daily, excel_daily, abs_daily_path, date_str)
            print(f"[{date_str}] File giornaliero '{daily_filename}' salvato con successo via Excel COM!")
        except Exception as ex:
            print(f"[{date_str}] Errore durante l'aggiornamento del file giornaliero via Excel COM: {ex}")
            raise ex
        finally:
            # ALWAYS release the daily workbook. Otherwise, in a month-long batch the shared
            # (invisible) Excel instance accumulates one open network workbook per day; since
            # Calculation is an app-level setting, each day then recalculates EVERY still-open
            # workbook, and the growing load wedged Save() around the ~20th file. Closing here
            # keeps exactly one daily workbook open at a time.
            try:
                if 'wb_daily' in locals() and wb_daily is not None:
                    wb_daily.Close(SaveChanges=False)
            except Exception:
                pass
            
        # Update mother file (skippable from the Opzioni Avanzate dialog)
        if not skip_mother_update:
            if self.sync_mother_var.get():
                self.sync_mother_file(calcolo_folder, year_val, month_val)
            else:
                print(">>> Sincronizzazione del file Madre disattivata nelle Opzioni Avanzate: saltata.")
            
        return df_result, calc_results

    def _pr_degradation_factor(self, year, month):
        """Contractual PR-target degradation (Allegato 9.1): the guaranteed PR degrades
        `rate` (0.4% by default) per year, compounding annually, from the plant start
        (Feb 2025 by default). Rate and start date are configurable in Opzioni Avanzate.
        Contract years run start_month..start_month-1 (Feb..Jan); Year 1 is undegraded
        (n=0). Returns (n, factor) with factor = (1-rate)**n."""
        start_year = int(self.cfg["deg_start_year"])
        start_month = int(self.cfg["deg_start_month"])
        rate = float(self.cfg["deg_rate"])
        anchor_year = year if month >= start_month else year - 1
        n = max(0, anchor_year - start_year)
        return n, (1.0 - rate) ** n

    def find_scada_pr_file(self, folder):
        """Locate the SCADA KPI daily PR report in `folder` or its subfolders."""
        if not folder or not os.path.exists(folder):
            return None
        patterns = [
            "KPI_Report_Daily.xls",
            "KPI_Report_Daily.xlsx",
            "KPI_Report*.xls*",
            "*KPI_Report*.xls*",
            "*SCADA*PR*.xls*",
            "*KPI*.xls*",
            "*SCADA*.xls*",
            "*KPI*.csv",
            "*SCADA*.csv",
        ]
        # 1. Search in target folder directly
        for pat in patterns:
            matches = glob.glob(os.path.join(folder, pat))
            if matches:
                return matches[0]
            try:
                for item in os.listdir(folder):
                    item_path = os.path.join(folder, item)
                    if os.path.isfile(item_path):
                        core = pat.replace("*", "").lower()
                        if core and core in item.lower():
                            return item_path
            except Exception:
                pass
        # 2. Search subfolders (excluding hidden / PR CALCOLO FILE)
        try:
            for subd in os.listdir(folder):
                subd_path = os.path.join(folder, subd)
                if os.path.isdir(subd_path) and not subd.startswith("_") and subd != "PR CALCOLO FILE":
                    for pat in patterns:
                        matches = glob.glob(os.path.join(subd_path, pat))
                        if matches:
                            return matches[0]
        except Exception:
            pass
        return None

    def find_vcom_pr_file(self, folder):
        """Locate the VCOM Performance Ratio CSV/Excel export in `folder` or its subfolders."""
        if not folder or not os.path.exists(folder):
            return None
        patterns = [
            "Performance_ratio_vcom.csv",
            "Performance_ratio_*.csv",
            "*Performance_ratio*.csv",
            "*vcom*pr*.csv",
            "*PR*vcom*.csv",
            "*vcom*.csv",
            "*Performance_ratio*.xls*",
            "*vcom*.xls*",
        ]
        # 1. Search in target folder directly
        for pat in patterns:
            matches = glob.glob(os.path.join(folder, pat))
            if matches:
                return matches[0]
            try:
                for item in os.listdir(folder):
                    item_path = os.path.join(folder, item)
                    if os.path.isfile(item_path):
                        core = pat.replace("*", "").lower()
                        if core and core in item.lower():
                            return item_path
            except Exception:
                pass
        # 2. Search subfolders
        try:
            for subd in os.listdir(folder):
                subd_path = os.path.join(folder, subd)
                if os.path.isdir(subd_path) and not subd.startswith("_") and subd != "PR CALCOLO FILE":
                    for pat in patterns:
                        matches = glob.glob(os.path.join(subd_path, pat))
                        if matches:
                            return matches[0]
        except Exception:
            pass
        return None

    def _read_scada_daily_pr(self, month_folder_or_path, year=None, month=None):
        """Daily SCADA PR (%) from 'KPI_Report_Daily.xls' (or matching file/folder).
        The KPI export lists Month, Date and the daily PR%. Returns {day:int -> pr:float}."""
        if not month_folder_or_path:
            return {}
        if os.path.isdir(month_folder_or_path):
            path = self.find_scada_pr_file(month_folder_or_path)
        else:
            path = month_folder_or_path
        if not path or not os.path.exists(path):
            return {}
        out = {}
        try:
            if path.lower().endswith(".csv"):
                with open(path, "rb") as f:
                    raw = f.read()
                txt = None
                if raw.startswith((b'\xff\xfe', b'\xfe\xff')) or b'\x00' in raw[:1024]:
                    try:
                        txt = raw.decode("utf-16")
                    except Exception:
                        pass
                if txt is None:
                    for enc in ["utf-8-sig", "utf-8", "cp1252", "latin1"]:
                        try:
                            txt = raw.decode(enc)
                            break
                        except Exception:
                            continue
                if txt is None:
                    txt = raw.decode("utf-8", errors="replace")

                import re
                for line in txt.splitlines():
                    for sep in ["\t", ";", ","]:
                        parts = [c.strip().strip('"') for c in line.split(sep)]
                        if len(parts) >= 3 and re.fullmatch(r"\d{1,2}", parts[0]):
                            try:
                                out[int(parts[0])] = round(float(parts[2].replace(",", ".")), 3)
                                break
                            except Exception:
                                pass
            else:
                import pandas as pd
                df = pd.read_excel(path, header=None)
                for _, row in df.iterrows():
                    try:
                        ts = pd.to_datetime(row[2])
                        pr = float(str(row[3]).replace(",", "."))
                    except Exception:
                        continue
                    if (year is None or ts.year == year) and (month is None or ts.month == month):
                        out[int(ts.day)] = round(pr, 3)
        except Exception as e:
            print(f"DEBUG: lettura SCADA '{os.path.basename(path)}' fallita: {e}")
        return out

    def _read_vcom_daily_pr(self, month_folder_or_path):
        """Daily VCOM PR (%) from 'Performance_ratio*.csv' (or matching file/folder).
        The file is UTF-16 / UTF-8, tab/comma-separated, Italian decimals; the 'Data' column
        holds the day-of-month and the 3rd column the PR%. Returns {day:int -> pr:float}."""
        if not month_folder_or_path:
            return {}
        if os.path.isdir(month_folder_or_path):
            path = self.find_vcom_pr_file(month_folder_or_path)
        else:
            path = month_folder_or_path
        if not path or not os.path.exists(path):
            return {}
        import re
        out = {}
        try:
            with open(path, "rb") as f:
                raw = f.read()
            txt = None
            if raw.startswith((b'\xff\xfe', b'\xfe\xff')) or b'\x00' in raw[:1024]:
                try:
                    txt = raw.decode("utf-16")
                except Exception:
                    pass
            if txt is None:
                for enc in ["utf-8-sig", "utf-8", "cp1252", "latin1"]:
                    try:
                        txt = raw.decode(enc)
                        break
                    except Exception:
                        continue
            if txt is None:
                txt = raw.decode("utf-8", errors="replace")

            for line in txt.splitlines():
                for sep in ["\t", ";", ","]:
                    parts = [c.strip().strip('"') for c in line.split(sep)]
                    if len(parts) >= 3 and re.fullmatch(r"\d{1,2}", parts[0]):
                        try:
                            out[int(parts[0])] = round(float(parts[2].replace(",", ".")), 3)
                            break
                        except Exception:
                            continue
        except Exception as e:
            print(f"DEBUG: lettura VCOM '{os.path.basename(path)}' fallita: {e}")
        return out

    def sync_scada_pr(self):
        """Action handler for 'Sync SCADA PR' button."""
        threading.Thread(target=self._sync_vendor_pr_worker, args=("SCADA",), daemon=True).start()

    def sync_vcom_pr(self):
        """Action handler for 'Sync VCOM PR' button."""
        threading.Thread(target=self._sync_vendor_pr_worker, args=("VCOM",), daemon=True).start()

    def _sync_vendor_pr_worker(self, vendor_type):
        """Worker to locate SCADA or VCOM file and synchronize into Mother Excel file."""
        import calendar
        import datetime
        import shutil
        import glob

        self.root.after(0, lambda: [
            self.btn_calculate.config(state="disabled"),
            self.btn_stop.config(state="disabled"),
            self.btn_sync_scada.config(state="disabled"),
            self.btn_sync_vcom.config(state="disabled"),
            self.lbl_status.config(text=f"Sincronizzazione PR {vendor_type} in corso...", foreground=self.warn_color)
        ])

        try:
            folder = self.folder_path_var.get().strip()
            if not folder or not os.path.exists(folder):
                # Prompt user to select month folder
                folder = filedialog.askdirectory(title="Seleziona la Cartella Mese (formato AAAA MM)")
                if not folder:
                    print(f">>> Sincronizzazione {vendor_type} annullata: nessuna cartella selezionata.")
                    return
                self.folder_path_var.set(folder)

            # If user selected a subfolder (like 'PR CALCOLO FILE' or a day folder '01'), step up to month folder
            bname = os.path.basename(folder.rstrip("\\/"))
            if bname == "PR CALCOLO FILE" or bname.isdigit():
                folder = os.path.dirname(os.path.abspath(folder))
                bname = os.path.basename(folder.rstrip("\\/"))

            # Determine Year and Month
            parts = bname.split()
            year_val = None
            month_val = None
            if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
                year_val = int(parts[0])
                month_val = int(parts[1])
            else:
                date_str = self.date_var.get().strip()
                if date_str:
                    try:
                        d_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d")
                        year_val = d_obj.year
                        month_val = d_obj.month
                    except Exception:
                        pass
            if year_val is None or month_val is None:
                now = datetime.datetime.now()
                year_val = now.year
                month_val = now.month

            month_abbrs = {
                1: "gen", 2: "feb", 3: "mar", 4: "apr", 5: "mag", 6: "giu",
                7: "lug", 8: "ago", 9: "set", 10: "ott", 11: "nov", 12: "dic"
            }
            italian_months_4 = {
                1: "GENN", 2: "FEBB", 3: "MARZ", 4: "APRL", 5: "MAGG", 6: "GIUG",
                7: "LUGL", 8: "AGOS", 9: "SETT", 10: "OTTO", 11: "NOVE", 12: "DICE"
            }
            month_name = month_abbrs[month_val]
            expected_mother_filename = f"00 PR_recalculation_{italian_months_4[month_val]}.xlsx"

            alt_mother_filename = f"00_PR_recalculation_{italian_months_4[month_val]}.xlsx"

            # 1. Locate vendor PR file
            print(f">>> Ricerca file {vendor_type} per {month_val:02d}/{year_val} in '{folder}'...")
            if vendor_type == "SCADA":
                vendor_file = self.find_scada_pr_file(folder)
                if not vendor_file:
                    vendor_file = filedialog.askopenfilename(
                        title="Seleziona file SCADA KPI (es. KPI_Report_Daily.xls)",
                        filetypes=[("File Excel/CSV SCADA", "*.xls;*.xlsx;*.csv"), ("Tutti i file", "*.*")]
                    )
            else:
                vendor_file = self.find_vcom_pr_file(folder)
                if not vendor_file:
                    vendor_file = filedialog.askopenfilename(
                        title="Seleziona file VCOM PR (es. Performance_ratio_vcom.csv)",
                        filetypes=[("File CSV/Excel VCOM", "*.csv;*.xlsx;*.xls"), ("Tutti i file", "*.*")]
                    )

            if not vendor_file or not os.path.exists(vendor_file):
                print(f">>> Sincronizzazione {vendor_type} annullata: file non trovato o non selezionato.")
                return

            print(f">>> File {vendor_type} trovato: '{os.path.basename(vendor_file)}'")

            # 2. Read daily PR data
            if vendor_type == "SCADA":
                pr_data = self._read_scada_daily_pr(vendor_file, year_val, month_val)
            else:
                pr_data = self._read_vcom_daily_pr(vendor_file)

            if not pr_data:
                msg_err = f"Nessun dato PR valido trovato per {month_val:02d}/{year_val} in:\n{vendor_file}"
                print(f">>> ERRORE: {msg_err}")
                self.root.after(0, lambda: messagebox.showwarning("Dati non trovati", msg_err))
                return

            print(f">>> Dati PR {vendor_type} estratti: {len(pr_data)} giorni (es. {sorted(pr_data.keys())[:5]}...)")

            # 3. Locate Mother file in 'PR CALCOLO FILE'
            calcolo_folder = os.path.join(folder, "PR CALCOLO FILE")
            os.makedirs(calcolo_folder, exist_ok=True)
            mother_path = os.path.join(calcolo_folder, expected_mother_filename)

            initialized_new = False
            if not os.path.exists(mother_path) and os.path.exists(os.path.join(calcolo_folder, alt_mother_filename)):
                mother_path = os.path.join(calcolo_folder, alt_mother_filename)
            elif not os.path.exists(mother_path):
                existing_mothers = glob.glob(os.path.join(calcolo_folder, "00*PR_recalculation_*.xlsx"))
                if existing_mothers:
                    mother_path = existing_mothers[0]
                else:
                    original_format_dir = get_resource_path("original_format")
                    orig_mothers = glob.glob(os.path.join(original_format_dir, "00*PR_recalculation_*.xlsx"))
                    if orig_mothers:
                        orig_mother_path = orig_mothers[0]
                        print(f">>> Inizializzazione template Madre: '{os.path.basename(orig_mother_path)}' -> '{expected_mother_filename}'")
                        shutil.copy(orig_mother_path, mother_path)
                        initialized_new = True
                    else:
                        raise FileNotFoundError(f"Template Madre non trovato per inizializzare '{expected_mother_filename}'")

            # 4. Optional backup
            if self.backup_mother_var.get():
                try:
                    backup_dir = os.path.join(calcolo_folder, "_backups")
                    os.makedirs(backup_dir, exist_ok=True)
                    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    backup_path = os.path.join(backup_dir, f"{os.path.splitext(os.path.basename(mother_path))[0]}_{vendor_type}_{ts}.xlsx")
                    shutil.copy2(mother_path, backup_path)
                    print(f">>> Backup file Madre creato: {os.path.basename(backup_path)}")
                except Exception as bk_err:
                    print(f">>> Warning backup file Madre: {bk_err}")

            # 5. Open Mother file and update target column
            excel = None
            wb_mother = None
            try:
                excel = get_excel_app()
                abs_mother_path = os.path.abspath(mother_path).replace('/', '\\')
                print(f">>> Apertura file Madre per aggiornamento {vendor_type}: {abs_mother_path}")
                wb_mother = self._open_workbook_writable(excel, abs_mother_path)

                try:
                    excel.Calculation = -4135  # xlCalculationManual
                except Exception:
                    pass

                ws_mother = wb_mother.Sheets('PR_Calc')
                num_days = calendar.monthrange(year_val, month_val)[1]

                # Find target column in Row 4
                target_col = None
                col_name_found = ""
                for col in range(2, 65):
                    hv = str(ws_mother.Cells(4, col).Value or "").strip().lower()
                    if vendor_type == "SCADA" and "pr scada" in hv:
                        target_col = col
                        col_name_found = str(ws_mother.Cells(4, col).Value or "").strip()
                        break
                    elif vendor_type == "VCOM" and ("pr vcom" in hv or ("pr total" in hv and not target_col)):
                        target_col = col
                        col_name_found = str(ws_mother.Cells(4, col).Value or "").strip()
                        if "pr vcom" in hv:
                            break

                if target_col is None:
                    target_col = 7 if vendor_type == "SCADA" else 8
                    col_name_found = f"Colonna {target_col} (Default {vendor_type})"

                print(f">>> Aggiornamento colonna '{col_name_found}' (Col {target_col}) nel file Madre...")

                updated_count = 0
                for day_num in range(1, num_days + 1):
                    r = 5 + day_num - 1
                    if day_num in pr_data:
                        val = pr_data[day_num]
                        ws_mother.Cells(r, target_col).Value = val
                        updated_count += 1

                try:
                    excel.Calculation = -4105  # xlCalculationAutomatic
                    excel.CalculateBeforeSave = True
                    wb_mother.Calculate()
                except Exception:
                    pass

                self._save_workbook_resilient(wb_mother, excel, abs_mother_path, f"MADRE_{vendor_type}_SYNC")
                try:
                    wb_mother.Close(SaveChanges=False)
                except Exception:
                    pass

                succ_msg = (
                    f"Sincronizzazione PR {vendor_type} completata con successo!\n\n"
                    f"File sorgente: {os.path.basename(vendor_file)}\n"
                    f"File Madre: {os.path.basename(mother_path)}\n"
                    f"Colonna aggiornata: {col_name_found} (Col {target_col})\n"
                    f"Giorni aggiornati: {updated_count} / {num_days}"
                )
                print(f">>> {succ_msg.replace(chr(10), ' ')}")
                self.root.after(0, lambda: [
                    self.lbl_status.config(text=f"Sync {vendor_type} completato ({updated_count} giorni)!", foreground=self.success_color),
                    messagebox.showinfo(f"Sync {vendor_type} Riuscito", succ_msg)
                ])

            except Exception as ex:
                print(f">>> Errore durante l'aggiornamento del file Madre ({vendor_type}): {ex}")
                if wb_mother:
                    try:
                        wb_mother.Close(SaveChanges=False)
                    except Exception:
                        pass
                err_msg = f"Impossibile completare la sincronizzazione {vendor_type}:\n{ex}"
                self.root.after(0, lambda: [
                    self.lbl_status.config(text=f"Errore sync {vendor_type}!", foreground="red"),
                    messagebox.showerror(f"Errore Sync {vendor_type}", err_msg)
                ])

        except Exception as ex_all:
            print(f">>> Errore generale sync {vendor_type}: {ex_all}")
            self.root.after(0, lambda: [
                self.lbl_status.config(text=f"Errore sync {vendor_type}!", foreground="red"),
                messagebox.showerror(f"Errore Sync {vendor_type}", str(ex_all))
            ])
        finally:
            self.root.after(0, self._reset_run_buttons)

    def sync_mother_file(self, calcolo_folder, year_val, month_val, vcom_days=None):
        import calendar
        import openpyxl
        import re
        import shutil
        import glob
        
        month_abbrs = {
            1: "gen", 2: "feb", 3: "mar", 4: "apr", 5: "mag", 6: "giu",
            7: "lug", 8: "ago", 9: "set", 10: "ott", 11: "nov", 12: "dic"
        }
        italian_months_4 = {
            1: "GENN", 2: "FEBB", 3: "MARZ", 4: "APRL", 5: "MAGG", 6: "GIUG",
            7: "LUGL", 8: "AGOS", 9: "SETT", 10: "OTTO", 11: "NOVE", 12: "DICE"
        }
        
        month_name = month_abbrs[month_val]
        expected_mother_filename = f"00 PR_recalculation_{italian_months_4[month_val]}.xlsx"

        alt_mother_filename = f"00_PR_recalculation_{italian_months_4[month_val]}.xlsx"
        mother_path = os.path.join(calcolo_folder, expected_mother_filename)
        
        initialized_new = False
        if not os.path.exists(mother_path) and os.path.exists(os.path.join(calcolo_folder, alt_mother_filename)):
            mother_path = os.path.join(calcolo_folder, alt_mother_filename)
        elif not os.path.exists(mother_path):
            original_format_dir = get_resource_path("original_format")
            orig_mothers = glob.glob(os.path.join(original_format_dir, "00*PR_recalculation_*.xlsx"))
            if orig_mothers:
                orig_mother_path = orig_mothers[0]
                print(f"Copia e inizializzazione del template Madre: '{os.path.basename(orig_mother_path)}' -> '{expected_mother_filename}'")
                shutil.copy(orig_mother_path, mother_path)
                initialized_new = True
            else:
                raise FileNotFoundError(f"Template Madre originale non trovato in original_format per inizializzare '{expected_mother_filename}'!")
        elif self.backup_mother_var.get():
            # v11 safety: snapshot the existing Madre file before we modify/overwrite
            # it, so a crash mid-save (or a bad sync) can be rolled back. Keep the last
            # few timestamped copies in a hidden _backups folder next to the file.
            try:
                backup_dir = os.path.join(calcolo_folder, "_backups")
                os.makedirs(backup_dir, exist_ok=True)
                ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_path = os.path.join(backup_dir, f"{os.path.splitext(expected_mother_filename)[0]}_{ts}.xlsx")
                shutil.copy2(mother_path, backup_path)
                existing = sorted(glob.glob(os.path.join(backup_dir, f"{os.path.splitext(expected_mother_filename)[0]}_*.xlsx")))
                for stale in existing[:-5]:
                    try:
                        os.remove(stale)
                    except Exception:
                        pass
                print(f"DEBUG: Backup file Madre creato: {os.path.basename(backup_path)}")
            except Exception as bk_err:
                print(f"DEBUG Warning: backup file Madre non riuscito (non fatale): {bk_err}")

        excel = None
        wb_mother = None
        try:
            excel = get_excel_app()
            abs_mother_path = os.path.abspath(mother_path).replace('/', '\\')
            print(f"DEBUG: Apertura file Madre: {abs_mother_path}")
            wb_mother = self._open_workbook_writable(excel, abs_mother_path)
            
            try:
                excel.Calculation = -4135  # xlCalculationManual
            except Exception:
                pass
                
            ws_mother = wb_mother.Sheets('PR_Calc')
            num_days = calendar.monthrange(year_val, month_val)[1]
            target_summary_row = 5 + num_days
            
            # Programmatically ensure "Irradiance TX1" and "Irradiance TX3" are present at Columns B and C
            b4_val = ws_mother.Cells(4, 2).Value
            if not (b4_val and "Irradiance TX1" in str(b4_val)):
                print("DEBUG: Irradiance TX1 column not found at Column B. Programmatically inserting Irradiance TX1 and Irradiance TX3 columns...")
                ws_mother.Columns(2).Insert()
                ws_mother.Columns(2).Insert()
                ws_mother.Cells(4, 2).Value = "Irradiance TX1"
                ws_mother.Cells(4, 3).Value = "Irradiance TX3"

            # Ensure number format for Irradiance columns is set to 4 decimal places (not date format inherited from Column A)
            ws_mother.Range(f"B5:D{5+num_days-1}").NumberFormatLocal = "0,0000"

            # Ensure B4, C4, I4, J4 header cells have the exact same style and borders as D4
            ref_cell = ws_mother.Cells(4, 4)
            for rng_str in ["B4:C4", "I4:J4"]:
                target_hdr = ws_mother.Range(rng_str)
                try:
                    target_hdr.Interior.Color = ref_cell.Interior.Color
                    target_hdr.Font.Name = ref_cell.Font.Name
                    target_hdr.Font.Size = ref_cell.Font.Size
                    target_hdr.Font.Bold = ref_cell.Font.Bold
                    target_hdr.Font.Color = ref_cell.Font.Color
                    target_hdr.HorizontalAlignment = ref_cell.HorizontalAlignment
                    target_hdr.VerticalAlignment = ref_cell.VerticalAlignment
                    target_hdr.WrapText = ref_cell.WrapText
                    
                    # Copy borders
                    for b_id in [7, 8, 9, 10, 11, 12]:
                        try:
                            target_hdr.Borders(b_id).LineStyle = ref_cell.Borders(b_id).LineStyle
                            target_hdr.Borders(b_id).Weight = ref_cell.Borders(b_id).Weight
                            target_hdr.Borders(b_id).Color = ref_cell.Borders(b_id).Color
                        except Exception:
                            pass
                except Exception as fmt_err:
                    print(f"DEBUG Warning: non-fatal header style copy error for {rng_str}: {fmt_err}")

            # Remove solid green background fill from day rows for Irradiance TX1 and TX3 (make transparent)
            try:
                ws_mother.Range(f"B5:C{5+num_days-1}").Interior.ColorIndex = -4142 # xlNone = -4142
            except Exception as fill_err:
                print(f"DEBUG Warning: non-fatal fill reset error: {fill_err}")

            # Programmatically ensure "Meter Reading" column is present before "Energy (day)"
            meter_reading_col = None
            energy_col_idx = None
            for c_idx in range(2, 20):
                v_str = str(ws_mother.Cells(4, c_idx).Value or "").strip().lower()
                if "meter reading" in v_str or "lettura contatore" in v_str:
                    meter_reading_col = c_idx
                elif "energy" in v_str and "loss" not in v_str and "perdita" not in v_str:
                    if not energy_col_idx:
                        energy_col_idx = c_idx
                        
            if not meter_reading_col and energy_col_idx:
                print(f"DEBUG: Inserimento colonna 'Meter Reading' prima di Energy (day) alla colonna {energy_col_idx}...")
                ws_mother.Columns(energy_col_idx).Insert()
                ws_mother.Cells(4, energy_col_idx).Value = "Meter Reading\n[MWh]"
                try:
                    ws_mother.Range(f"{openpyxl.utils.get_column_letter(energy_col_idx)}5:{openpyxl.utils.get_column_letter(energy_col_idx)}{5+num_days-1}").NumberFormatLocal = "0,000"
                except Exception:
                    pass

            # Programmatically ensure "PR VCOM", "PR Compensated", and "External Availability [%]" columns are present at Columns H, I, and J
            h4_val = ws_mother.Cells(4, 8).Value
            if not (h4_val and "VCOM" in str(h4_val)):
                print("DEBUG: PR VCOM column not found at Column H. Programmatically inserting...")
                ws_mother.Columns(8).Insert()
                ws_mother.Cells(4, 8).Value = "PR VCOM"
                
            i4_val = ws_mother.Cells(4, 9).Value
            if not (i4_val and ("compensated" in str(i4_val).lower() or "compensato" in str(i4_val).lower())):
                print("DEBUG: PR Compensated column not found at Column I. Programmatically inserting...")
                ws_mother.Columns(9).Insert()
                ws_mother.Cells(4, 9).Value = "PR Compensated"
                
            j4_val = ws_mother.Cells(4, 10).Value
            if not (j4_val and ("availability" in str(j4_val).lower() or "disponibilità" in str(j4_val).lower())):
                print("DEBUG: External Availability column not found at Column J. Programmatically inserting...")
                ws_mother.Columns(10).Insert()
                ws_mother.Cells(4, 10).Value = "External Availability\n[%]"
                
            # Dynamically format and adjust summary row in existing Mother file if needed!
            current_summary_row = None
            for r in range(30, 42):
                c1_val = str(ws_mother.Cells(r, 1).Value or "")
                if "-" not in c1_val and "/" not in c1_val:
                    for c_chk in range(2, 15):
                        f_text = str(ws_mother.Cells(r, c_chk).Formula or "").upper()
                        if "AVERAGE" in f_text or "SUM" in f_text:
                            current_summary_row = r
                            break
                    if current_summary_row is not None:
                        break
                    
            if current_summary_row is not None:
                if current_summary_row < target_summary_row:
                    rows_to_insert = target_summary_row - current_summary_row
                    for _ in range(rows_to_insert):
                        ws_mother.Rows(current_summary_row).Insert()
                        ws_mother.Rows(current_summary_row - 1).Copy(ws_mother.Rows(current_summary_row))
                        excel.CutCopyMode = False
                        for c in range(1, 65):
                            ws_mother.Cells(current_summary_row, c).Value = None
                elif current_summary_row > target_summary_row:
                    rows_to_delete = current_summary_row - target_summary_row
                    del_start = target_summary_row
                    del_end = current_summary_row - 1
                    ws_mother.Rows(f"{del_start}:{del_end}").Delete()
                    
            ws_mother.Cells(target_summary_row, 2).Formula = f"=SUM(B5:B{target_summary_row-1})"
            ws_mother.Cells(target_summary_row, 3).Formula = f"=SUM(C5:C{target_summary_row-1})"
            ws_mother.Cells(target_summary_row, 4).Formula = f"=SUM(D5:D{target_summary_row-1})"
            ws_mother.Cells(target_summary_row, 5).Formula = f"=MAX(E5:E{target_summary_row-1})"
            ws_mother.Cells(target_summary_row, 6).Formula = f"=SUM(F5:F{target_summary_row-1})"
            ws_mother.Cells(target_summary_row, 7).Formula = f"=AVERAGE(G5:G{target_summary_row-1})"
            ws_mother.Cells(target_summary_row, 8).Formula = f"=AVERAGE(H5:H{target_summary_row-1})"
            # PR Compensated monthly average — matches v10 (plain AVERAGE of day rows;
            # AVERAGE ignores blank/unprocessed days but includes any 0-value days).
            ws_mother.Cells(target_summary_row, 9).Formula = f"=AVERAGE(I5:I{target_summary_row-1})"
            ws_mother.Cells(target_summary_row, 10).Formula = f"=SUMIF(J5:J{target_summary_row-1},\"<>0\")/COUNTIF(J5:J{target_summary_row-1},\"<>0\")"
            ws_mother.Cells(target_summary_row, 11).Formula = f"=SUM(K5:K{target_summary_row-1})"
            ws_mother.Cells(target_summary_row, 12).Formula = f"=SUM(L5:L{target_summary_row-1})"
            ws_mother.Cells(target_summary_row, 13).Formula = f"=SUM(M5:M{target_summary_row-1})"
            
            # Inverter PR columns averages (columns 14 to 49)
            for inv_col in range(14, 50):
                col_let = openpyxl.utils.get_column_letter(inv_col)
                ws_mother.Cells(target_summary_row, inv_col).Formula = f"=AVERAGE({col_let}5:{col_let}{target_summary_row-1})" 
            
            # Format summary row cells as numbers (prevent date formatting)
            ws_mother.Range(f"B{target_summary_row}:D{target_summary_row}").NumberFormatLocal = "0,0000"
            ws_mother.Cells(target_summary_row, 5).NumberFormatLocal = "0,00"
            ws_mother.Range(f"K{target_summary_row}:M{target_summary_row}").NumberFormatLocal = "0,00"
            
            # Write/update formulas for External Availability in day rows (5 to 4 + num_days) after summary row has been adjusted
            for r in range(5, 5 + num_days):
                ws_mother.Cells(r, 10).Formula = f"=IF(F{r}=\"\",0,(F{r}/(F{r}+K{r}+L{r}+M{r}))*100)"
            
            # Change links natively via Excel to avoid openpyxl corruption if initialized new
            if initialized_new:
                links = wb_mother.LinkSources(1) # xlExcelLinks
                if links:
                    for link in links:
                        match = re.search(r"PR_recalculation_(\d+)_", link, re.IGNORECASE)
                        if match:
                            day_num = int(match.group(1))
                            if day_num <= num_days:
                                chk_daily_filename = f"PR_recalculation_{day_num:02d}_{month_name}.xlsx"
                                chk_daily_path = os.path.abspath(os.path.join(calcolo_folder, chk_daily_filename)).replace('/', '\\')
                                wb_mother.ChangeLink(Name=link, NewName=chk_daily_path, Type=1)

            # Ensure all day rows have correct literal dates for this month
            for r in range(5, 5 + num_days):
                day_num = r - 4
                ws_mother.Cells(r, 1).Value = f"{year_val}-{month_val:02d}-{day_num:02d}"
                
            # Dynamically map mother sheet columns based on header names (Row 4)
            header_mapping = {}
            for col in range(2, 65):
                val = ws_mother.Cells(4, col).Value
                if not val:
                    continue
                val_str = str(val).strip().lower()
                if "irradiance tx1" in val_str:
                    header_mapping[col] = "$D$111"
                elif "irradiance tx3" in val_str:
                    header_mapping[col] = "$F$111"
                elif "irradiance" in val_str or "irraggiamento" in val_str:
                    header_mapping[col] = "$I$111"
                elif "meter reading" in val_str or "lettura contatore" in val_str:
                    header_mapping[col] = "$L$110"
                elif "energy" in val_str and "loss" not in val_str and "perdita" not in val_str:
                    header_mapping[col] = "$M$111"
                elif "pr total" in val_str or "pr vcom" in val_str:
                    header_mapping[col] = "$BA$5*100"
                elif "pr scada" in val_str:
                    header_mapping[col] = "$BH$8"
                elif "pr compensated" in val_str or "pr compensato" in val_str:
                    header_mapping[col] = "$BH$11"
                elif "tx1 - energy loss" in val_str or "tx1 - perdita" in val_str or "perdita energia tx1" in val_str:
                    header_mapping[col] = "$AA$111"
                elif "tx2 - energy loss" in val_str or "tx2 - perdita" in val_str or "perdita energia tx2" in val_str:
                    header_mapping[col] = "$AN$111"
                elif "tx3 - energy loss" in val_str or "tx3 - perdita" in val_str or "perdita energia tx3" in val_str:
                    header_mapping[col] = "$BA$111"
                elif "pr tx1-inv-" in val_str:
                    try:
                        inv_num = int(val_str.split("inv-")[-1])
                        daily_col = 14 + inv_num
                        header_mapping[col] = openpyxl.utils.get_column_letter(daily_col) + "$111"
                    except Exception:
                        pass
                elif "pr tx2-inv-" in val_str:
                    try:
                        inv_num = int(val_str.split("inv-")[-1])
                        daily_col = 27 + inv_num
                        header_mapping[col] = openpyxl.utils.get_column_letter(daily_col) + "$111"
                    except Exception:
                        pass
                elif "pr tx3-inv-" in val_str:
                    try:
                        inv_num = int(val_str.split("inv-")[-1])
                        daily_col = 40 + inv_num
                        header_mapping[col] = openpyxl.utils.get_column_letter(daily_col) + "$111"
                    except Exception:
                        pass
                
            # --- Optional external vendor PR sources ---------------------------------
            # If the month folder (the '/YYYY MM' folder that contains 'PR CALCOLO FILE')
            # holds the SCADA KPI export and/or the VCOM export, their per-day PR values
            # REPLACE the formula-linked PR SCADA / PR VCOM columns in the Mother file.
            month_folder = os.path.dirname(os.path.abspath(calcolo_folder))
            scada_pr = self._read_scada_daily_pr(month_folder, year_val, month_val)
            vcom_pr = self._read_vcom_daily_pr(month_folder)
            scada_col = vcom_col = None
            for col in range(2, 65):
                hv = str(ws_mother.Cells(4, col).Value or "").strip().lower()
                if "pr scada" in hv:
                    scada_col = col
                elif "pr vcom" in hv:
                    vcom_col = col
            # When a vendor file is present, manage that column as direct values: drop its
            # child-link formula so the value we write is not overwritten on the next sync.
            if scada_pr and scada_col:
                header_mapping.pop(scada_col, None)
                print(f"PR SCADA da 'KPI_Report_Daily.xls': {len(scada_pr)} giorni -> colonna Madre {scada_col}")
            if vcom_pr and vcom_col:
                header_mapping.pop(vcom_col, None)
                print(f"PR VCOM da 'Performance_ratio_vcom.csv': {len(vcom_pr)} giorni -> colonna Madre {vcom_col}")

            daily_dir = os.path.abspath(calcolo_folder).replace('/', '\\')
            
            # Determine which days are computed from VCOM data for highlighting
            if vcom_days is not None:
                effective_vcom_days = set(vcom_days)
            else:
                effective_vcom_days = set()
                if hasattr(self, 'vcom_days_processed') and self.vcom_days_processed:
                    effective_vcom_days.update(self.vcom_days_processed)
                
                mode_src = self.data_source_var.get() if hasattr(self, 'data_source_var') else "scada"
                if mode_src == "vcom":
                    effective_vcom_days.update(range(1, num_days + 1))
                elif mode_src == "misto" and hasattr(self, '_parse_vcom_days_set'):
                    effective_vcom_days.update(self._parse_vcom_days_set())

            sync_count = 0
            for day_num in range(1, num_days + 1):
                chk_daily_filename = f"PR_recalculation_{day_num:02d}_{month_name}.xlsx"
                chk_daily_path = os.path.join(calcolo_folder, chk_daily_filename)
                r = 5 + day_num - 1

                # Check if this day is a VCOM day
                is_vcom_day = (day_num in effective_vcom_days)

                if os.path.exists(chk_daily_path):
                    prefix = f"='{daily_dir}\\[{chk_daily_filename}]PR_Calc'"

                    formulas_to_write = {}
                    for col, addr in header_mapping.items():
                        expected_formula = f"{prefix}!{addr}"
                        try:
                            curr_f = ws_mother.Cells(r, col).Formula
                        except Exception:
                            curr_f = ""
                        if curr_f != expected_formula:
                            formulas_to_write[col] = expected_formula

                    if formulas_to_write:
                        print(f"Sincronizzazione formule per il giorno {day_num}...")
                        for col, f_val in formulas_to_write.items():
                            ws_mother.Cells(r, col).Formula = f_val
                        sync_count += 1

                    # Vendor PR overrides (written as values, not formulas) when present.
                    if scada_pr and scada_col:
                        ws_mother.Cells(r, scada_col).Value = scada_pr.get(day_num)
                    if vcom_pr and vcom_col:
                        ws_mother.Cells(r, vcom_col).Value = vcom_pr.get(day_num)

                    # Highlight VCOM rows in Light Orange (RGB: 255, 224, 178 / 11722975) and add VCOM difference comment
                    try:
                        row_range = ws_mother.Range(ws_mother.Cells(r, 1), ws_mother.Cells(r, 64))
                        vcom_note = ("Nota: Giorno elaborato con dati VCOM. È presente una differenza di circa 300 kW "
                                     "(kWh/giorno) nei dati di energia giornalieri tra SCADA e VCOM.")
                        date_cell = ws_mother.Cells(r, 1)
                        if is_vcom_day:
                            row_range.Interior.Color = 11722975  # Soft Light Orange (#FFE0B2)
                            try:
                                if date_cell.Comment is not None:
                                    date_cell.Comment.Delete()
                                date_cell.AddComment(vcom_note)
                                date_cell.Comment.Visible = False
                            except Exception:
                                pass
                        else:
                            if row_range.Interior.Color == 11722975:
                                row_range.Interior.ColorIndex = -4142  # Clear orange fill if re-calculated via SCADA
                            try:
                                if date_cell.Comment is not None:
                                    date_cell.Comment.Delete()
                            except Exception:
                                pass
                    except Exception as clr_ex:
                        pass
                else:
                    # Daily file not yet processed — clear any stale data/formulas left
                    # over from a previous month's template so the row stays blank.
                    for col in header_mapping.keys():
                        try:
                            cell = ws_mother.Cells(r, col)
                            if cell.Value is not None or cell.Formula:
                                cell.Value = None
                        except Exception:
                            pass
                        
            try:
                excel.Calculation = -4105  # xlCalculationAutomatic
                excel.CalculateBeforeSave = True
                wb_mother.Calculate()
            except Exception:
                pass
                
            self._save_workbook_resilient(wb_mother, excel, abs_mother_path, "MADRE")
            try:
                wb_mother.Close(SaveChanges=False)  # already persisted by resilient save
            except Exception:
                pass
            print(f"Sincronizzazione completata! Giorni aggiornati nel file Madre: {sync_count}")
        except Exception as ex:
            print(f"Errore durante l'aggiornamento del file Madre via Excel COM: {ex}")
            if wb_mother:
                try:
                    wb_mother.Close(SaveChanges=False)
                except Exception:
                    pass
            raise RuntimeError(f"Impossibile salvare o aggiornare il file Madre '{expected_mother_filename}'.\n"
                               f"Assicurarsi che il file non sia aperto in un'altra finestra di Excel o bloccato da un altro utente.\n"
                               f"Dettaglio errore: {ex}")

    def run_calculation(self, folder, date_str, pvsyst_pr, threshold, diff_threshold=0.10, poa_method="condmax", day_filter=None):
        try:
            import datetime
            # Scan immediate subdirectories to check if this is a month folder (e.g. contains subdirectories '01', '02', '25', '26' etc.)
            subdirs = [d for d in os.listdir(folder) if os.path.isdir(os.path.join(folder, d))]
            numerical_subdirs = sorted([d for d in subdirs if d.isdigit() and 1 <= int(d) <= 31], key=lambda x: int(x))
            
            is_batch = len(numerical_subdirs) > 0
            
            # v13: restrict the batch to the days the user selected. Everything downstream
            # (missing-data pre-scan, VCOM download/conversion, the processing loop) keys off
            # numerical_subdirs, so filtering here scopes the whole run in one place.
            if is_batch and day_filter is not None:
                available = list(numerical_subdirs)
                numerical_subdirs = [d for d in available if int(d) in day_filter]
                if not numerical_subdirs:
                    wanted = ", ".join(str(d) for d in sorted(day_filter))
                    month_label = os.path.basename(folder.rstrip("\\/"))
                    raise RuntimeError(
                        f"Nessuna cartella giornaliera corrisponde all'ambito selezionato (giorni: {wanted}).\n"
                        f"Giorni disponibili in '{month_label}': {', '.join(available) or 'nessuno'}."
                    )
                print(f">>> Ambito ristretto a {len(numerical_subdirs)} giorno/i su {len(available)}: "
                      f"{', '.join(numerical_subdirs)}.")
            
            self.vcom_days_processed = set()
            if is_batch:
                # Batch Month Processing mode
                calcolo_folder = os.path.join(folder, "PR CALCOLO FILE")
                os.makedirs(calcolo_folder, exist_ok=True)
                
                # Dynamically parse Year & Month from folder name
                basename = os.path.basename(folder.rstrip("\\/"))
                parts = basename.split()
                year_val = datetime.datetime.now().year
                month_val = datetime.datetime.now().month
                if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
                    year_val = int(parts[0])
                    month_val = int(parts[1])
                    
                month_abbrs = {
                    1: "gen", 2: "feb", 3: "mar", 4: "apr", 5: "mag", 6: "giu",
                    7: "lug", 8: "ago", 9: "set", 10: "ott", 11: "nov", 12: "dic"
                }
                month_name = month_abbrs[month_val]
                
                # The scope only supplies day numbers; year/month always come from the folder
                # name. Flag the mismatch so a stale Data field can't look like a wrong result.
                if day_filter is not None and not date_str.startswith(f"{year_val:04d}-{month_val:02d}"):
                    print(f">>> ATTENZIONE: il campo Data ({date_str}) non appartiene al mese della "
                          f"cartella ({year_val:04d}-{month_val:02d}). Vengono elaborati i giorni "
                          f"{', '.join(numerical_subdirs)} di {month_val:02d}/{year_val}.")
                
                reprocess_all = self.reprocess_all_var.get()
                processed_count = 0
                skipped_count = 0
                
                last_df_result = None
                last_calc_results = None
                self.all_days_results = []

                # Pre-scan: check that every day that will actually be processed has all
                # required SCADA files. Days whose daily file already exists (and reprocess
                # is off) are skipped anyway, so they are not checked. If any day is
                # incomplete, ask the user whether to bypass it and process only the
                # complete days, rather than aborting the whole month.
                bypassed_days = {}
                vcom_convertible = {}
                vcom_downloadable = []
                mode_src = self.data_source_var.get()
                vcom_days_set = self._parse_vcom_days_set()
                auto_approved_vcom = True

                for day_str in numerical_subdirs:
                    day_val = int(day_str)
                    daily_filename = f"PR_recalculation_{day_val:02d}_{month_name}.xlsx"
                    if os.path.exists(os.path.join(calcolo_folder, daily_filename)) and not reprocess_all:
                        continue
                    
                    day_path = os.path.join(folder, day_str)
                    is_vcom_req = (mode_src == "vcom") or (mode_src == "misto" and day_val in vcom_days_set)
                    vcom_dir = self._find_vcom_folder_for_day(day_path)
                    
                    if is_vcom_req:
                        if vcom_dir:
                            vcom_convertible[day_str] = vcom_dir
                        else:
                            vcom_downloadable.append(day_str)
                            bypassed_days[day_str] = ["Dati VCOM non presenti"]
                    else:
                        miss = self._missing_files_for_day(day_path)
                        if miss:
                            bypassed_days[day_str] = miss
                            auto_approved_vcom = False
                            if vcom_dir:
                                vcom_convertible[day_str] = vcom_dir
                            else:
                                vcom_downloadable.append(day_str)

                if bypassed_days:
                    # If some missing days do not have VCOM locally, ask to download them
                    if vcom_downloadable:
                        selected_days = self._ask_select_days_on_gui(
                            "Dati mancanti - Download VCOM",
                            "I file SCADA e VCOM per i seguenti giorni sono mancanti.\n"
                            "Seleziona i giorni per cui scaricare i dati VCOM da meteocontrol:",
                            vcom_downloadable
                        )
                        if selected_days:
                            print(f"[VCOM-Downloader] Avvio download dati VCOM per {len(selected_days)} giorno/i...")
                            batch_pairs = []
                            for d_str in selected_days:
                                d_path = os.path.join(folder, d_str)
                                vcom_dir = os.path.join(d_path, "vcom")
                                target_date_str = f"{year_val:04d}-{month_val:02d}-{int(d_str):02d}"
                                batch_pairs.append((target_date_str, vcom_dir))
                            batch_results = self._download_vcom_data_batch(batch_pairs)
                            for d_str in selected_days:
                                target_date_str = f"{year_val:04d}-{month_val:02d}-{int(d_str):02d}"
                                if batch_results.get(target_date_str, False):
                                    vcom_convertible[d_str] = os.path.join(folder, d_str, "vcom")
                                else:
                                    print(f"[VCOM-Downloader] Download fallito per il Giorno {d_str}")
                        else:
                            print("[VCOM-Downloader] Download annullato dall'utente.")

                    # If we have VCOM data available (either pre-existing or just downloaded), offer automatic conversion
                    if vcom_convertible:
                        convert_list = ", ".join(sorted(vcom_convertible.keys(), key=lambda x: int(x)))
                        non_vcom = [d for d in bypassed_days if d not in vcom_convertible]
                        
                        if auto_approved_vcom and mode_src in ["vcom", "misto"]:
                            convert_approved = True
                        else:
                            msg = "I file SCADA per alcuni giorni sono mancanti/incompleti, ma i dati VCOM sono stati rilevati o scaricati:\n\n"
                            msg += f"   • Giorni con VCOM pronti: {convert_list}\n\n"
                            if non_vcom:
                                msg += f"NOTA: Per i giorni {', '.join(sorted(non_vcom, key=lambda x: int(x)))}, i dati VCOM non sono disponibili.\n\n"
                            msg += "Vuoi convertire automaticamente i dati VCOM per calcolare il PR?"
                            
                            convert_approved = self._ask_yes_no_on_gui(
                                "Rilevati dati VCOM per giorni mancanti",
                                msg
                            )
                        
                        if convert_approved:
                            from VCOM_to_SCADA import convert_vcom_to_scada
                            import shutil
                            for d_str, v_dir in vcom_convertible.items():
                                if self.stop_requested.is_set():
                                    print("[VCOM-Pre-calcolo] Arresto richiesto: conversione interrotta.")
                                    break
                                d_path = os.path.join(folder, d_str)
                                target_date_str = f"{year_val:04d}-{month_val:02d}-{int(d_str):02d}"
                                print(f"[VCOM-Pre-calcolo] Generazione dati pseudo-SCADA per il Giorno {d_str} ({target_date_str})...")
                                try:
                                    convert_vcom_to_scada(v_dir, d_path, target_date_str)
                                    
                                    # Re-evaluate missing files
                                    new_miss = self._missing_files_for_day(d_path)
                                    if not new_miss:
                                        bypassed_days.pop(d_str, None)
                                except Exception as conv_err:
                                    print(f"[VCOM-Pre-calcolo] Errore conversione per il Giorno {d_str}: {conv_err}")

                    # If there are still days with missing files
                    if bypassed_days:
                        ordered = sorted(bypassed_days, key=lambda d: int(d))
                        detail = "\n".join(f"   • Giorno {d}: mancano {', '.join(bypassed_days[d])}" for d in ordered)
                        
                        no_vcom_days = [d for d in ordered if d not in vcom_convertible]
                        guide_hint = ""
                        if no_vcom_days:
                            guide_hint = f"\n\nPer i giorni {', '.join(sorted(no_vcom_days, key=lambda x: int(x)))}, puoi scaricare i dati VCOM inserendoli nella cartella 'vcom/' del relativo giorno."
                            
                        proceed = self._ask_yes_no_on_gui(
                            "Elaborazione Giorni Disponibili",
                            f"I giorni con dati completi (SCADA o VCOM) verranno elaborati e sincronizzati sul file Madre.\n\n"
                            f"I seguenti giorni non hanno file SCADA/VCOM e verranno saltati:\n{detail}{guide_hint}\n\n"
                            "Vuoi procedere con il calcolo dei giorni disponibili?\n\n"
                            "[Sì] = Elabora i giorni disponibili e aggiorna il file Madre\n"
                            "[No] = Annulla l'elaborazione"
                        )
                        if not proceed:
                            raise RuntimeError(
                                "Elaborazione annullata dall'utente: "
                                f"{len(bypassed_days)} giorno/i senza tutti i file richiesti ({', '.join(ordered)})."
                            )
                        for d in ordered:
                            print(f"[Batch] Giorno {d} SALTATO per dati mancanti: {', '.join(bypassed_days[d])}.")

                bypassed_count = 0
                stopped_early = False
                for idx, day_str in enumerate(numerical_subdirs):
                    # Safe-stop checkpoint: the previous day is fully written at this point,
                    # so breaking here leaves the output set consistent.
                    if self.stop_requested.is_set():
                        stopped_early = True
                        print(f"[Batch] Arresto richiesto: interrotto dopo {processed_count} giorno/i elaborato/i.")
                        break
                    day_val = int(day_str)
                    target_date_str = f"{year_val:04d}-{month_val:02d}-{day_val:02d}"
                    daily_filename = f"PR_recalculation_{day_val:02d}_{month_name}.xlsx"
                    daily_file_path = os.path.join(calcolo_folder, daily_filename)

                    # Skip days the user chose to bypass (incomplete data).
                    if day_str in bypassed_days:
                        bypassed_count += 1
                        continue

                    # If daily file already exists and we are not forcing reprocess, skip this day folder!
                    if os.path.exists(daily_file_path) and not reprocess_all:
                        skipped_count += 1
                        continue
                        
                    # Update status securely from the non-GUI thread
                    status_text = f"Modalità Batch: Elaborazione giorno {day_val} di {len(numerical_subdirs)}..."
                    self.root.after(0, lambda t=status_text: self.lbl_status.config(text=t, foreground=self.warn_color))
                    self.root.after(0, lambda d=idx, t=len(numerical_subdirs): self._set_progress(d, t))
                    
                    day_folder = os.path.join(folder, day_str)
                    # Determine whether this day uses VCOM or SCADA
                    mode_src = self.data_source_var.get()
                    vcom_days_set = self._parse_vcom_days_set()
                    use_vcom = False
                    if mode_src == "vcom":
                        use_vcom = True
                    elif mode_src == "misto":
                        if day_val in vcom_days_set:
                            use_vcom = True
                    
                    # Auto-fallback: if day folder has vcom/ subfolder and SCADA files are missing, use VCOM
                    v_dir = self._find_vcom_folder_for_day(day_folder)
                    if not use_vcom and v_dir and self._missing_files_for_day(day_folder):
                        use_vcom = True

                    src_label = "SCADA"
                    if use_vcom:
                        if not v_dir:
                            v_dir = self._find_vcom_folder_for_day(day_folder)
                        if v_dir:
                            print(f"[Batch] Giorno {day_str} ({target_date_str}): Generazione dati da VCOM ({os.path.basename(v_dir)})...")
                            from VCOM_to_SCADA import convert_vcom_to_scada
                            convert_vcom_to_scada(v_dir, day_folder, target_date_str)
                            src_label = "VCOM"
                            self.vcom_days_processed.add(int(day_str))
                        else:
                            raise FileNotFoundError(f"Giorno {day_str}: configurato per VCOM ma nessun file VCOM trovato in '{day_folder}' o 'vcom/'")

                    try:
                        last_df_result, last_calc_results = self.calculate_single_day(
                            day_folder, target_date_str, pvsyst_pr, threshold, diff_threshold, calcolo_folder, skip_mother_update=True, poa_method=poa_method
                        )
                        last_calc_results['data_source'] = src_label
                        self.all_days_results.append(last_calc_results)
                        processed_count += 1
                        comp_pr_val = last_calc_results.get('comp_raw_pr', 0.0)
                        print(f"[Batch] Giorno {day_str} ({target_date_str}): Elaborato con sorgente [{src_label}] -> PR Compensato = {comp_pr_val:.2f}%")
                    except Exception as day_err:
                        raise ValueError(f"Errore nel Giorno {day_str} (sorgente {src_label}): {day_err}")
                        
                # Sincronizza il file Madre alla fine del batch (aggiorna anche i giorni saltati).
                # Disattivabile dalle Opzioni Avanzate: è il passo più lento del calcolo e per
                # un ricalcolo di pochi giorni può essere rimandato a fine mese.
                if self.sync_mother_var.get():
                    status_text = "Sincronizzazione finale del file Madre..."
                    self.root.after(0, lambda t=status_text: self.lbl_status.config(text=t, foreground=self.warn_color))
                    self.sync_mother_file(calcolo_folder, year_val, month_val)
                    mother_txt = " Sincronizzato file Madre."
                else:
                    print(">>> Sincronizzazione del file Madre disattivata nelle Opzioni Avanzate: saltata.")
                    mother_txt = " File Madre NON sincronizzato (opzione disattivata)."
                
                if processed_count == 0:
                    bypass_txt = f" {bypassed_count} saltati per dati mancanti." if bypassed_count else ""
                    stop_txt = " Elaborazione interrotta dall'utente." if stopped_early else ""
                    status_msg = f"Nessun nuovo giorno elaborato ({skipped_count} già presenti).{bypass_txt}{stop_txt}{mother_txt}"
                    self.root.after(0, lambda m=status_msg: self.lbl_status.config(text=m, foreground=self.success_color))
                    self.root.after(0, self._reset_run_buttons)
                    return
                    
                # Display results of the last processed day on GUI dashboard
                self.df_result = last_df_result
                self.calc_results = last_calc_results
                
                bypass_txt = f", {bypassed_count} saltati per dati mancanti" if bypassed_count else ""
                if stopped_early:
                    success_msg = (f"Elaborazione interrotta dall'utente. Completati: {processed_count} giorni"
                                   f"{bypass_txt}.{mother_txt}")
                else:
                    success_msg = f"Calcolo completato! Elaborati: {processed_count} giorni{bypass_txt}.{mother_txt}"
                self.root.after(0, lambda m=success_msg: self.lbl_status.config(text=m, foreground=self.success_color))
                self.root.after(0, self.update_ui_on_success)
                
            else:
                # Single Day Processing mode
                self.root.after(0, lambda: self.lbl_status.config(text="Calcolo giorno singolo in corso...", foreground=self.warn_color))
                
                # Determine if Single Day should use VCOM
                mode_src = self.data_source_var.get()
                vcom_days_set = self._parse_vcom_days_set()
                day_num = int(date_str.split("-")[2]) if "-" in date_str else 1
                use_vcom = (mode_src == "vcom") or (mode_src == "misto" and day_num in vcom_days_set)
                
                vcom_dir = self._find_vcom_folder_for_day(folder)
                if not use_vcom and (self._folder_contains_vcom_csvs(folder) or (vcom_dir and self._missing_files_for_day(folder))):
                    use_vcom = True

                if use_vcom:
                    if not vcom_dir:
                        vcom_dir = self._find_vcom_folder_for_day(folder)
                    if not vcom_dir:
                        # VCOM is not present locally, ask to download it!
                        dl_approved = self._ask_yes_no_on_gui(
                            "Dati VCOM mancanti",
                            f"I file VCOM per il giorno selezionato ({date_str}) non sono stati trovati in '{folder}'.\n\n"
                            "Vuoi avviare il download automatico dei dati VCOM da meteocontrol?"
                        )
                        if dl_approved:
                            vcom_dir = os.path.join(folder, "vcom")
                            print(f"[VCOM-Downloader] Download in corso per il Giorno Singolo ({date_str})...")
                            success = self._download_vcom_data(date_str, vcom_dir)
                            if not success:
                                raise RuntimeError(f"Download dati VCOM fallito per la data {date_str}.")
                        else:
                            raise RuntimeError(f"Calcolo annullato. Dati VCOM mancanti per la data {date_str}.")
                    
                    if vcom_dir:
                        print(f"[Giorno Singolo] ({date_str}): Generazione dati pseudo-SCADA da VCOM ({vcom_dir})...")
                        from VCOM_to_SCADA import convert_vcom_to_scada
                        convert_vcom_to_scada(vcom_dir, folder, date_str)
                        if not hasattr(self, "vcom_days_processed"): self.vcom_days_processed = set()
                        self.vcom_days_processed.add(day_num)
                else:
                    # Check for missing SCADA files
                    miss = self._missing_files_for_day(folder)
                    if miss:
                        vcom_dir = self._find_vcom_folder_for_day(folder)
                        if vcom_dir:
                            convert_approved = self._ask_yes_no_on_gui(
                                "Dati SCADA mancanti - VCOM Rilevato",
                                f"I file SCADA per il giorno selezionato sono incompleti, ma i dati VCOM sono stati rilevati nella cartella:\n\n{vcom_dir}\n\nVuoi convertire automaticamente i dati VCOM per questo giorno e procedere con il calcolo?"
                            )
                            if convert_approved:
                                print(f"[VCOM-Pre-calcolo] Generazione dati pseudo-SCADA per il Giorno Singolo ({date_str})...")
                                from VCOM_to_SCADA import convert_vcom_to_scada
                                convert_vcom_to_scada(vcom_dir, folder, date_str)
                            else:
                                raise RuntimeError(f"Calcolo annullato. File SCADA mancanti: {', '.join(miss)}")
                        else:
                            raise RuntimeError(f"Calcolo annullato. File SCADA mancanti: {', '.join(miss)}")

                df_res, calc_res = self.calculate_single_day(folder, date_str, pvsyst_pr, threshold, diff_threshold, poa_method=poa_method)
                
                self.df_result = df_res
                self.calc_results = calc_res
                self.all_days_results = [calc_res]
                
                mother_txt = "aggiornato file Madre" if self.sync_mother_var.get() else "file Madre non sincronizzato"
                success_msg = f"Calcolo completato! Creato file giornaliero & {mother_txt}."
                self.root.after(0, lambda m=success_msg: self.lbl_status.config(text=m, foreground=self.success_color))
                self.root.after(0, self.update_ui_on_success)
                
        except Exception as e:
            err_msg = str(e)
            self.root.after(0, lambda: self.update_ui_on_failure(err_msg))
        finally:
            quit_excel_app()
            
    def _reset_run_buttons(self):
        """Return the run controls to idle: Calcola enabled, Interrompi disabled, bar hidden."""
        try:
            self.btn_calculate.config(state="normal")
            self.btn_stop.config(state="disabled")
            self.progress.pack_forget()
        except Exception:
            pass

    def _set_progress(self, done, total):
        """Show the batch progress bar and move it. Call via root.after from the worker."""
        try:
            if not self.progress.winfo_ismapped():
                self.progress.pack(fill="x", pady=(2, 4), before=self.lbl_status)
            self.progress.config(value=(100.0 * done / total) if total else 0)
        except Exception:
            pass

    def update_ui_on_success(self):
        res = self.calc_results
        
        # Enable controls
        self._reset_run_buttons()
        self.btn_export.config(state="normal")
        
        # Display main metrics
        self.lbl_avg_pr_val.config(text=f"{res['avg_inv_pr']:.3f} %".replace(".", ","))
        self.lbl_comp_pr_val.config(text=f"{res['comp_raw_pr']:.3f} %".replace(".", ","))
        try:
            val = float(self.pvsyst_pr_var.get().replace(",", "."))
            self.lbl_pvsyst_target_val.config(text=f"{val * 100:.3f} %".replace(".", ","))
        except ValueError:
            self.lbl_pvsyst_target_val.config(text="-- %")
        
        self.lbl_irrad_summary.config(
            text=f"Irradiazione giornaliera totale: {res['h_sum_kwh']:.4f} kWh/m² ({'Conditional MAX' if self.poa_method_var.get()=='condmax' else 'Media'} POA >= {self.threshold_var.get()} W/m²)".replace(".", ",")
        )
        
        # Fill treeview (Inverters)
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        for row in res['inverter_table_data']:
            self.tree.insert("", "end", values=row)
            
        # Fill treeview (Days)
        for item in self.tree_days.get_children():
            self.tree_days.delete(item)
            
        for day_res in getattr(self, "all_days_results", []):
            day_row = (
                day_res['date_str'],
                f"{day_res['h_sum_kwh']:.4f}",
                f"{day_res['uncomp_pr']:.3f}",
                f"{day_res['comp_raw_pr']:.3f}",
                f"{day_res['avg_inv_pr']:.3f}"
            )
            self.tree_days.insert("", "end", values=day_row)
            
        # Select appropriate tab
        if hasattr(self, "all_days_results") and len(self.all_days_results) > 1:
            self.notebook.select(1) # Go to Daily Summary tab
        else:
            self.notebook.select(0) # Go to Detailed Inverters tab
            
        self.lbl_status.config(text="Calcolo completato con successo!", foreground=self.success_color)
        
        # Build detailed success report showing PR of each day processed
        msg = f"Calcolo del PR per il {res['date_str']} terminato con successo!\n\n"
        if hasattr(self, "all_days_results") and len(self.all_days_results) > 1:
            msg += "Riepilogo PR Giornaliero (Media PR Inverter):\n"
            for d in self.all_days_results:
                msg += f"- {d['date_str']}: {d['avg_inv_pr']:.3f}%\n"
        else:
            msg += f"Media PR Inverter: {res['avg_inv_pr']:.3f}%"
            
        messagebox.showinfo("Successo", msg)
        
    def update_ui_on_failure(self, error_message):
        self._reset_run_buttons()
        self.btn_export.config(state="disabled")
        self.lbl_status.config(text="Calcolo fallito!", foreground="red")
        messagebox.showerror("Errore di Calcolo", error_message)
        
    def export_to_excel(self):
        if self.df_result is None or self.calc_results is None:
            messagebox.showerror("Errore", "Nessun risultato disponibile da esportare!")
            return
            
        # Ask where to save the file
        default_name = f"PR_recalculation_{self.calc_results['date_str'].replace('-', '_')}.xlsx"
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("File Excel", "*.xlsx")],
            initialfile=default_name,
            title="Salva Report PR"
        )
        
        if not save_path:
            return
            
        try:
            self.lbl_export_status.config(text="Esportazione in corso... attendere...", foreground=self.warn_color)
            self.btn_export.config(state="disabled")
            
            # Run export in background thread
            thread = threading.Thread(target=self.run_export, args=(save_path,))
            thread.start()
        except Exception as e:
            self.lbl_export_status.config(text="Esportazione fallita!", foreground="red")
            messagebox.showerror("Errore Esportazione", str(e))
            self.btn_export.config(state="normal")
            
    def run_export(self, save_path):
        try:
            res = self.calc_results
            df_result = self.df_result
            
            # We will export 3 sheets
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                # 1. Summary Sheet
                summary_data = {
                    "Parametro": [
                        "Data di Calcolo",
                        "Irradiazione Giornaliera Totale (kWh/m²)",
                        "Target PR PVSyst Mensile",
                        "Soglia Irraggiamento Minimo (W/m²)",
                        "Tolleranza Differenza Irraggiamento (%)",
                        "PR Grezzo Non Compensato (%)",
                        "PR Grezzo Compensato (%)",
                        "Media dei 36 PR Inverter Compensati (%)"
                    ],
                    "Valore": [
                        res["date_str"],
                        res["h_sum_kwh"],
                        float(self.pvsyst_pr_var.get().replace(",", ".")),
                        float(self.threshold_var.get().replace(",", ".")),
                        float(self.diff_threshold_var.get().replace(",", ".")),
                        res["uncomp_pr"],
                        res["comp_raw_pr"],
                        res["avg_inv_pr"]
                    ]
                }
                df_sum = pd.DataFrame(summary_data)
                df_sum.to_excel(writer, sheet_name="Riepilogo", index=False)
                
                # 2. Inverter PRs Sheet
                inverter_cols = ["Codice Inverter", "Trasformatore", "Potenza CC Nominale (kW)", "Energia Prodotta (kWh)", "Perdita Stimata (kWh)", "PR Compensato (%)"]
                df_inv_prs = pd.DataFrame(res["inverter_table_data"], columns=inverter_cols)
                df_inv_prs.to_excel(writer, sheet_name="PR_Inverter", index=False)
                
                # 3. Complete Timeslot Data
                cols_to_export = [
                    "time", "poa1", "poa3", "poa1_kwh", "poa3_kwh", "poa_avg_kwh", "poa_avg_w", "h", 
                    "diff_pct", "poa_cond_max_kwh", "limit_ratio", "meter_reading", "active_energy_prod",
                    "TX1_Average_Power", "TX2_Average_Power", "TX3_Average_Power",
                    "TX1_Total_Loss", "TX2_Total_Loss", "TX3_Total_Loss"
                ]
                for inv_id in self.dc_powers:
                    cols_to_export.append(inv_id)
                    cols_to_export.append(f"{inv_id}_loss")
                    
                df_ts = df_result[cols_to_export].copy()
                df_ts.to_excel(writer, sheet_name="Dettaglio_Quarti_Ora", index=False)
                
            # Success UI trigger
            self.root.after(0, lambda: self.on_export_success(save_path))
        except Exception as e:
            self.root.after(0, lambda: self.on_export_failure(str(e)))
            
    def on_export_success(self, save_path):
        self.btn_export.config(state="normal")
        self.lbl_export_status.config(text="Esportazione completata con successo!", foreground=self.success_color)
        messagebox.showinfo("Esportazione Riuscita", f"Report PR dettagliato salvato in:\n\n{save_path}")
        
    def on_export_failure(self, error_message):
        self.btn_export.config(state="normal")
        self.lbl_export_status.config(text="Esportazione fallita!", foreground="red")
        messagebox.showerror("Errore Esportazione", f"Impossibile salvare il file Excel:\n\n{error_message}")

    def on_close(self):
        save_settings(self._collect_settings())
        sys.stdout = sys.__stdout__
        sys.stderr = sys.__stderr__
        self.root.destroy()


if __name__ == "__main__":
    # Enable Windows DPI awareness so fonts and widgets render at the native
    # display scale instead of being bitmap-stretched (fixes "shrunken text").
    try:
        import ctypes
        ctypes.windll.shcore.SetProcessDpiAwareness(1)   # PROCESS_SYSTEM_DPI_AWARE
    except Exception:
        pass
    root = tk.Tk()
    app = PRCalculatorGUI(root)
    root.mainloop()
