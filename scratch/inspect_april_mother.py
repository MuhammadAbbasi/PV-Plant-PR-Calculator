import os, glob
import openpyxl

target_dir = r"\\S01\get\2025.01 Mazara 01 A2A\03 - REPORT\Report\01 Daily Reports\2026 04\PR CALCOLO FILE"
mothers = glob.glob(os.path.join(target_dir, "00 PR_recalculation_*.xlsx"))

if not mothers:
    print(f"Nessun file Madre trovato in {target_dir}")
else:
    mother_file = mothers[0]
    print(f"Analisi file Madre: {mother_file}\n")
    
    wb = openpyxl.load_workbook(mother_file, data_only=False)
    ws = wb['PR_Calc']
    
    wb_val = openpyxl.load_workbook(mother_file, data_only=True)
    ws_val = wb_val['PR_Calc']
    
    # Print headers of some key columns
    print("Riga | Data | PR Tot (B) Form/Val | Inverter 1 (I) Form/Val | PR SCADA (E) Val")
    print("-" * 80)
    for r in range(5, 36):
        data_val = ws_val.cell(row=r, column=1).value
        pr_tot_f = ws.cell(row=r, column=2).value
        pr_tot_v = ws_val.cell(row=r, column=2).value
        inv1_f = ws.cell(row=r, column=9).value
        inv1_v = ws_val.cell(row=r, column=9).value
        pr_scada_v = ws_val.cell(row=r, column=5).value
        
        # format formula snippet
        pr_tot_f_short = str(pr_tot_f)[:30] if pr_tot_f else "None"
        inv1_f_short = str(inv1_f)[:30] if inv1_f else "None"
        
        print(f"R{r:02d} | {str(data_val)[:10]} | {pr_tot_f_short} ({pr_tot_v}) | {inv1_f_short} ({inv1_v}) | {pr_scada_v}")
