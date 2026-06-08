import os
import glob
import sys
import threading
import datetime
import pandas as pd
# pyrefly: ignore [missing-import]
import numpy as np
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
_excel_app = None

def get_resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

def get_excel_app():
    global _excel_app
    if _excel_app is None:
        import win32com.client
        _excel_app = win32com.client.DispatchEx("Excel.Application")
        _excel_app.Visible = False
        _excel_app.DisplayAlerts = False
        try:
            _excel_app.Calculation = -4135  # xlCalculationManual
            _excel_app.CalculateBeforeSave = False
        except Exception:
            pass
    return _excel_app

def quit_excel_app():
    global _excel_app
    if _excel_app is not None:
        try:
            _excel_app.Quit()
        except Exception:
            pass
        _excel_app = None

class RedirectText:
    def __init__(self, root, text_widget, original_stream):
        self.root = root
        self.text_widget = text_widget
        self.original_stream = original_stream

    def write(self, string):
        try:
            self.original_stream.write(string)
        except Exception:
            pass
        try:
            self.root.after(0, self._insert_text, string)
        except Exception:
            pass

    def flush(self):
        try:
            self.original_stream.flush()
        except Exception:
            pass

    def _insert_text(self, string):
        try:
            self.text_widget.config(state="normal")
            self.text_widget.insert("end", string)
            self.text_widget.see("end")
            self.text_widget.config(state="disabled")
        except Exception:
            pass

class PRCalculatorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Calcolatore Performance Ratio (PR) Fotovoltaico Mazara 01 - v9.0")
        
        # Center the window on the screen for a soft, user-oriented launch
        self.root.withdraw()
        self.root.update_idletasks()
        width = 1040
        height = 920
        x = (self.root.winfo_screenwidth() - width) // 2
        y = (self.root.winfo_screenheight() - height) // 2
        self.root.geometry(f"{width}x{height}+{x}+{y}")
        self.root.deiconify()
        self.root.configure(bg="#f8f9fa")
        
        # Ensure Windows taskbar and task manager correctly display the custom GET logo icon
        try:
            import ctypes
            myappid = 'get.srl.prcalculator.v9'
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
        except Exception:
            pass
            
        icon_path = get_resource_path(os.path.join("assets", "logo.ico"))
        if os.path.exists(icon_path):
            try:
                self.root.iconbitmap(icon_path)
            except Exception:
                pass
                
        logo_png_path = get_resource_path(os.path.join("assets", "logo.png"))
        if os.path.exists(logo_png_path):
            try:
                img_icon = tk.PhotoImage(file=logo_png_path)
                self.root.iconphoto(True, img_icon)
            except Exception:
                pass
        
        # Apply style
        self.style = ttk.Style()
        self.style.theme_use("clam")
        
        # Define Color Palette (Google Styled Clean Light Theme)
        self.bg_color = "#f8f9fa"      # Clean Light Gray background
        self.card_bg = "#ffffff"       # Pure White Surface
        self.accent_color = "#1a73e8"  # Google Blue Primary Accent
        self.accent_hover = "#1557b0"  # Darker Blue for hovers
        self.text_color = "#202124"    # Dark Charcoal for maximum readability
        self.muted_text = "#5f6368"    # Soft Slate Grey
        self.success_color = "#137333" # Google Green
        self.warn_color = "#b06000"    # Google Amber/Orange for secondary/uncompensated
        self.border_color = "#dadce0"  # Google Light Gray border
        
        # Custom Widget Configurations
        self.style.configure("TFrame", background=self.bg_color)
        self.style.configure("Card.TFrame", background=self.card_bg, relief="flat")
        self.style.configure("TLabel", background=self.bg_color, foreground=self.text_color, font=("Segoe UI", 10))
        self.style.configure("Card.TLabel", background=self.card_bg, foreground=self.text_color, font=("Segoe UI", 10))
        self.style.configure("Title.TLabel", background=self.bg_color, foreground=self.text_color, font=("Segoe UI Semibold", 16, "bold"))
        self.style.configure("Section.TLabel", background=self.card_bg, foreground=self.accent_color, font=("Segoe UI Semibold", 12, "bold"))
        self.style.configure("MetricVal.TLabel", background=self.card_bg, foreground=self.success_color, font=("Segoe UI", 22, "bold"))
        self.style.configure("MetricLbl.TLabel", background=self.card_bg, foreground=self.muted_text, font=("Segoe UI Semibold", 9))
        
        # Entry Field styling
        self.style.configure("TEntry", fieldbackground="#ffffff", background="#ffffff", foreground=self.text_color, borderwidth=1, bordercolor="#dadce0", padding=6)
        self.style.map("TEntry", 
                       bordercolor=[("focus", self.accent_color), ("active", self.accent_color)],
                       lightcolor=[("focus", self.accent_color)],
                       darkcolor=[("focus", self.accent_color)])
        
        # Button styling
        self.style.configure("TButton", background=self.accent_color, foreground="#ffffff", borderwidth=0, font=("Segoe UI Semibold", 10, "bold"), padding=[16, 8])
        self.style.map("TButton", background=[("active", self.accent_hover), ("disabled", "#f1f3f4")], foreground=[("disabled", "#94a3b8")])
        
        self.style.configure("Secondary.TButton", background="#ffffff", foreground=self.accent_color, bordercolor="#dadce0", darkcolor="#dadce0", lightcolor="#dadce0", borderwidth=1, font=("Segoe UI Semibold", 10), padding=[12, 6])
        self.style.map("Secondary.TButton", background=[("active", "#f8f9fa"), ("disabled", "#ffffff")], foreground=[("disabled", "#94a3b8")], bordercolor=[("active", self.accent_color)])
        
        self.style.configure("Action.TButton", background=self.accent_color, foreground="#ffffff", font=("Segoe UI Semibold", 11, "bold"), padding=[20, 10])
        self.style.map("Action.TButton", background=[("active", self.accent_hover)])
        
        # Checkbutton styling
        self.style.configure("TCheckbutton", background=self.card_bg, foreground=self.text_color, font=("Segoe UI", 10))
        self.style.map("TCheckbutton", background=[("active", self.card_bg)], foreground=[("active", self.text_color)])
        
        # Treeview styling (for the inverter list)
        self.style.configure("Treeview", background="#ffffff", foreground=self.text_color, fieldbackground="#ffffff", rowheight=28, borderwidth=0, font=("Segoe UI", 10))
        self.style.configure("Treeview.Heading", background="#f8f9fa", foreground=self.text_color, font=("Segoe UI Semibold", 10), borderwidth=1, bordercolor="#dadce0")
        self.style.map("Treeview", background=[("selected", "#e8f0fe")], foreground=[("selected", self.accent_color)])
        self.style.map("Treeview.Heading", background=[("active", "#e8f0fe")], foreground=[("active", self.accent_color)])
        
        # Notebook styling
        self.style.configure("TNotebook", background=self.bg_color, borderwidth=0)
        self.style.configure("TNotebook.Tab", background="#f1f3f4", foreground=self.muted_text, font=("Segoe UI Semibold", 10), padding=[20, 8, 20, 8], borderwidth=0)
        self.style.map("TNotebook.Tab", background=[("selected", "#ffffff"), ("active", "#e8f0fe")], foreground=[("selected", self.accent_color), ("active", self.accent_color)])
        
        # Inverter Nominal capacities
        self.dc_powers = {
            # TX1
            "TX1-INV-1": 343.75, "TX1-INV-2": 343.75, "TX1-INV-3": 343.75, "TX1-INV-4": 343.75,
            "TX1-INV-5": 343.75, "TX1-INV-6": 343.75, "TX1-INV-7": 343.75, "TX1-INV-8": 359.375,
            "TX1-INV-9": 359.375, "TX1-INV-10": 359.375, "TX1-INV-11": 359.375, "TX1-INV-12": 343.75,
            # TX2
            "TX2-INV-1": 328.125, "TX2-INV-2": 343.75, "TX2-INV-3": 343.75, "TX2-INV-4": 328.125,
            "TX2-INV-5": 328.125, "TX2-INV-6": 359.375, "TX2-INV-7": 343.75, "TX2-INV-8": 359.375,
            "TX2-INV-9": 359.375, "TX2-INV-10": 343.75, "TX2-INV-11": 359.375, "TX2-INV-12": 359.375,
            # TX3
            "TX3-INV-1": 359.375, "TX3-INV-2": 359.375, "TX3-INV-3": 359.375, "TX3-INV-4": 359.375,
            "TX3-INV-5": 359.375, "TX3-INV-6": 359.375, "TX3-INV-7": 359.375, "TX3-INV-8": 359.375,
            "TX3-INV-9": 359.375, "TX3-INV-10": 359.375, "TX3-INV-11": 359.375, "TX3-INV-12": 328.125
        }
        self.ac_power_all = 320.0
        
        # State variables
        self.folder_path_var = tk.StringVar()
        
        # Determine current local date and corresponding PVSyst PR default value
        today = datetime.date.today()
        current_date_str = today.strftime("%Y-%m-%d")
        current_month = today.month
        pvsyst_defaults = {
            1: 0.904, 2: 0.896, 3: 0.897, 4: 0.868, 5: 0.832, 6: 0.833,
            7: 0.820, 8: 0.828, 9: 0.852, 10: 0.876, 11: 0.894, 12: 0.900
        }
        default_pr = f"{pvsyst_defaults.get(current_month, 0.868):.3f}"
        
        self.date_var = tk.StringVar(value=current_date_str)
        self.pvsyst_pr_var = tk.StringVar(value=default_pr)
        self.threshold_var = tk.StringVar(value="50")
        self.reprocess_all_var = tk.BooleanVar(value=False)
        
        # Register a trace on date_var to auto-update the PVSyst PR default value
        self.date_var.trace_add("write", self.on_date_changed)
        # Register a trace on pvsyst_pr_var to update the metric card in real time
        self.pvsyst_pr_var.trace_add("write", self.on_pvsyst_pr_changed)
        
        # Placeholders for results
        self.calc_results = None
        self.df_result = None
        self.all_days_results = []
        
        # Setup GUI elements
        self.create_layout()
        
        # Select active month row in PVSyst reference treeview on startup
        try:
            self.pvsyst_tree.selection_set(f"m{current_month}")
            self.pvsyst_tree.focus(f"m{current_month}")
            self.pvsyst_tree.see(f"m{current_month}")
        except Exception:
            pass
        
        # Redirect stdout/stderr to Live Log Widget
        self.stdout_redirector = RedirectText(self.root, self.log_widget, sys.stdout)
        self.stderr_redirector = RedirectText(self.root, self.log_widget, sys.stderr)
        sys.stdout = self.stdout_redirector
        sys.stderr = self.stderr_redirector
        
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        
        print(">>> GET SRL - Motore di calcolo Performance Ratio inizializzato.")
        print(">>> Pronto per l'analisi del giorno singolo o della cartella mensile batch.")
        
    def create_card(self, parent, padding=16):
        # Border container: light grey outline
        border_card = tk.Frame(parent, bg="#dadce0", bd=0, highlightthickness=0)
        # White surface
        inner_card = tk.Frame(border_card, bg="#ffffff", bd=0, highlightthickness=0)
        inner_card.pack(fill="both", expand=True, padx=1, pady=1)
        
        # Internal content frame with padding
        content_frame = tk.Frame(inner_card, bg="#ffffff", bd=0, highlightthickness=0)
        content_frame.pack(fill="both", expand=True, padx=padding, pady=padding)
        return border_card, content_frame

    def create_layout(self):
        # Top Header Bar spanning full width
        header_bar = tk.Frame(self.root, bg="#ffffff", height=60, bd=0, highlightthickness=0)
        header_bar.pack(side="top", fill="x")
        header_bar.pack_propagate(False)
        
        # Subtle bottom border to the header bar
        bottom_border = tk.Frame(self.root, bg="#dadce0", height=1)
        bottom_border.pack(side="top", fill="x")
        
        # Header contents
        header_inner = tk.Frame(header_bar, bg="#ffffff")
        header_inner.pack(fill="both", expand=True, padx=24)
        
        logo_path = get_resource_path(os.path.join("assets", "logo.png"))
        if os.path.exists(logo_path):
            try:
                from PIL import Image, ImageTk
                img = Image.open(logo_path)
                img = img.resize((120, 38), Image.Resampling.LANCZOS)
                self.logo_photo = ImageTk.PhotoImage(img)
                lbl_logo = tk.Label(header_inner, image=self.logo_photo, bg="#ffffff", bd=0)
                lbl_logo.pack(side="left", pady=10)
            except Exception:
                pass
                
        # Vertical divider line in header
        divider = tk.Frame(header_inner, bg="#dadce0", width=1, height=24)
        divider.pack(side="left", padx=16, pady=18)
        
        title_lbl = tk.Label(header_inner, text="Motore di Calcolo Performance Ratio - Mazara 01", bg="#ffffff", fg="#1f2124", font=("Segoe UI Semibold", 13))
        title_lbl.pack(side="left", pady=10)
        
        # Version badge
        version_badge = tk.Label(header_inner, text="v9.0", bg="#e8f0fe", fg="#1a73e8", font=("Segoe UI Semibold", 9), padx=8, pady=2)
        version_badge.pack(side="left", padx=12)
        
        # Guide button on the top right with a ? logo
        btn_guide = ttk.Button(header_inner, text="? Guida d'Uso", style="Secondary.TButton", command=self.show_guide)
        btn_guide.pack(side="right", pady=10)
        
        # Main container for body (with nice margins/padding)
        main_frame = tk.Frame(self.root, bg="#f8f9fa")
        main_frame.pack(side="top", fill="both", expand=True, padx=24, pady=20)
        
        # Top Grid: Inputs (Left) and Metrics (Right)
        top_grid = tk.Frame(main_frame, bg=self.bg_color)
        top_grid.pack(fill="x", pady=(0, 15))
        top_grid.columnconfigure(0, weight=4, minsize=400)
        top_grid.columnconfigure(1, weight=5, minsize=450)
        
        # 1. Inputs Frame (Card style using helper)
        inputs_card_border, inputs_card = self.create_card(top_grid, padding=15)
        inputs_card_border.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        
        lbl_sec_in = tk.Label(inputs_card, text="Impostazioni di Calcolo", bg="#ffffff", fg=self.accent_color, font=("Segoe UI Semibold", 12, "bold"))
        lbl_sec_in.pack(anchor="w", pady=(0, 4))
        
        # Add descriptive help text for inputs card
        lbl_desc_in = tk.Label(inputs_card, text="Seleziona la cartella contenente i dati SCADA e configura i parametri.", bg="#ffffff", fg=self.muted_text, font=("Segoe UI", 10))
        lbl_desc_in.pack(anchor="w", pady=(0, 10))
        
        # Folder row
        folder_frame = tk.Frame(inputs_card, bg="#ffffff")
        folder_frame.pack(fill="x", pady=5)
        lbl_f = tk.Label(folder_frame, text="Cartella File SCADA (Input):", bg="#ffffff", fg=self.text_color, font=("Segoe UI Semibold", 9))
        lbl_f.pack(anchor="w")
        
        folder_entry_frame = tk.Frame(folder_frame, bg="#ffffff")
        folder_entry_frame.pack(fill="x", pady=2)
        
        self.entry_folder = ttk.Entry(folder_entry_frame, textvariable=self.folder_path_var, font=("Segoe UI", 10))
        self.entry_folder.pack(side="left", fill="x", expand=True, padx=(0, 8))
        
        btn_browse = ttk.Button(folder_entry_frame, text="Sfoglia...", style="Secondary.TButton", command=self.browse_folder)
        btn_browse.pack(side="right")
        
        # Parameters Grid (Date, PVSyst PR, Threshold)
        params_grid = tk.Frame(inputs_card, bg="#ffffff")
        params_grid.pack(fill="x", pady=10)
        params_grid.columnconfigure(0, weight=1)
        params_grid.columnconfigure(1, weight=1)
        params_grid.columnconfigure(2, weight=1)
        
        # Date Input
        date_frame = tk.Frame(params_grid, bg="#ffffff")
        date_frame.grid(row=0, column=0, padx=(0, 5), sticky="ew")
        tk.Label(date_frame, text="Data (AAAA-MM-GG):", bg="#ffffff", fg=self.text_color, font=("Segoe UI Semibold", 9)).pack(anchor="w")
        self.entry_date = ttk.Entry(date_frame, textvariable=self.date_var, width=12, font=("Segoe UI", 10))
        self.entry_date.pack(anchor="w", pady=2)
        
        # PVSyst PR Input
        pr_frame = tk.Frame(params_grid, bg="#ffffff")
        pr_frame.grid(row=0, column=1, padx=5, sticky="ew")
        tk.Label(pr_frame, text="PR Mensile PVSyst:", bg="#ffffff", fg=self.text_color, font=("Segoe UI Semibold", 9)).pack(anchor="w")
        self.entry_pr = ttk.Entry(pr_frame, textvariable=self.pvsyst_pr_var, width=10, font=("Segoe UI", 10))
        self.entry_pr.pack(anchor="w", pady=2)
        
        # Irradiance Threshold
        thresh_frame = tk.Frame(params_grid, bg="#ffffff")
        thresh_frame.grid(row=0, column=2, padx=(5, 0), sticky="ew")
        tk.Label(thresh_frame, text="Irraggiamento Min (W/m²):", bg="#ffffff", fg=self.text_color, font=("Segoe UI Semibold", 9)).pack(anchor="w")
        self.entry_thresh = ttk.Entry(thresh_frame, textvariable=self.threshold_var, width=10, font=("Segoe UI", 10))
        self.entry_thresh.pack(anchor="w", pady=2)
        
        # Force reprocess checkbox (Batch mode)
        chk_frame = tk.Frame(inputs_card, bg="#ffffff")
        chk_frame.pack(anchor="w", pady=(8, 8), fill="x")
        self.chk_reprocess = ttk.Checkbutton(
            chk_frame, 
            variable=self.reprocess_all_var,
            text="Ricalcola forzatamente i giorni già elaborati (Modalità Batch)"
        )
        self.chk_reprocess.pack(side="left")
        
        # Action button
        self.btn_calculate = ttk.Button(inputs_card, text="Calcola Performance Ratio", style="Action.TButton", command=self.start_calculation)
        self.btn_calculate.pack(fill="x", pady=(5, 5))
        
        # Progress/Status
        self.lbl_status = tk.Label(inputs_card, text="Pronto. Seleziona la cartella e clicca su Calcola.", bg="#ffffff", fg=self.muted_text, font=("Segoe UI", 10))
        self.lbl_status.pack(anchor="w", pady=(2, 0))
        
        # 2. Metrics Frame (Right Card - now containing PVSyst Table)
        metrics_card_border, self.metrics_card = self.create_card(top_grid, padding=15)
        metrics_card_border.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        
        # Dummy parent for keeping label references without packing them to avoid breaking update logic
        self.dummy_parent = tk.Frame(self.root)
        self.lbl_avg_pr_val = tk.Label(self.dummy_parent)
        self.lbl_comp_pr_val = tk.Label(self.dummy_parent)
        self.lbl_pvsyst_target_val = tk.Label(self.dummy_parent)
        self.lbl_irrad_summary = tk.Label(self.dummy_parent)
        
        lbl_sec_me = tk.Label(self.metrics_card, text="Tabella Riferimento Target PVSyst", bg="#ffffff", fg=self.accent_color, font=("Segoe UI Semibold", 12, "bold"))
        lbl_sec_me.pack(anchor="w", pady=(0, 2))
        
        # Add descriptive help text for PVSyst reference table
        lbl_desc_me = tk.Label(self.metrics_card, text="Valori teorici mensili di Target PR. Il mese attivo viene evidenziato in automatico.", bg="#ffffff", fg=self.muted_text, font=("Segoe UI", 10))
        lbl_desc_me.pack(anchor="w", pady=(0, 8))
        
        # Table frame with 1px border
        pvsyst_table_border = tk.Frame(self.metrics_card, bg="#dadce0")
        pvsyst_table_border.pack(fill="both", expand=True)
        
        pvsyst_table_frame = tk.Frame(pvsyst_table_border, bg="#ffffff")
        pvsyst_table_frame.pack(fill="both", expand=True, padx=1, pady=1)
        
        cols_pvsyst = ("Mese", "Target PR", "Target PR (%)")
        self.pvsyst_tree = ttk.Treeview(pvsyst_table_frame, columns=cols_pvsyst, show="headings", height=10)
        self.pvsyst_tree.pack(fill="both", expand=True)
        
        for c in cols_pvsyst:
            self.pvsyst_tree.heading(c, text=c)
            align = "w" if c == "Mese" else "center"
            self.pvsyst_tree.column(c, width=120, anchor=align)
            
        months_data = [
            ("Gennaio", "0.904", "90.4%"),
            ("Febbraio", "0.896", "89.6%"),
            ("Marzo", "0.897", "89.7%"),
            ("Aprile", "0.868", "86.8%"),
            ("Maggio", "0.832", "83.2%"),
            ("Giugno", "0.833", "83.3%"),
            ("Luglio", "0.820", "82.0%"),
            ("Agosto", "0.828", "82.8%"),
            ("Settembre", "0.852", "85.2%"),
            ("Ottobre", "0.876", "87.6%"),
            ("Novembre", "0.894", "89.4%"),
            ("Dicembre", "0.900", "90.0%")
        ]
        for idx, (m, dec, pct) in enumerate(months_data, start=1):
            self.pvsyst_tree.insert("", "end", iid=f"m{idx}", values=(m, dec, pct))
        
        # Bottom Grid: Results (Card)
        bottom_card_border, bottom_card = self.create_card(main_frame, padding=15)
        bottom_card_border.pack(fill="both", expand=True, pady=(15, 0))
        
        self.notebook = ttk.Notebook(bottom_card)
        self.notebook.pack(fill="both", expand=True)
        
        # Tab 1: Detailed Inverters
        tab1 = ttk.Frame(self.notebook, style="Card.TFrame")
        self.notebook.add(tab1, text="Dettaglio Inverter (Ultimo Giorno)")
        
        lbl_sec_det = tk.Label(tab1, text="Dettaglio Performance Ratio Compensato (36 Inverter)", bg="#ffffff", fg=self.accent_color, font=("Segoe UI Semibold", 12, "bold"))
        lbl_sec_det.pack(anchor="w", pady=(10, 2), padx=5)
        
        lbl_desc_det = tk.Label(tab1, text="Analisi dettagliata per ciascuno dei 36 inverter attivi nell'impianto per l'ultimo giorno elaborato.", bg="#ffffff", fg=self.muted_text, font=("Segoe UI", 10))
        lbl_desc_det.pack(anchor="w", pady=(0, 8), padx=5)
        
        # Table frame with 1px border
        table_border = tk.Frame(tab1, bg="#dadce0")
        table_border.pack(fill="both", expand=True, padx=5, pady=5)
        
        table_frame = tk.Frame(table_border, bg="#ffffff")
        table_frame.pack(fill="both", expand=True, padx=1, pady=1)
        
        scrollbar = ttk.Scrollbar(table_frame)
        scrollbar.pack(side="right", fill="y")
        
        cols = ("Codice Inverter", "Trasformatore", "Potenza CC Nominale (kW)", "Energia Prodotta (kWh)", "Perdita Stimata (kWh)", "PR Compensato (%)")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.tree.yview)
        
        for c in cols:
            self.tree.heading(c, text=c)
            align = "center" if c != "Codice Inverter" else "w"
            width = 110 if c != "Codice Inverter" else 150
            self.tree.column(c, width=width, anchor=align)
 
        # Tab 2: Daily Summary
        tab2 = ttk.Frame(self.notebook, style="Card.TFrame")
        self.notebook.add(tab2, text="Riepilogo Giorni Elaborati")
        
        lbl_sec_days = tk.Label(tab2, text="Risultati Giornalieri (Modalità Batch)", bg="#ffffff", fg=self.accent_color, font=("Segoe UI Semibold", 12, "bold"))
        lbl_sec_days.pack(anchor="w", pady=(10, 2), padx=5)
        
        lbl_desc_days = tk.Label(tab2, text="Cronologia dei risultati giornalieri calcolati e salvati per il mese selezionato.", bg="#ffffff", fg=self.muted_text, font=("Segoe UI", 10))
        lbl_desc_days.pack(anchor="w", pady=(0, 8), padx=5)
        
        # Table frame with 1px border
        days_border = tk.Frame(tab2, bg="#dadce0")
        days_border.pack(fill="both", expand=True, padx=5, pady=5)
        
        days_frame = tk.Frame(days_border, bg="#ffffff")
        days_frame.pack(fill="both", expand=True, padx=1, pady=1)
        
        days_scrollbar = ttk.Scrollbar(days_frame)
        days_scrollbar.pack(side="right", fill="y")
        
        days_cols = ("Data", "Irradiazione (kWh/m²)", "PR Non Comp. (%)", "PR Grezzo Comp. (%)", "Media PR Inv. (%)")
        self.tree_days = ttk.Treeview(days_frame, columns=days_cols, show="headings", yscrollcommand=days_scrollbar.set)
        self.tree_days.pack(side="left", fill="both", expand=True)
        days_scrollbar.config(command=self.tree_days.yview)
        
        for c in days_cols:
            self.tree_days.heading(c, text=c)
            self.tree_days.column(c, width=150, anchor="center")
            
        # Export Panel
        export_frame = tk.Frame(bottom_card, bg="#ffffff")
        export_frame.pack(fill="x", pady=(10, 0))
        
        self.btn_export = ttk.Button(export_frame, text="Esporta Dati Completi su Excel...", state="disabled", command=self.export_to_excel)
        self.btn_export.pack(side="right")
        
        self.lbl_export_status = tk.Label(export_frame, text="", bg="#ffffff", fg=self.muted_text, font=("Segoe UI", 10))
        self.lbl_export_status.pack(side="left", anchor="center")
        
        # 4. Live Log Console Panel (Bottom-most)
        log_card_border, log_card = self.create_card(main_frame, padding=12)
        log_card_border.pack(fill="both", expand=True, pady=(15, 0))
        
        lbl_sec_log = tk.Label(log_card, text="Console Live Log di Esecuzione", bg="#ffffff", fg=self.accent_color, font=("Segoe UI Semibold", 12, "bold"))
        lbl_sec_log.pack(anchor="w", pady=(0, 2))
        
        lbl_desc_log = tk.Label(log_card, text="Messaggi diagnostici in tempo reale sull'elaborazione dei file excel e calcoli.", bg="#ffffff", fg=self.muted_text, font=("Segoe UI", 10))
        lbl_desc_log.pack(anchor="w", pady=(0, 6))
        
        log_border = tk.Frame(log_card, bg="#dadce0")
        log_border.pack(fill="both", expand=True, pady=(5, 0))
        
        log_frame = tk.Frame(log_border, bg="#ffffff")
        log_frame.pack(fill="both", expand=True, padx=1, pady=1)
        
        log_scrollbar = ttk.Scrollbar(log_frame)
        log_scrollbar.pack(side="right", fill="y")
        
        self.log_widget = tk.Text(
            log_frame, 
            height=6, 
            bg="#f8f9fa", 
            fg=self.text_color, 
            insertbackground=self.text_color, 
            relief="flat", 
            bd=0,
            font=("Consolas", 9), 
            yscrollcommand=log_scrollbar.set,
            wrap="word",
            state="disabled"
        )
        self.log_widget.pack(side="left", fill="both", expand=True)
        log_scrollbar.config(command=self.log_widget.yview)
        
    def browse_folder(self):
        selected_folder = filedialog.askdirectory(title="Seleziona Cartella Mese (formato AAAA MM)")
        if selected_folder:
            self.folder_path_var.set(selected_folder)
            self.lbl_status.config(text="Cartella selezionata: " + os.path.basename(selected_folder), foreground=self.accent_color)
            # Try to auto-detect date from folder name or file names
            self.auto_detect_date(selected_folder)
            
    def auto_detect_date(self, folder):
        try:
            if not folder or not os.path.exists(folder):
                return
                
            # 1. Try to read from any Excel file inside the folder or its subdirectories
            excel_files = glob.glob(os.path.join(folder, "*.xlsx"))
            if not excel_files:
                # Check subdirectories (like '01' daily folder)
                for subd in os.listdir(folder):
                    subd_path = os.path.join(folder, subd)
                    if os.path.isdir(subd_path) and subd.isdigit():
                        excel_files = glob.glob(os.path.join(subd_path, "*.xlsx"))
                        if excel_files:
                            break
                            
            if excel_files:
                import openpyxl
                import datetime
                import re
                for file_path in excel_files[:3]:  # check up to 3 files to be fast
                    try:
                        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                        sheet = wb.active
                        found = False
                        for row in sheet.iter_rows(max_row=30, max_col=10, values_only=True):
                            for cell in row:
                                if isinstance(cell, (datetime.datetime, datetime.date)):
                                    self.date_var.set(f"{cell.year}-{cell.month:02d}-{cell.day:02d}")
                                    found = True
                                    break
                                if isinstance(cell, str):
                                    m = re.search(r'\b(\d{4})[-/](0[1-9]|1[0-2])[-/](\d{2,4})\b', cell)
                                    if m:
                                        self.date_var.set(f"{m.group(1)}-{m.group(2)}-{int(m.group(3)):02d}")
                                        found = True
                                        break
                                    m2 = re.search(r'\b(\d{1,2})[-/](0[1-9]|1[0-2])[-/](\d{4})\b', cell)
                                    if m2:
                                        self.date_var.set(f"{m2.group(3)}-{m2.group(2)}-{int(m2.group(1)):02d}")
                                        found = True
                                        break
                            if found:
                                return
                    except Exception:
                        continue

            # 2. Fallback to folder name / path parsing
            basename = os.path.basename(folder.rstrip("\\/"))
            parent = os.path.dirname(folder.rstrip("\\/"))
            parent_name = os.path.basename(parent)
            
            # If folder name is YYYY MM, e.g. "2026 05"
            parts = basename.split()
            if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit() and len(parts[0]) == 4:
                year = parts[0]
                month = parts[1]
                self.date_var.set(f"{year}-{month}-01")
                return
                
            if len(basename) == 2 and basename.isdigit():
                # Looks like a day, check parent
                if len(parent_name) == 7 and parent_name[4] == ' ' and parent_name[:4].isdigit() and parent_name[5:].isdigit():
                    year = parent_name[:4]
                    month = parent_name[5:]
                    day = basename
                    self.date_var.set(f"{year}-{month}-{day}")
                    return
                    
            # Or look at active power regulation file name
            files = os.listdir(folder)
            for f in files:
                if "regolazione_della_potenza_attiva_" in f.lower():
                    # Extract date from name like Regolazione_della_potenza_attiva_2026_04_26.xlsx
                    parts = f.replace(".xlsx", "").split("_")
                    # Look for 4 digit year
                    for idx, part in enumerate(parts):
                        if len(part) == 4 and part.isdigit() and idx <= len(parts) - 3:
                            if parts[idx+1].isdigit() and parts[idx+2].isdigit():
                                self.date_var.set(f"{parts[idx]}-{parts[idx+1]}-{parts[idx+2]}")
                                return
        except Exception:
            pass
            
    def show_guide(self):
        html_path = get_resource_path("Manuale_Utente_PR_Calculator.html")
        if os.path.exists(html_path):
            import webbrowser
            webbrowser.open(f"file:///{os.path.abspath(html_path)}")
        else:
            messagebox.showinfo(
                "Guida Utente",
                "Manuale Utente non trovato in locale.\n\n"
                "Procedura d'uso:\n"
                "1. Seleziona la cartella del mese (es. '2026 05').\n"
                "2. La data e il Target PR PVSyst verranno rilevati in automatico.\n"
                "3. Clicca su 'Calcola Performance Ratio' per elaborare i file SCADA.\n"
                "4. Esporta i risultati su Excel tramite il pulsante in basso."
            )
            
    def on_date_changed(self, *args):
        try:
            date_str = self.date_var.get().strip()
            parts = date_str.split("-")
            if len(parts) >= 2:
                month_val = int(parts[1])
                pvsyst_defaults = {
                    1: 0.904,   # January
                    2: 0.896,   # February
                    3: 0.897,   # March
                    4: 0.868,   # April
                    5: 0.832,   # May
                    6: 0.833,   # June
                    7: 0.820,   # July
                    8: 0.828,   # August
                    9: 0.852,   # September
                    10: 0.876,  # October
                    11: 0.894,  # November
                    12: 0.900   # December
                }
                if month_val in pvsyst_defaults:
                    self.pvsyst_pr_var.set(f"{pvsyst_defaults[month_val]:.3f}")
                    # Highlight matching row in the PVSyst reference treeview
                    try:
                        self.pvsyst_tree.selection_set(f"m{month_val}")
                        self.pvsyst_tree.focus(f"m{month_val}")
                        self.pvsyst_tree.see(f"m{month_val}")
                    except Exception:
                        pass
        except Exception:
            pass
            
    def on_pvsyst_pr_changed(self, *args):
        try:
            val = float(self.pvsyst_pr_var.get())
            self.lbl_pvsyst_target_val.config(text=f"{val * 100:.3f} %")
        except ValueError:
            self.lbl_pvsyst_target_val.config(text="-- %")
            
    def start_calculation(self):
        folder = self.folder_path_var.get()
        if not folder or not os.path.exists(folder):
            messagebox.showerror("Errore", "Per favore seleziona una cartella di input valida!")
            return
            
        try:
            pvsyst_pr = float(self.pvsyst_pr_var.get())
            threshold = float(self.threshold_var.get())
        except ValueError:
            messagebox.showerror("Errore", "PR PVSyst e soglia Irraggiamento devono essere numeri validi!")
            return
            
        date_str = self.date_var.get().strip()
        if len(date_str) != 10 or date_str[4] != '-' or date_str[7] != '-':
            messagebox.showerror("Errore", "La data deve essere nel formato AAAA-MM-GG!")
            return
            
        self.btn_calculate.config(state="disabled")
        self.btn_export.config(state="disabled")
        self.lbl_status.config(text="Calcolo del PR in corso... attendere prego...", foreground=self.warn_color)
        
        # Run calculation in a separate thread to keep UI active
        thread = threading.Thread(target=self.run_calculation, args=(folder, date_str, pvsyst_pr, threshold))
        thread.start()
        
    def find_file_by_patterns(self, folder, patterns):
        for pattern in patterns:
            # Simple direct match
            matches = glob.glob(os.path.join(folder, pattern))
            if matches:
                return matches[0]
            # Case insensitive match
            for file in os.listdir(folder):
                file_path = os.path.join(folder, file)
                if os.path.isfile(file_path):
                    # Check if pattern without asterisks is part of the filename
                    core_pat = pattern.replace("*", "").lower()
                    if core_pat in file.lower():
                        return file_path
        return None

    def calculate_single_day(self, folder, date_str, pvsyst_pr, threshold, calcolo_folder=None, skip_mother_update=False):
        import shutil
        import openpyxl
        import datetime
        
        # Format date components
        date_replaced = date_str.replace("-", "_") # e.g. 2026_04_26
        
        # Find required files
        reg_patterns = [f"Regolazione_della_potenza_attiva_{date_replaced}.xlsx", f"Regolazione_potenza_attiva_{date_replaced}.xlsx", "*potenza_attiva*.xlsx"]
        satac_patterns = ["SATAC_Meter_15Min.xlsx", "SATAC_Meter*.xlsx", "*SATAC*.xlsx"]
        ts1_w_patterns = ["TS_01_Weather_15Min.xlsx", "*Weather*01*.xlsx", "*TS_01*Weather*.xlsx"]
        ts3_w_patterns = ["TS_03_Weather_15Min.xlsx", "*Weather*03*.xlsx", "*TS_03*Weather*.xlsx"]
        ts1_i_patterns = ["TS_01_Inverter_15Min.xlsx", "*Inverter*01*.xlsx", "*TS_01*Inverter*.xlsx"]
        ts2_i_patterns = ["TS_02_Inverter_15Min.xlsx", "*Inverter*02*.xlsx", "*TS_02*Inverter*.xlsx"]
        ts3_i_patterns = ["TS_03_Inverter_15Min.xlsx", "*Inverter*03*.xlsx", "*TS_03*Inverter*.xlsx"]
        
        reg_file = self.find_file_by_patterns(folder, reg_patterns)
        satac_file = self.find_file_by_patterns(folder, satac_patterns)
        ts1_w_file = self.find_file_by_patterns(folder, ts1_w_patterns)
        ts3_w_file = self.find_file_by_patterns(folder, ts3_w_patterns)
        ts1_i_file = self.find_file_by_patterns(folder, ts1_i_patterns)
        ts2_i_file = self.find_file_by_patterns(folder, ts2_i_patterns)
        ts3_i_file = self.find_file_by_patterns(folder, ts3_i_patterns)
        
        missing_files = []
        if not reg_file: missing_files.append("Active Power Regulation")
        if not satac_file: missing_files.append("SATAC Meter Reading")
        if not ts1_w_file: missing_files.append("TS1 Weather Station")
        if not ts3_w_file: missing_files.append("TS3 Weather Station")
        if not ts1_i_file: missing_files.append("TS1 Inverters")
        if not ts2_i_file: missing_files.append("TS2 Inverters")
        if not ts3_i_file: missing_files.append("TS3 Inverters")
        
        if missing_files:
            err_msg = f"I seguenti file richiesti sono assenti nella cartella '{os.path.basename(folder)}':\n"
            for mf in missing_files:
                err_msg += f"- {mf}\n"
            raise FileNotFoundError(err_msg)
            
        # 1. Load active power regulation
        df_reg = None
        try:
            df_reg = pd.read_excel(reg_file)
        except Exception:
            for enc in ['utf-16', 'utf-16-le', 'utf-16-be', 'utf-8-sig', 'utf-8', 'latin-1']:
                try:
                    df_reg = pd.read_csv(reg_file, sep='\t', encoding=enc)
                    if len(df_reg.columns) > 1:
                        break
                except Exception:
                    continue
                    
        if df_reg is None or len(df_reg.columns) <= 1:
            raise ValueError(f"Could not read regulation file with any known encoding or format: {reg_file}")
            
        val_col = None
        for col in df_reg.columns:
            if 'potenza attiva' in col.lower() or 'valore nominale' in col.lower():
                val_col = col
                break
        if val_col is None:
            raise ValueError(f"Could not find 'potenza attiva' column in regulation file. Columns: {list(df_reg.columns)}")
        df_reg['limit_ratio'] = df_reg[val_col].astype(str).str.replace(',', '.').astype(float) / 100.0
        
        def normalize_columns(df):
            if 'Colonna2' not in df.columns and len(df.columns) >= 6:
                row_data = pd.DataFrame([df.columns.values], columns=[f"Colonna{i+1}" for i in range(len(df.columns))])
                df.columns = [f"Colonna{i+1}" for i in range(len(df.columns))]
                return pd.concat([row_data, df], ignore_index=True)
            return df

        # 2. Load weather data (TX1 and TX3)
        df_w1 = normalize_columns(pd.read_excel(ts1_w_file))
        df_w3 = normalize_columns(pd.read_excel(ts3_w_file))
        
        df_poa1 = df_w1[df_w1['Colonna2'].astype(str).str.strip() == "POA"].copy()
        df_poa3 = df_w3[df_w3['Colonna2'].astype(str).str.strip() == "POA"].copy()
        
        # 3. Load meter reading data
        df_m = normalize_columns(pd.read_excel(satac_file))
        df_meter = df_m[df_m['Colonna2'].astype(str).str.strip().str.startswith("Energia attiva prod")].copy()
        
        # Load previous day's meter reading if available
        df_meter_prev = None
        try:
            dt_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            prev_dt_obj = dt_obj - datetime.timedelta(days=1)
            
            # Find project root and construct previous day's folder
            month_folder = os.path.dirname(folder)
            project_root = os.path.dirname(month_folder)
            
            prev_month_folder = os.path.join(project_root, f"{prev_dt_obj.year} {prev_dt_obj.month:02d}")
            prev_day_folder = os.path.join(prev_month_folder, f"{prev_dt_obj.day:02d}")
            
            if os.path.exists(prev_day_folder):
                prev_satac_file = self.find_file_by_patterns(prev_day_folder, satac_patterns)
                if prev_satac_file:
                    df_m_prev = normalize_columns(pd.read_excel(prev_satac_file))
                    df_meter_prev = df_m_prev[df_m_prev['Colonna2'].astype(str).str.strip().str.startswith("Energia attiva prod")].copy()
                    print(f"[{date_str}] Letture contatore SATAC del giorno precedente caricate da '{os.path.basename(prev_satac_file)}'.")
        except Exception as ex:
            print(f"[{date_str}] Avvertenza: Impossibile caricare le letture SATAC del giorno precedente: {ex}")
        
        # 4. Load inverter data (TX1, TX2, TX3)
        df_i1 = normalize_columns(pd.read_excel(ts1_i_file))
        df_i2 = normalize_columns(pd.read_excel(ts2_i_file))
        df_i3 = normalize_columns(pd.read_excel(ts3_i_file))
        
        df_pa1 = df_i1[df_i1['Colonna2'].astype(str).str.strip() == "Potenza attiva"].copy()
        df_pa2 = df_i2[df_i2['Colonna2'].astype(str).str.strip() == "Potenza attiva"].copy()
        df_pa3 = df_i3[df_i3['Colonna2'].astype(str).str.strip() == "Potenza attiva"].copy()
        
        times = pd.date_range("00:00:00", "23:45:00", freq="15min").time
        time_strs = [t.strftime("%H:%M:%S") for t in times]
        
        def clean_float(val):
            if val is None or pd.isna(val): return 0.0
            if isinstance(val, (int, float)): return float(val)
            s = str(val).strip().replace(',', '.').rstrip('.')
            try: return float(s)
            except Exception: return 0.0

        calc_rows = []
        for idx, t_str in enumerate(time_strs):
            p1_row = df_poa1[df_poa1['Colonna6'].astype(str).str[:8] == t_str]
            p3_row = df_poa3[df_poa3['Colonna6'].astype(str).str[:8] == t_str]
            
            poa1 = clean_float(p1_row['Colonna3'].values[0]) if len(p1_row) > 0 else 0.0
            poa3 = clean_float(p3_row['Colonna3'].values[0]) if len(p3_row) > 0 else 0.0
            
            poa1_kwh = poa1 / 4000.0
            poa3_kwh = poa3 / 4000.0
            
            poa_avg_kwh = (poa1_kwh + poa3_kwh) / 2.0
            poa_avg_w = (poa1 + poa3) / 2.0
            h = poa_avg_w if poa_avg_w > threshold else 0.0
            
            diff_pct = 0.0
            if poa1_kwh > 0 and poa3_kwh > 0:
                avg_val = (poa1_kwh + poa3_kwh) / 2.0
                if avg_val != 0:
                    diff_pct = abs(poa1_kwh - poa3_kwh) / avg_val
            
            if poa1_kwh == 0 and poa3_kwh == 0:
                poa_cond_max_kwh = 0.0
            else:
                if diff_pct > 0.03:
                    poa_cond_max_kwh = max(poa1_kwh, poa3_kwh)
                else:
                    poa_cond_max_kwh = poa_avg_kwh
            
            reg_row = df_reg[df_reg['Unnamed: 1'].astype(str).str[:8] == t_str]
            limit_ratio = reg_row['limit_ratio'].values[0] if len(reg_row) > 0 else 0.876
            
            m_row = df_meter[df_meter['Colonna6'].astype(str).str[:8] == t_str]
            m_val = clean_float(m_row['Colonna3'].values[0]) if len(m_row) > 0 else 0.0
            
            if idx == 0:
                # First interval of today: set previous reading to yesterday's last reading (at 23:45:00)
                yesterday_last_val = None
                if df_meter_prev is not None:
                    last_row = df_meter_prev[df_meter_prev['Colonna6'].astype(str).str[:8] == "23:45:00"]
                    if len(last_row) > 0:
                        yesterday_last_val = clean_float(last_row['Colonna3'].values[0])
                    else:
                        yesterday_last_val = clean_float(df_meter_prev['Colonna3'].values[-1]) if len(df_meter_prev) > 0 else None
                m_val_prev = yesterday_last_val if yesterday_last_val is not None else m_val
            else:
                # Subsequent intervals: set previous reading to the previous interval's reading of today
                prev_t_str = time_strs[idx - 1]
                prev_m_row = df_meter[df_meter['Colonna6'].astype(str).str[:8] == prev_t_str]
                m_val_prev = clean_float(prev_m_row['Colonna3'].values[0]) if len(prev_m_row) > 0 else m_val
            
            pa1_t = df_pa1[df_pa1['Colonna6'].astype(str).str[:8] == t_str]
            pa2_t = df_pa2[df_pa2['Colonna6'].astype(str).str[:8] == t_str]
            pa3_t = df_pa3[df_pa3['Colonna6'].astype(str).str[:8] == t_str]
            
            inv_powers = {}
            for i in range(1, 13):
                inv_name = f"MW(EA,MW(17,Data_Mod_TS_01_Inverter_{i:02d}.I01))"
                inv_row = pa1_t[pa1_t['Colonna1'] == inv_name]
                inv_powers[f"TX1-INV-{i}"] = clean_float(inv_row['Colonna3'].values[0]) if len(inv_row) > 0 else 0.0
                
            for i in range(1, 13):
                inv_name = f"MW(EG,MW(18,Data_Mod_TS_02_Inverter_{i:02d}.I01))"
                inv_row = pa2_t[pa2_t['Colonna1'] == inv_name]
                inv_powers[f"TX2-INV-{i}"] = clean_float(inv_row['Colonna3'].values[0]) if len(inv_row) > 0 else 0.0
                
            for i in range(1, 13):
                inv_name = f"MW(EM,MW(19,Data_Mod_TS_03_Inverter_{i:02d}.I01))"
                inv_row = pa3_t[pa3_t['Colonna1'] == inv_name]
                inv_powers[f"TX3-INV-{i}"] = clean_float(inv_row['Colonna3'].values[0]) if len(inv_row) > 0 else 0.0
                
            row_data = {
                "time": t_str,
                "poa1": poa1,
                "poa3": poa3,
                "poa1_kwh": poa1_kwh,
                "poa3_kwh": poa3_kwh,
                "poa_avg_kwh": poa_avg_kwh,
                "poa_avg_w": poa_avg_w,
                "h": h,
                "diff_pct": diff_pct,
                "poa_cond_max_kwh": poa_cond_max_kwh,
                "limit_ratio": limit_ratio,
                "meter_reading": m_val,
                "meter_prev_reading": m_val_prev,
                **inv_powers
            }
            calc_rows.append(row_data)
            
        df_result = pd.DataFrame(calc_rows)
        df_result['active_energy_prod'] = (df_result['meter_reading'] - df_result['meter_prev_reading']) * 1000.0
        
        for tx in ["TX1", "TX2", "TX3"]:
            cols = [f"{tx}-INV-{i}" for i in range(1, 13)]
            df_result[f"{tx}_Average_Power"] = df_result[cols].apply(lambda r: r[r > 0].mean() if len(r[r > 0]) > 0 else 0.0, axis=1)
            
        new_cols = {}
        for inv_id in self.dc_powers:
            dc = self.dc_powers[inv_id]
            tx_name = inv_id.split("-")[0]
            
            if tx_name in ["TX1", "TX3"]:
                dt_loss_s = np.where(
                    (df_result['h'] > threshold) & (df_result[inv_id] <= 0),
                    np.where(
                        df_result[f"{tx_name}_Average_Power"] > 0,
                        df_result[f"{tx_name}_Average_Power"] * 0.25,
                        (df_result['h'] / 1000.0) * dc * pvsyst_pr * 0.25
                    ),
                    0.0
                )
                curt_loss_s = np.where(
                    df_result['limit_ratio'] < 0.875,
                    np.maximum(0.0, np.minimum((df_result['h'] / 1000.0) * dc * pvsyst_pr, self.ac_power_all * 0.876) - self.ac_power_all * df_result['limit_ratio']) * 0.25,
                    0.0
                )
                new_cols[f"{inv_id}_dt_loss"] = dt_loss_s
                new_cols[f"{inv_id}_curt_loss"] = curt_loss_s
                new_cols[f"{inv_id}_loss"] = dt_loss_s + curt_loss_s
            else:
                # Nested logic for TX2:
                # If downtime condition AND average power <= 0:
                # loss = min(irrad_expected_power, ac_power_all * 0.876) * 0.25 (no curtailment added)
                # Else:
                # loss = (average_power * 0.25 if downtime and active <= 0 else 0) + curtailment_loss
                dt_loss_avg_zero = np.minimum((df_result['h'] / 1000.0) * dc * pvsyst_pr, self.ac_power_all * 0.876) * 0.25
                dt_loss_avg_pos = np.where(
                    (df_result['h'] > threshold) & (df_result[inv_id] <= 0),
                    df_result[f"{tx_name}_Average_Power"] * 0.25,
                    0.0
                )
                curt_loss_s = np.where(
                    df_result['limit_ratio'] < 0.875,
                    np.maximum(0.0, np.minimum((df_result['h'] / 1000.0) * dc * pvsyst_pr, self.ac_power_all * 0.876) - self.ac_power_all * df_result['limit_ratio']) * 0.25,
                    0.0
                )
                dt_loss_s = np.where(
                    (df_result['h'] > threshold) & (df_result[inv_id] <= 0) & (df_result[f"{tx_name}_Average_Power"] <= 0),
                    dt_loss_avg_zero,
                    dt_loss_avg_pos
                )
                loss_s = np.where(
                    (df_result['h'] > threshold) & (df_result[inv_id] <= 0) & (df_result[f"{tx_name}_Average_Power"] <= 0),
                    dt_loss_avg_zero,
                    dt_loss_avg_pos + curt_loss_s
                )
                new_cols[f"{inv_id}_dt_loss"] = dt_loss_s
                new_cols[f"{inv_id}_curt_loss"] = curt_loss_s
                new_cols[f"{inv_id}_loss"] = loss_s
            
        for tx in ["TX1", "TX2", "TX3"]:
            loss_cols = [f"{tx}-INV-{i}_loss" for i in range(1, 13)]
            new_cols[f"{tx}_Total_Loss"] = sum(new_cols[col] for col in loss_cols)
            
        df_result = pd.concat([df_result, pd.DataFrame(new_cols, index=df_result.index)], axis=1)
            
        h_sum = df_result['h'].sum()
        h_sum_kwh = h_sum / 4000.0
        
        inv_prs = {}
        inverter_table_data = []
        for inv_id in self.dc_powers:
            dc = self.dc_powers[inv_id]
            energy_gen = df_result[inv_id].sum() * 0.25
            total_loss = df_result[f"{inv_id}_loss"].sum()
            numerator = energy_gen + total_loss
            denominator = dc * h_sum_kwh
            pr_val = (numerator / denominator * 100.0) if denominator > 0 else 0.0
            inv_prs[inv_id] = pr_val
            tx_name = inv_id.split("-")[0]
            inverter_table_data.append((inv_id, tx_name, dc, f"{energy_gen:.2f}", f"{total_loss:.2f}", f"{pr_val:.3f}"))
            
        avg_inv_pr = np.mean(list(inv_prs.values()))
        total_energy = sum(df_result[inv_id].sum() * 0.25 for inv_id in self.dc_powers)
        uncomp_pr = total_energy / (12625.0 * h_sum_kwh) * 100.0
        total_losses_all = sum(df_result[f"{tx}_Total_Loss"].sum() for tx in ["TX1", "TX2", "TX3"])
        comp_raw_pr = (total_energy + total_losses_all) / (12625.0 * h_sum_kwh) * 100.0
        
        calc_results = {
            "avg_inv_pr": avg_inv_pr,
            "comp_raw_pr": comp_raw_pr,
            "uncomp_pr": uncomp_pr,
            "h_sum_kwh": h_sum_kwh,
            "inverter_table_data": inverter_table_data,
            "date_str": date_str,
            "dc_powers": self.dc_powers
        }
        
        # Determine output folder
        if not calcolo_folder:
            parent_folder = os.path.dirname(folder)
            calcolo_folder = os.path.join(parent_folder, "PR CALCOLO FILE")
            
        os.makedirs(calcolo_folder, exist_ok=True)
        
        month_abbrs = {
            1: "gen", 2: "feb", 3: "mar", 4: "apr", 5: "mag", 6: "giu",
            7: "lug", 8: "ago", 9: "set", 10: "ott", 11: "nov", 12: "dic"
        }
        italian_months_4 = {
            1: "GENN", 2: "FEBB", 3: "MARZ", 4: "APRL", 5: "MAGG", 6: "GIUG",
            7: "LUGL", 8: "AGOS", 9: "SETT", 10: "OTTO", 11: "NOVE", 12: "DICE"
        }
        
        date_parts = date_str.split("-")
        year_val = int(date_parts[0])
        month_val = int(date_parts[1])
        day_val = int(date_parts[2])
        month_name = month_abbrs[month_val]
        
        daily_filename = f"PR_recalculation_{day_val:02d}_{month_name}.xlsx"
        daily_file_path = os.path.join(calcolo_folder, daily_filename)
        
        # Identify pristine template daily file from original_format to avoid openpyxl cumulative corruption
        original_format_dir = get_resource_path("original_format")
        orig_templates = glob.glob(os.path.join(original_format_dir, "PR_recalculation_*.xlsx"))
        template_file = None
        for tf in orig_templates:
            bname = os.path.basename(tf)
            if "00 PR_recalculation" not in bname and "_test_" not in bname:
                template_file = tf
                break
        
        # If the file does not exist, copy from pristine template
        if not os.path.exists(daily_file_path):
            if template_file:
                print(f"[{date_str}] Copia del template Excel giornaliero pulito: '{os.path.basename(template_file)}' -> '{daily_filename}'")
                shutil.copy(template_file, daily_file_path)
            else:
                raise FileNotFoundError("Template Excel giornaliero originale non trovato nella cartella original_format!")
        else:
            print(f"[{date_str}] Il file giornaliero di destinazione '{daily_filename}' esiste già; aggiornamento in corso.")
            
        # Open and write values to daily workbook natively via Excel COM to prevent openpyxl table corruption!
        excel_daily = None
        try:
            import win32com.client
            excel_daily = get_excel_app()
            abs_daily_path = os.path.abspath(daily_file_path).replace('/', '\\')
            print(f"[{date_str}] DEBUG: Apertura cartella di lavoro giornaliera: {abs_daily_path}")
            wb_daily = excel_daily.Workbooks.Open(abs_daily_path, UpdateLinks=0)
            print(f"[{date_str}] DEBUG: Cartella di lavoro aperta con successo!")
            
            try:
                excel_daily.Calculation = -4135  # xlCalculationManual
                excel_daily.CalculateBeforeSave = False
            except Exception:
                pass
                
            ws_calc = wb_daily.Sheets('PR_Calc')
            ws_inv = wb_daily.Sheets('Inverter_data')
            
            dt_obj = datetime.datetime(year_val, month_val, day_val, 0, 0)
            
            # Prepare fast list assignments for 96 rows (Row 15 to 110)
            # Find matching df_result rows for each slot
            time_slots = []
            for r in range(15, 111):
                time_val = ws_calc.Cells(r, 2).Value
                time_str_prefix = str(time_val)[:8] if time_val else ""
                time_slots.append(time_str_prefix)
                
            print(f"[{date_str}] DEBUG: Lettura Colonna B completata (96 righe).")
                
            calc_rows_ordered = []
            for t_prefix in time_slots:
                match = df_result[df_result['time'].str.startswith(t_prefix)]
                if len(match) > 0:
                    calc_rows_ordered.append(match.iloc[0].to_dict())
                else:
                    # fallback dummy row
                    calc_rows_ordered.append({
                        'poa1': 0.0, 'poa1_kwh': 0.0, 'poa3': 0.0, 'poa3_kwh': 0.0,
                        'meter_prev_reading': 0.0, 'meter_reading': 0.0, 'limit_ratio': 0.876,
                        **{f"TX1-INV-{i}": 0.0 for i in range(1, 13)},
                        **{f"TX2-INV-{i}": 0.0 for i in range(1, 13)},
                        **{f"TX3-INV-{i}": 0.0 for i in range(1, 13)},
                        **{f"TX1-INV-{i}_loss": 0.0 for i in range(1, 13)},
                        **{f"TX2-INV-{i}_loss": 0.0 for i in range(1, 13)},
                        **{f"TX3-INV-{i}_loss": 0.0 for i in range(1, 13)},
                    })
                    
            print(f"[{date_str}] DEBUG: Creazione calc_rows_ordered completata.")
            
            # Set today's date in Column A as a fast date string (avoids sluggish COM datetime serialization)
            date_col_vals = [[f"{year_val:04d}-{month_val:02d}-{day_val:02d}"] for _ in range(96)]
            print(f"[{date_str}] DEBUG: Scrittura date nella Colonna A di ws_calc...")
            ws_calc.Range("A15:A110").Value = date_col_vals
            print(f"[{date_str}] DEBUG: Scrittura date nella Colonna A di ws_inv...")
            ws_inv.Range("A15:A110").Value = date_col_vals
            print(f"[{date_str}] DEBUG: Scrittura date completata. Scrittura colonne dati PR_Calc...")
            
            # Write WS_Calc columns
            ws_calc.Range("C15:C110").Value = [[float(r['poa1'])] for r in calc_rows_ordered]
            ws_calc.Range("D15:D110").Value = [[float(r['poa1_kwh'])] for r in calc_rows_ordered]
            print(f"[{date_str}] DEBUG: Scritte colonne C e D di WS_Calc.")
            ws_calc.Range("E15:E110").Value = [[float(r['poa3'])] for r in calc_rows_ordered]
            ws_calc.Range("F15:F110").Value = [[float(r['poa3_kwh'])] for r in calc_rows_ordered]
            print(f"[{date_str}] DEBUG: Scritte colonne E e F di WS_Calc.")
            ws_calc.Range("K15:K110").Value = [[float(r['meter_prev_reading'])] for r in calc_rows_ordered]
            ws_calc.Range("L15:L110").Value = [[float(r['meter_reading'])] for r in calc_rows_ordered]
            ws_calc.Range("N15:N110").Value = [[float(r['limit_ratio'])] for r in calc_rows_ordered]
            
            # Permanently eliminate #DIV/0! errors from daily workbook formula columns
            for r in range(15, 111):
                ws_calc.Cells(r, 7).Formula = f"=IFERROR((D{r}+F{r})/2, 0)"
                ws_calc.Cells(r, 8).Formula = f"=IFERROR(IF(AVERAGE(C{r},E{r})>$BA$7,AVERAGE(C{r},E{r}),0), 0)"
                ws_calc.Cells(r, 9).Formula = f"=IFERROR(IF(AND(D{r}=0,F{r}=0),0,IF(J{r}>$BA$6,MAX(D{r},F{r}),G{r})), 0)"
                ws_calc.Cells(r, 10).Formula = f"=IFERROR(IF(AND(D{r}>0,F{r}>0),ABS(D{r}-F{r})/AVERAGE(D{r},F{r}),0), 0)"
                ws_calc.Cells(r, 13).Formula = f"=IFERROR((L{r}-K{r})*1000, 0)"
                
            # Programmatically enforce correct inverter PR formulas in row 111 (without losses)
            import openpyxl.utils
            for i in range(1, 13):
                # TX1 (Columns O to Z -> 15 to 26)
                col_calc = 14 + i
                col_inv = 2 + i
                col_calc_letter = openpyxl.utils.get_column_letter(col_calc)
                inv_col_letter = openpyxl.utils.get_column_letter(col_inv)
                ws_calc.Cells(111, col_calc).Formula2 = f"=SUM((Inverter_data!{inv_col_letter}15:{inv_col_letter}110)*0.25)/({col_calc_letter}$10*(SUM(PR_Calc!$H$15:$H$110)/4000))"
                
                # TX2 (Columns AB to AM -> 28 to 39)
                col_calc = 27 + i
                col_inv = 17 + i
                col_calc_letter = openpyxl.utils.get_column_letter(col_calc)
                inv_col_letter = openpyxl.utils.get_column_letter(col_inv)
                ws_calc.Cells(111, col_calc).Formula2 = f"=SUM((Inverter_data!{inv_col_letter}15:{inv_col_letter}110)*0.25)/({col_calc_letter}$10*(SUM(PR_Calc!$H$15:$H$110)/4000))"
                
                # TX3 (Columns AO to AZ -> 41 to 52)
                col_calc = 40 + i
                col_inv = 32 + i
                col_calc_letter = openpyxl.utils.get_column_letter(col_calc)
                inv_col_letter = openpyxl.utils.get_column_letter(col_inv)
                ws_calc.Cells(111, col_calc).Formula2 = f"=SUM((Inverter_data!{inv_col_letter}15:{inv_col_letter}110)*0.25)/({col_calc_letter}$10*(SUM(PR_Calc!$H$15:$H$110)/4000))"
                
            print(f"[{date_str}] DEBUG: Scrittura dati PR_Calc e formule riga 111 completata.")
            
            # Write Inverter_data columns
            tx1_vals = [[float(r[f"TX1-INV-{i}"]) for i in range(1, 13)] for r in calc_rows_ordered]
            ws_inv.Range("C15:N110").Value = tx1_vals
            
            tx2_vals = [[float(r[f"TX2-INV-{i}"]) for i in range(1, 13)] for r in calc_rows_ordered]
            ws_inv.Range("R15:AC110").Value = tx2_vals
            
            tx3_vals = [[float(r[f"TX3-INV-{i}"]) for i in range(1, 13)] for r in calc_rows_ordered]
            ws_inv.Range("AG15:AR110").Value = tx3_vals
            print(f"[{date_str}] DEBUG: Scrittura potenze Inverter_data completata.")

            # Write Python-computed per-inverter energy losses directly to PR_Calc (cols O-Z, AB-AM, AO-AZ)
            # This overrides the Excel array formulas and ensures losses are exactly what Python calculated,
            # independent of BA4 or any other Excel formula parameter.
            tx1_loss_vals = [[float(r.get(f"TX1-INV-{i}_loss", 0.0)) for i in range(1, 13)] for r in calc_rows_ordered]
            ws_calc.Range("O15:Z110").Value = tx1_loss_vals
            tx2_loss_vals = [[float(r.get(f"TX2-INV-{i}_loss", 0.0)) for i in range(1, 13)] for r in calc_rows_ordered]
            ws_calc.Range("AB15:AM110").Value = tx2_loss_vals
            tx3_loss_vals = [[float(r.get(f"TX3-INV-{i}_loss", 0.0)) for i in range(1, 13)] for r in calc_rows_ordered]
            ws_calc.Range("AO15:AZ110").Value = tx3_loss_vals
            print(f"[{date_str}] DEBUG: Scrittura perdite energia per inverter completata (O-Z, AB-AM, AO-AZ).")

            # Update titles and side tables
            italian_months_full_upper = {
                1: "GENNAIO", 2: "FEBBRAIO", 3: "MARZO", 4: "APRILE", 5: "MAGGIO", 6: "GIUGNO",
                7: "LUGLIO", 8: "AGOSTO", 9: "SETTEMBRE", 10: "OTTOBRE", 11: "NOVEMBRE", 12: "DICEMBRE"
            }
            english_months_full = {
                1: "January", 2: "February", 3: "March", 4: "April", 5: "May", 6: "June",
                7: "July", 8: "August", 9: "September", 10: "October", 11: "November", 12: "December"
            }
            
            ws_calc.Cells(1, 1).Value = "PR CALCULATION"
            ws_calc.Cells(2, 1).Value = f"{day_val:02d} {italian_months_full_upper[month_val]} {year_val}"
            ws_calc.Cells(8, 1).Value = "NOMINAL VALUES"
            
            # Write nominal parameters in Column BA (53)
            ws_calc.Cells(4, 53).Value = float(pvsyst_pr)           # BA4 = decimal PR (e.g. 0.897)
            ws_calc.Cells(5, 53).Formula = "=((SUM(Tabella01MarzoInverter[[Active power TX1-INV-1]:[Active power TX1-INV-12]]," \
                                           "Tabella01MarzoInverter[[Active power TX2-INV-1]:[Active power TX2-INV-12]]," \
                                           "Tabella01MarzoInverter[[Active power TX3-INV-1]:[Active power TX3-INV-12]]) * 0.25) )" \
                                           " / (12625 * (SUM($H$15:$H$110) * 0.25 / 1000))"
            ws_calc.Cells(6, 53).Value = 0.03                     # BA6 is irradiance acceptance limit ratio
            ws_calc.Cells(7, 53).Value = float(threshold)          # BA7 is irradiance minimum value (e.g. 50)
            
            # Write English PR calculation header in Column BD (56) Row 2
            ws_calc.Cells(2, 56).Value = f"{day_val} {english_months_full[month_val]} {year_val} PR Calculation"
            
            # Write PR from SCADA in Column BH (60) Row 8 (BH8 expects percentage e.g. 81.743)
            ws_calc.Cells(8, 60).Value = float(uncomp_pr)
            print(f"[{date_str}] DEBUG: Aggiornamento tabelle laterali completato. Salvataggio cartella di lavoro in corso...")
            
            try:
                excel_daily.Calculation = -4105  # xlCalculationAutomatic
            except Exception:
                pass
                
            wb_daily.Save()
            print(f"[{date_str}] File giornaliero '{daily_filename}' salvato con successo via Excel COM!")
        except Exception as ex:
            print(f"[{date_str}] Errore durante l'aggiornamento del file giornaliero via Excel COM: {ex}")
            if 'wb_daily' in locals() and wb_daily:
                try:
                    wb_daily.Close(SaveChanges=False)
                except Exception:
                    pass
            raise ex
        finally:
            pass
            
        # Update mother file
        if not skip_mother_update:
            self.sync_mother_file(calcolo_folder, year_val, month_val)
            
        return df_result, calc_results

    def sync_mother_file(self, calcolo_folder, year_val, month_val):
        import calendar
        import openpyxl
        import re
        import shutil
        import glob
        
        month_abbrs = {
            1: "gen", 2: "feb", 3: "mar", 4: "apr", 5: "mag", 6: "giu",
            7: "lug", 8: "ago", 9: "set", 10: "ott", 11: "nov", 12: "dic"
        }
        italian_months_4 = {
            1: "GENN", 2: "FEBB", 3: "MARZ", 4: "APRL", 5: "MAGG", 6: "GIUG",
            7: "LUGL", 8: "AGOS", 9: "SETT", 10: "OTTO", 11: "NOVE", 12: "DICE"
        }
        
        month_name = month_abbrs[month_val]
        expected_mother_filename = f"00 PR_recalculation_{italian_months_4[month_val]}.xlsx"
        mother_path = os.path.join(calcolo_folder, expected_mother_filename)
        
        initialized_new = False
        if not os.path.exists(mother_path):
            original_format_dir = get_resource_path("original_format")
            orig_mothers = glob.glob(os.path.join(original_format_dir, "00 PR_recalculation_*.xlsx"))
            if orig_mothers:
                orig_mother_path = orig_mothers[0]
                print(f"Copia e inizializzazione del template Madre: '{os.path.basename(orig_mother_path)}' -> '{expected_mother_filename}'")
                shutil.copy(orig_mother_path, mother_path)
                initialized_new = True
            else:
                raise FileNotFoundError(f"Template Madre originale non trovato in original_format per inizializzare '{expected_mother_filename}'!")

        excel = None
        wb_mother = None
        try:
            excel = get_excel_app()
            abs_mother_path = os.path.abspath(mother_path).replace('/', '\\')
            print(f"DEBUG: Apertura file Madre: {abs_mother_path}")
            wb_mother = excel.Workbooks.Open(abs_mother_path, UpdateLinks=0)
            
            try:
                excel.Calculation = -4135  # xlCalculationManual
            except Exception:
                pass
                
            ws_mother = wb_mother.Sheets('PR_Calc')
            num_days = calendar.monthrange(year_val, month_val)[1]
            target_summary_row = 5 + num_days
            
            # Programmatically ensure "Irradiance TX1" and "Irradiance TX3" are present at Columns B and C
            b4_val = ws_mother.Cells(4, 2).Value
            if not (b4_val and "Irradiance TX1" in str(b4_val)):
                print("DEBUG: Irradiance TX1 column not found at Column B. Programmatically inserting Irradiance TX1 and Irradiance TX3 columns...")
                ws_mother.Columns(2).Insert()
                ws_mother.Columns(2).Insert()
                ws_mother.Cells(4, 2).Value = "Irradiance TX1"
                ws_mother.Cells(4, 3).Value = "Irradiance TX3"

            # Programmatically ensure "PR VCOM" and "External Availability [%]" columns are present at Columns H and I
            h4_val = ws_mother.Cells(4, 8).Value
            if not (h4_val and "VCOM" in str(h4_val)):
                print("DEBUG: PR VCOM column not found at Column H. Programmatically inserting PR VCOM and External Availability columns...")
                ws_mother.Columns(8).Insert()
                ws_mother.Columns(8).Insert()
                ws_mother.Cells(4, 8).Value = "PR VCOM"
                ws_mother.Cells(4, 9).Value = "External Availability\n[%]"
                
            # Dynamically format and adjust summary row in existing Mother file if needed!
            current_summary_row = None
            for r in range(30, 42):
                f_text = ws_mother.Cells(r, 6).Formula
                if isinstance(f_text, str) and 'AVERAGE' in f_text.upper():
                    current_summary_row = r
                    break
                    
            if current_summary_row is not None:
                if current_summary_row < target_summary_row:
                    rows_to_insert = target_summary_row - current_summary_row
                    for _ in range(rows_to_insert):
                        ws_mother.Rows(current_summary_row).Insert()
                        ws_mother.Rows(current_summary_row - 1).Copy(ws_mother.Rows(current_summary_row))
                        excel.CutCopyMode = False
                        for c in range(1, 65):
                            ws_mother.Cells(current_summary_row, c).Value = None
                elif current_summary_row > target_summary_row:
                    rows_to_delete = current_summary_row - target_summary_row
                    del_start = target_summary_row
                    del_end = current_summary_row - 1
                    ws_mother.Rows(f"{del_start}:{del_end}").Delete()
                    
                ws_mother.Cells(target_summary_row, 6).Formula = f"=AVERAGE(F5:F{target_summary_row-1})"
                ws_mother.Cells(target_summary_row, 7).Formula = f"=AVERAGE(G5:G{target_summary_row-1})"
                ws_mother.Cells(target_summary_row, 8).Formula = f"=AVERAGE(H5:H{target_summary_row-1})"
                ws_mother.Cells(target_summary_row, 9).Formula = f"=SUMIF(I5:I{target_summary_row-1},\"<>0\")/COUNTIF(I5:I{target_summary_row-1},\"<>0\")"
            
            # Write/update formulas for External Availability in day rows (5 to 4 + num_days) after summary row has been adjusted
            for r in range(5, 5 + num_days):
                ws_mother.Cells(r, 9).Formula = f"=IF(E{r}=\"\",0,(E{r}/(E{r}+J{r}+K{r}+L{r}))*100)"
            
            # Change links natively via Excel to avoid openpyxl corruption if initialized new
            if initialized_new:
                links = wb_mother.LinkSources(1) # xlExcelLinks
                if links:
                    for link in links:
                        match = re.search(r"PR_recalculation_(\d+)_", link, re.IGNORECASE)
                        if match:
                            day_num = int(match.group(1))
                            if day_num <= num_days:
                                chk_daily_filename = f"PR_recalculation_{day_num:02d}_{month_name}.xlsx"
                                chk_daily_path = os.path.abspath(os.path.join(calcolo_folder, chk_daily_filename)).replace('/', '\\')
                                wb_mother.ChangeLink(Name=link, NewName=chk_daily_path, Type=1)

            # Ensure all day rows have correct literal dates for this month
            for r in range(5, 5 + num_days):
                day_num = r - 4
                ws_mother.Cells(r, 1).Value = f"{year_val}-{month_val:02d}-{day_num:02d}"
                
            # Dynamically map mother sheet columns based on header names (Row 4)
            header_mapping = {}
            for col in range(2, 65):
                val = ws_mother.Cells(4, col).Value
                if not val:
                    continue
                val_str = str(val).strip().lower()
                if "irradiance tx1" in val_str:
                    header_mapping[col] = "$D$111"
                elif "irradiance tx3" in val_str:
                    header_mapping[col] = "$F$111"
                elif "irradiance" in val_str or "irraggiamento" in val_str:
                    header_mapping[col] = "$I$111"
                elif "energy" in val_str and "loss" not in val_str and "perdita" not in val_str:
                    header_mapping[col] = "$M$111"
                elif "pr total" in val_str or "pr vcom" in val_str:
                    header_mapping[col] = "$BA$5*100"
                elif "pr scada" in val_str:
                    header_mapping[col] = "$BH$8"
                elif "tx1 - energy loss" in val_str or "tx1 - perdita" in val_str or "perdita energia tx1" in val_str:
                    header_mapping[col] = "$AA$111"
                elif "tx2 - energy loss" in val_str or "tx2 - perdita" in val_str or "perdita energia tx2" in val_str:
                    header_mapping[col] = "$AN$111"
                elif "tx3 - energy loss" in val_str or "tx3 - perdita" in val_str or "perdita energia tx3" in val_str:
                    header_mapping[col] = "$BA$111"
                elif "pr tx1-inv-" in val_str:
                    try:
                        inv_num = int(val_str.split("inv-")[-1])
                        daily_col = 14 + inv_num
                        header_mapping[col] = openpyxl.utils.get_column_letter(daily_col) + "$111"
                    except Exception:
                        pass
                elif "pr tx2-inv-" in val_str:
                    try:
                        inv_num = int(val_str.split("inv-")[-1])
                        daily_col = 27 + inv_num
                        header_mapping[col] = openpyxl.utils.get_column_letter(daily_col) + "$111"
                    except Exception:
                        pass
                elif "pr tx3-inv-" in val_str:
                    try:
                        inv_num = int(val_str.split("inv-")[-1])
                        daily_col = 40 + inv_num
                        header_mapping[col] = openpyxl.utils.get_column_letter(daily_col) + "$111"
                    except Exception:
                        pass
                
            daily_dir = os.path.abspath(calcolo_folder).replace('/', '\\')
            
            sync_count = 0
            for day_num in range(1, num_days + 1):
                chk_daily_filename = f"PR_recalculation_{day_num:02d}_{month_name}.xlsx"
                chk_daily_path = os.path.join(calcolo_folder, chk_daily_filename)
                r = 5 + day_num - 1

                if os.path.exists(chk_daily_path):
                    prefix = f"='{daily_dir}\\[{chk_daily_filename}]PR_Calc'"

                    formulas_to_write = {}
                    for col, addr in header_mapping.items():
                        expected_formula = f"{prefix}!{addr}"
                        try:
                            curr_f = ws_mother.Cells(r, col).Formula
                        except Exception:
                            curr_f = ""
                        if curr_f != expected_formula:
                            formulas_to_write[col] = expected_formula

                    if formulas_to_write:
                        print(f"Sincronizzazione formule per il giorno {day_num}...")
                        for col, f_val in formulas_to_write.items():
                            ws_mother.Cells(r, col).Formula = f_val
                        sync_count += 1
                else:
                    # Daily file not yet processed — clear any stale data/formulas left
                    # over from a previous month's template so the row stays blank.
                    for col in header_mapping.keys():
                        try:
                            cell = ws_mother.Cells(r, col)
                            if cell.Value is not None or cell.Formula:
                                cell.Value = None
                        except Exception:
                            pass
                        
            try:
                excel.Calculation = -4105  # xlCalculationAutomatic
                excel.CalculateBeforeSave = True
                wb_mother.Calculate()
            except Exception:
                pass
                
            wb_mother.Save()
            wb_mother.Close(SaveChanges=True)
            print(f"Sincronizzazione completata! Giorni aggiornati nel file Madre: {sync_count}")
        except Exception as ex:
            print(f"Errore durante l'aggiornamento del file Madre via Excel COM: {ex}")
            if wb_mother:
                try:
                    wb_mother.Close(SaveChanges=False)
                except Exception:
                    pass
            raise RuntimeError(f"Impossibile salvare o aggiornare il file Madre '{expected_mother_filename}'.\n"
                               f"Assicurarsi che il file non sia aperto in un'altra finestra di Excel o bloccato da un altro utente.\n"
                               f"Dettaglio errore: {ex}")

    def run_calculation(self, folder, date_str, pvsyst_pr, threshold):
        try:
            import datetime
            # Scan immediate subdirectories to check if this is a month folder (e.g. contains subdirectories '01', '02', '25', '26' etc.)
            subdirs = [d for d in os.listdir(folder) if os.path.isdir(os.path.join(folder, d))]
            numerical_subdirs = sorted([d for d in subdirs if d.isdigit() and 1 <= int(d) <= 31], key=lambda x: int(x))
            
            is_batch = len(numerical_subdirs) > 0
            
            if is_batch:
                # Batch Month Processing mode
                calcolo_folder = os.path.join(folder, "PR CALCOLO FILE")
                os.makedirs(calcolo_folder, exist_ok=True)
                
                # Dynamically parse Year & Month from folder name
                basename = os.path.basename(folder.rstrip("\\/"))
                parts = basename.split()
                year_val = datetime.datetime.now().year
                month_val = datetime.datetime.now().month
                if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
                    year_val = int(parts[0])
                    month_val = int(parts[1])
                    
                month_abbrs = {
                    1: "gen", 2: "feb", 3: "mar", 4: "apr", 5: "mag", 6: "giu",
                    7: "lug", 8: "ago", 9: "set", 10: "ott", 11: "nov", 12: "dic"
                }
                month_name = month_abbrs[month_val]
                
                reprocess_all = self.reprocess_all_var.get()
                processed_count = 0
                skipped_count = 0
                
                last_df_result = None
                last_calc_results = None
                self.all_days_results = []
                
                for idx, day_str in enumerate(numerical_subdirs):
                    day_val = int(day_str)
                    target_date_str = f"{year_val:04d}-{month_val:02d}-{day_val:02d}"
                    daily_filename = f"PR_recalculation_{day_val:02d}_{month_name}.xlsx"
                    daily_file_path = os.path.join(calcolo_folder, daily_filename)
                    
                    # If daily file already exists and we are not forcing reprocess, skip this day folder!
                    if os.path.exists(daily_file_path) and not reprocess_all:
                        skipped_count += 1
                        continue
                        
                    # Update status securely from the non-GUI thread
                    status_text = f"Modalità Batch: Elaborazione giorno {day_val} di {len(numerical_subdirs)}..."
                    self.root.after(0, lambda t=status_text: self.lbl_status.config(text=t, foreground=self.warn_color))
                    
                    day_folder = os.path.join(folder, day_str)
                    try:
                        last_df_result, last_calc_results = self.calculate_single_day(
                            day_folder, target_date_str, pvsyst_pr, threshold, calcolo_folder, skip_mother_update=True
                        )
                        self.all_days_results.append(last_calc_results)
                        processed_count += 1
                    except Exception as day_err:
                        raise ValueError(f"Errore nel Giorno {day_str}: {day_err}")
                        
                # Sincronizza sempre il file Madre alla fine del batch (aggiorna anche i giorni saltati)
                status_text = "Sincronizzazione finale del file Madre..."
                self.root.after(0, lambda t=status_text: self.lbl_status.config(text=t, foreground=self.warn_color))
                self.sync_mother_file(calcolo_folder, year_val, month_val)
                
                if processed_count == 0:
                    status_msg = f"Tutti i {skipped_count} giorni erano già elaborati. Sincronizzato il file Madre."
                    self.root.after(0, lambda m=status_msg: self.lbl_status.config(text=m, foreground=self.success_color))
                    self.root.after(0, lambda: self.btn_calculate.config(state="normal"))
                    return
                    
                # Display results of the last processed day on GUI dashboard
                self.df_result = last_df_result
                self.calc_results = last_calc_results
                
                success_msg = f"Batch completato! Elaborati: {processed_count} giorni, Sincronizzato file Madre."
                self.root.after(0, lambda m=success_msg: self.lbl_status.config(text=m, foreground=self.success_color))
                self.root.after(0, self.update_ui_on_success)
                
            else:
                # Single Day Processing mode
                self.root.after(0, lambda: self.lbl_status.config(text="Calcolo giorno singolo in corso...", foreground=self.warn_color))
                df_res, calc_res = self.calculate_single_day(folder, date_str, pvsyst_pr, threshold)
                
                self.df_result = df_res
                self.calc_results = calc_res
                self.all_days_results = [calc_res]
                
                success_msg = f"Calcolo completato! Creato file giornaliero & aggiornato file Madre."
                self.root.after(0, lambda m=success_msg: self.lbl_status.config(text=m, foreground=self.success_color))
                self.root.after(0, self.update_ui_on_success)
                
        except Exception as e:
            err_msg = str(e)
            self.root.after(0, lambda: self.update_ui_on_failure(err_msg))
        finally:
            quit_excel_app()
            
    def update_ui_on_success(self):
        res = self.calc_results
        
        # Enable controls
        self.btn_calculate.config(state="normal")
        self.btn_export.config(state="normal")
        
        # Display main metrics
        self.lbl_avg_pr_val.config(text=f"{res['avg_inv_pr']:.3f} %")
        self.lbl_comp_pr_val.config(text=f"{res['comp_raw_pr']:.3f} %")
        try:
            val = float(self.pvsyst_pr_var.get())
            self.lbl_pvsyst_target_val.config(text=f"{val * 100:.3f} %")
        except ValueError:
            self.lbl_pvsyst_target_val.config(text="-- %")
        
        self.lbl_irrad_summary.config(
            text=f"Irradiazione giornaliera totale: {res['h_sum_kwh']:.4f} kWh/m² (Media POA > {self.threshold_var.get()} W/m²)"
        )
        
        # Fill treeview (Inverters)
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        for row in res['inverter_table_data']:
            self.tree.insert("", "end", values=row)
            
        # Fill treeview (Days)
        for item in self.tree_days.get_children():
            self.tree_days.delete(item)
            
        for day_res in getattr(self, "all_days_results", []):
            day_row = (
                day_res['date_str'],
                f"{day_res['h_sum_kwh']:.4f}",
                f"{day_res['uncomp_pr']:.3f}",
                f"{day_res['comp_raw_pr']:.3f}",
                f"{day_res['avg_inv_pr']:.3f}"
            )
            self.tree_days.insert("", "end", values=day_row)
            
        # Select appropriate tab
        if hasattr(self, "all_days_results") and len(self.all_days_results) > 1:
            self.notebook.select(1) # Go to Daily Summary tab
        else:
            self.notebook.select(0) # Go to Detailed Inverters tab
            
        self.lbl_status.config(text="Calcolo completato con successo!", foreground=self.success_color)
        
        # Build detailed success report showing PR of each day processed
        msg = f"Calcolo del PR per il {res['date_str']} terminato con successo!\n\n"
        if hasattr(self, "all_days_results") and len(self.all_days_results) > 1:
            msg += "Riepilogo PR Giornaliero (Media PR Inverter):\n"
            for d in self.all_days_results:
                msg += f"- {d['date_str']}: {d['avg_inv_pr']:.3f}%\n"
        else:
            msg += f"Media PR Inverter: {res['avg_inv_pr']:.3f}%"
            
        messagebox.showinfo("Successo", msg)
        
    def update_ui_on_failure(self, error_message):
        self.btn_calculate.config(state="normal")
        self.btn_export.config(state="disabled")
        self.lbl_status.config(text="Calcolo fallito!", foreground="red")
        messagebox.showerror("Errore di Calcolo", error_message)
        
    def export_to_excel(self):
        if self.df_result is None or self.calc_results is None:
            messagebox.showerror("Errore", "Nessun risultato disponibile da esportare!")
            return
            
        # Ask where to save the file
        default_name = f"PR_recalculation_{self.calc_results['date_str'].replace('-', '_')}.xlsx"
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("File Excel", "*.xlsx")],
            initialfile=default_name,
            title="Salva Report PR"
        )
        
        if not save_path:
            return
            
        try:
            self.lbl_export_status.config(text="Esportazione in corso... attendere...", foreground=self.warn_color)
            self.btn_export.config(state="disabled")
            
            # Run export in background thread
            thread = threading.Thread(target=self.run_export, args=(save_path,))
            thread.start()
        except Exception as e:
            self.lbl_export_status.config(text="Esportazione fallita!", foreground="red")
            messagebox.showerror("Errore Esportazione", str(e))
            self.btn_export.config(state="normal")
            
    def run_export(self, save_path):
        try:
            res = self.calc_results
            df_result = self.df_result
            
            # We will export 3 sheets
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                # 1. Summary Sheet
                summary_data = {
                    "Parametro": [
                        "Data di Calcolo",
                        "Irradiazione Giornaliera Totale (kWh/m²)",
                        "Target PR PVSyst Mensile",
                        "Soglia Irraggiamento Minimo (W/m²)",
                        "PR Grezzo Non Compensato (%)",
                        "PR Grezzo Compensato (%)",
                        "Media dei 36 PR Inverter Compensati (%)"
                    ],
                    "Valore": [
                        res["date_str"],
                        res["h_sum_kwh"],
                        float(self.pvsyst_pr_var.get()),
                        float(self.threshold_var.get()),
                        res["uncomp_pr"],
                        res["comp_raw_pr"],
                        res["avg_inv_pr"]
                    ]
                }
                df_sum = pd.DataFrame(summary_data)
                df_sum.to_excel(writer, sheet_name="Riepilogo", index=False)
                
                # 2. Inverter PRs Sheet
                inverter_cols = ["Codice Inverter", "Trasformatore", "Potenza CC Nominale (kW)", "Energia Prodotta (kWh)", "Perdita Stimata (kWh)", "PR Compensato (%)"]
                df_inv_prs = pd.DataFrame(res["inverter_table_data"], columns=inverter_cols)
                df_inv_prs.to_excel(writer, sheet_name="PR_Inverter", index=False)
                
                # 3. Complete Timeslot Data
                cols_to_export = [
                    "time", "poa1", "poa3", "poa1_kwh", "poa3_kwh", "poa_avg_kwh", "poa_avg_w", "h", 
                    "diff_pct", "poa_cond_max_kwh", "limit_ratio", "meter_reading", "active_energy_prod",
                    "TX1_Average_Power", "TX2_Average_Power", "TX3_Average_Power",
                    "TX1_Total_Loss", "TX2_Total_Loss", "TX3_Total_Loss"
                ]
                for inv_id in self.dc_powers:
                    cols_to_export.append(inv_id)
                    cols_to_export.append(f"{inv_id}_loss")
                    
                df_ts = df_result[cols_to_export].copy()
                df_ts.to_excel(writer, sheet_name="Dettaglio_Quarti_Ora", index=False)
                
            # Success UI trigger
            self.root.after(0, lambda: self.on_export_success(save_path))
        except Exception as e:
            self.root.after(0, lambda: self.on_export_failure(str(e)))
            
    def on_export_success(self, save_path):
        self.btn_export.config(state="normal")
        self.lbl_export_status.config(text="Esportazione completata con successo!", foreground=self.success_color)
        messagebox.showinfo("Esportazione Riuscita", f"Report PR dettagliato salvato in:\n\n{save_path}")
        
    def on_export_failure(self, error_message):
        self.btn_export.config(state="normal")
        self.lbl_export_status.config(text="Esportazione fallita!", foreground="red")
        messagebox.showerror("Errore Esportazione", f"Impossibile salvare il file Excel:\n\n{error_message}")

    def on_close(self):
        sys.stdout = sys.__stdout__
        sys.stderr = sys.__stderr__
        self.root.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    app = PRCalculatorGUI(root)
    root.mainloop()
