"""Data Completeness Checker - Mazara 01

Verifica che ogni giorno di un mese sia completo su entrambi gli archivi:
 - Daily Reports: 24 ore di dati SCADA (96 intervalli da 15 min) in tutti i
   file richiesti dal PR Calculator;
 - Tracker: 24 file orari con data e ora coerenti con il nome del file.

Uso:  python Data_Completeness_Checker.py
      python Data_Completeness_Checker.py --selftest
"""
import os
import re
import sys
import csv
import datetime
import threading
import queue
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import openpyxl

# 96 slot da 15 minuti: 00:00:00 -> 23:45:00
EXPECTED_TIMES = [f"{h:02d}:{m:02d}:00" for h in range(24) for m in (0, 15, 30, 45)]
N_INVERTERS = 12

# (label, pattern di ricerca, filtro sul nome segnale Colonna2, raggruppa per inverter)
CHECKS = [
    ("Contatore SATAC", ["SATAC_Meter_15Min.xlsx", "*SATAC*.xlsx"],
     lambda s: s.startswith("Energia attiva prod"), False),
    ("Meteo TS1", ["TS_01_Weather_15Min.xlsx", "*TS_01*Weather*.xlsx"],
     lambda s: s == "POA", False),
    ("Meteo TS3", ["TS_03_Weather_15Min.xlsx", "*TS_03*Weather*.xlsx"],
     lambda s: s == "POA", False),
    ("Inverter TS1", ["TS_01_Inverter_15Min.xlsx", "*TS_01*Inverter*.xlsx"],
     lambda s: s == "Potenza attiva", True),
    ("Inverter TS2", ["TS_02_Inverter_15Min.xlsx", "*TS_02*Inverter*.xlsx"],
     lambda s: s == "Potenza attiva", True),
    ("Inverter TS3", ["TS_03_Inverter_15Min.xlsx", "*TS_03*Inverter*.xlsx"],
     lambda s: s == "Potenza attiva", True),
]
REG_PATTERNS = ["Regolazione_della_potenza_attiva_*.xlsx", "*potenza_attiva*.xlsx"]


def time_key(v):
    """HH:MM:SS da una cella oraria (datetime, time o stringa); None se non e' un orario."""
    if isinstance(v, (datetime.datetime, datetime.time)):
        return v.strftime("%H:%M:%S")
    m = re.match(r"^(\d{1,2}):(\d{2}):(\d{2})", str(v).strip())
    return f"{int(m.group(1)):02d}:{m.group(2)}:{m.group(3)}" if m else None


def date_key(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", str(v).strip())
    return m.group(0) if m else None


def find_file(folder, patterns):
    """Cerca solo file Excel: CSV/TXT vengono ignorati."""
    import glob
    if not os.path.isdir(folder):
        return None
    xlsx = [f for f in os.listdir(folder)
            if f.lower().endswith((".xlsx", ".xlsm", ".xls"))
            and os.path.isfile(os.path.join(folder, f))]
    for pat in patterns:
        hits = glob.glob(os.path.join(folder, pat))
        if hits:
            return hits[0]
        core = pat.replace("*", "").lower()
        for f in xlsx:
            if core in f.lower():
                return os.path.join(folder, f)
    return None


def summarize_gaps(present, label=""):
    """Descrive gli slot mancanti/vuoti rispetto ai 96 attesi."""
    missing = [t for t in EXPECTED_TIMES if t not in present]
    if not missing:
        return None
    head = ", ".join(t[:5] for t in missing[:8])
    more = f" ... (+{len(missing) - 8})" if len(missing) > 8 else ""
    pre = f"{label}: " if label else ""
    return f"{pre}{len(missing)}/96 intervalli mancanti [{head}{more}]"


def read_signal(path, sig_filter, by_inverter, expected_date):
    """Legge il file SCADA e restituisce (mappa slot presenti, date trovate).

    Gli slot sono contati SOLO per le righe della data attesa, cosi' un file
    del giorno sbagliato non viene mai scambiato per un giorno completo.
    La mappa e' {chiave: set(HH:MM:SS)} - una sola chiave "" se non si
    raggruppa per inverter. Un valore vuoto (None) conta come dato mancante.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    present, dates = {}, set()
    try:
        for row in ws.iter_rows(min_row=1, max_col=6, values_only=True):
            sig = row[1]
            if not isinstance(sig, str) or not sig_filter(sig.strip()):
                continue
            t = time_key(row[5])
            if t is None:
                continue
            d = date_key(row[4])
            if d:
                dates.add(d)
                if d != expected_date:
                    continue        # riga di un altro giorno: non conta come dato del giorno
            if row[2] is None:      # slot presente ma senza valore
                continue
            key = ""
            if by_inverter:
                m = re.search(r"Inverter_?(\d+)", str(row[0]))
                key = f"INV-{int(m.group(1)):02d}" if m else "INV-?"
            present.setdefault(key, set()).add(t)
    finally:
        wb.close()
    return present, dates


def filename_date(path):
    """Data YYYY-MM-DD estratta dal nome file (es. ..._2026_05_01.xlsx), se presente."""
    m = re.search(r"(20\d{2})[ _-](\d{2})[ _-](\d{2})", os.path.basename(path))
    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}" if m else None


def date_problem(dates, expected_date):
    """Descrive un disallineamento fra le date dentro al file e il giorno atteso."""
    if not dates:
        return None
    others = sorted(d for d in dates if d != expected_date)
    if expected_date not in dates:
        return f"FILE DEL GIORNO SBAGLIATO: contiene {', '.join(others)} invece di {expected_date}"
    if others:
        return f"contiene anche righe di altre date: {', '.join(others)}"
    return None


def check_regulation(path):
    """Il file di regolazione e' su griglia a 5 minuti: verifica la copertura 24h."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    times = set()
    try:
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row or row[-1] is None:
                continue
            t = time_key(row[0]) or (time_key(row[1]) if len(row) > 1 else None)
            if t:
                times.add(t)
    finally:
        wb.close()
    if not times:
        return "nessun dato orario leggibile"
    step = 5 if len(times) > 150 else 15
    expected = [f"{h:02d}:{m:02d}:00" for h in range(24) for m in range(0, 60, step)]
    missing = [t for t in expected if t not in times]
    if missing:
        head = ", ".join(t[:5] for t in missing[:8])
        more = f" ... (+{len(missing) - 8})" if len(missing) > 8 else ""
        return f"{len(missing)}/{len(expected)} intervalli da {step} min mancanti [{head}{more}]"
    return None


def check_day(day_folder, expected_date):
    """Restituisce la lista dei problemi del giorno: (file, problema).

    Solo file Excel: CSV e TXT (inclusa la sottocartella CSV/) sono ignorati.
    """
    problems = []
    found = {label: find_file(day_folder, pats) for label, pats, _, _ in CHECKS}
    if not any(found.values()):
        # Nessun file SCADA: una riga sola invece di sei identiche.
        problems.append(("SCADA", f"tutti i {len(CHECKS)} file Excel SCADA assenti"))

    for label, patterns, sig_filter, by_inv in CHECKS:
        path = found[label]
        if not path:
            if not any(found.values()):
                continue
            problems.append((label, "FILE ASSENTE"))
            continue
        try:
            present, dates = read_signal(path, sig_filter, by_inv, expected_date)
        except Exception as e:
            problems.append((label, f"errore di lettura: {e}"))
            continue
        fn_date = filename_date(path)
        if fn_date and fn_date != expected_date:
            problems.append((label, f"nome file di un altro giorno ({fn_date}): {os.path.basename(path)}"))
        dp = date_problem(dates, expected_date)
        if dp:
            problems.append((label, dp))
        if expected_date not in dates and dates:
            continue        # file del giorno sbagliato: inutile elencare 96 slot mancanti
        if not present:
            problems.append((label, "nessuna riga del segnale atteso nel file"))
            continue
        if by_inv:
            if len(present) != N_INVERTERS:
                problems.append((label, f"{len(present)}/{N_INVERTERS} inverter presenti "
                                        f"(mancanti: {N_INVERTERS - len(present)})"))
            # Un buco comune a tutti gli inverter e' un solo problema, non 12.
            if len(set(map(frozenset, present.values()))) == 1:
                gap = summarize_gaps(next(iter(present.values())), f"tutti i {len(present)} inverter")
                if gap:
                    problems.append((label, gap))
            else:
                for key in sorted(present):
                    gap = summarize_gaps(present[key], key)
                    if gap:
                        problems.append((label, gap))
        else:
            gap = summarize_gaps(present[""])
            if gap:
                problems.append((label, gap))

    reg = find_file(day_folder, REG_PATTERNS)
    if not reg:
        problems.append(("Regolazione Potenza Attiva", "FILE ASSENTE"))
    else:
        # Il file di regolazione contiene solo orari: la data sta nel nome file.
        fn_date = filename_date(reg)
        if fn_date and fn_date != expected_date:
            problems.append(("Regolazione Potenza Attiva",
                             f"FILE DEL GIORNO SBAGLIATO ({fn_date}): {os.path.basename(reg)}"))
        try:
            issue = check_regulation(reg)
        except Exception as e:
            issue = f"errore di lettura: {e}"
        if issue:
            problems.append(("Regolazione Potenza Attiva", issue))
    return problems


# ---------------------------------------------------------------- TRACKER
# Un giorno = 24 file orari CSV/TXT (UTF-16LE, ";") da ~550 MB l'uno:
# si leggono solo tre finestre per file (inizio, meta', fine), mai tutto.
TRACKER_NAME = re.compile(r"^(\d{2})_(\d{2})_(\d{4})_(\d{2})_(\d{2})", re.I)
TRACKER_TS = re.compile(r"(\d{2})/(\d{2})/(\d{4})\s+(\d{2}):(\d{2}):(\d{2})")
CHUNK = 96 * 1024       # ~400 righe: basta per il blocco del primo segnale (61 righe)


def tracker_ts(line):
    """Data/ora dall'ultima colonna della riga (DD/MM/YYYY HH:MM:SS.mmm)."""
    m = TRACKER_TS.search(line.rsplit(";", 1)[-1])
    if not m:
        return None
    d, mo, y, h, mi, s = (int(x) for x in m.groups())
    try:
        return datetime.datetime(y, mo, d, h, mi, s)
    except ValueError:
        return None


def read_chunk(path, offset, size=CHUNK):
    """Righe decodificate da una finestra del file (offset negativo = dalla fine)."""
    with open(path, "rb") as f:
        f.seek(0, 2)
        total = f.tell()
        pos = max(0, total + offset) if offset < 0 else min(offset, total)
        f.seek(pos - (pos % 2))     # UTF-16: sempre su confine di carattere
        raw = f.read(size)
    return raw.decode("utf-16-le", errors="replace").replace("\ufeff", "").splitlines()


def missing_hours(spans):
    """Ore 0-23 non coperte dai file, con spans = [(ora_inizio, ora_fine), ...].

    Un file 01_03 (ora legale di marzo) copre 01 e 02: nessun buco segnalato.
    """
    covered = set()
    for h0, h1 in spans:
        span = (h1 - h0) % 24 or 24
        covered |= {(h0 + i) % 24 for i in range(span)}
    return sorted(set(range(24)) - covered)


def check_tracker_file(path, start, end):
    """Problemi del singolo file orario rispetto all'intervallo atteso [start, end]."""
    problems = []
    rows = [l for l in read_chunk(path, 0) if ";" in l]
    if not rows:
        return ["file vuoto o illeggibile"]
    first = tracker_ts(rows[0])
    if first is None:
        return ["prima riga senza data/ora leggibile"]
    if first != start:
        problems.append(f"inizia alle {first:%d/%m/%Y %H:%M} invece di {start:%d/%m/%Y %H:%M}")

    # Il file e' ordinato per segnale: il primo blocco copre da solo tutta l'ora.
    sig = rows[0].split(";")[0]
    block = []
    for l in rows:
        if l.split(";")[0] != sig:
            break
        t = tracker_ts(l)
        if t:
            block.append(t)
    have = {t for t in block if t < end}
    want = [start + datetime.timedelta(minutes=i) for i in range(60)]
    lost = [t for t in want if t not in have]
    if lost and len(block) > 1:     # se il blocco non e' entrato nella finestra letta, non concludo
        head = ", ".join(f"{t:%H:%M}" for t in lost[:8])
        more = f" ... (+{len(lost) - 8})" if len(lost) > 8 else ""
        problems.append(f"{len(lost)}/60 minuti mancanti [{head}{more}]")

    # Campione a meta' file: intercetta blocchi di un'altra ora incollati dentro.
    with open(path, "rb") as f:
        f.seek(0, 2)
        size = f.tell()
    if size > 4 * CHUNK:
        out = [t for t in (tracker_ts(l) for l in read_chunk(path, size // 2)[1:])
               if t and not start <= t <= end]
        if out:
            problems.append(f"dati fuori intervallo a meta' file (es. {out[0]:%d/%m/%Y %H:%M})")

    tail = [t for t in (tracker_ts(l) for l in read_chunk(path, -CHUNK)) if t]
    if not tail:
        problems.append("ultima riga senza data/ora leggibile")
    elif tail[-1] != end:
        problems.append(f"finisce alle {tail[-1]:%d/%m/%Y %H:%M} invece di {end:%d/%m/%Y %H:%M}")
    return problems


def check_tracker_day(day_folder, expected_date):
    """Problemi del giorno tracker: (file, problema)."""
    day = datetime.datetime.strptime(expected_date, "%Y-%m-%d")
    problems, files = [], {}
    for name in sorted(os.listdir(day_folder)):
        if not name.lower().endswith((".csv", ".txt")):
            continue
        m = TRACKER_NAME.match(name)
        if not m:
            problems.append((name, "nome file non riconosciuto"))
            continue
        d, mo, y, h0, h1 = (int(x) for x in m.groups())
        if (y, mo, d) != (day.year, day.month, day.day):
            problems.append((name, f"nome file di un altro giorno ({d:02d}/{mo:02d}/{y})"))
            continue
        if h0 > 23 or h1 > 23:
            problems.append((name, f"ore non valide nel nome file ({h0:02d}-{h1:02d})"))
            continue
        files.setdefault(h0, []).append((os.path.join(day_folder, name), h1))

    if not files:
        return problems + [("-", "nessun file orario presente")]

    lost = missing_hours([(h0, h1) for h0, e in files.items() for _, h1 in e])
    if lost:
        problems.append(("-", f"mancano {len(lost)}/24 ore: "
                              + ", ".join(f"{h:02d}" for h in lost)))
    for h0, entries in sorted(files.items()):
        if len(entries) > 1:
            problems.append((f"ora {h0:02d}", "file duplicati: "
                             + ", ".join(os.path.basename(p) for p, _ in entries)))

    paths = [(h0, p, h1) for h0, e in sorted(files.items()) for p, h1 in e]
    sizes = {p: os.path.getsize(p) for _, p, _ in paths}
    # ponytail: euristica sulla dimensione, unico modo di vedere un export troncato
    # senza rileggere 550 MB; se un giorno intero e' parziale non se ne accorge.
    median = sorted(sizes.values())[len(sizes) // 2]
    for h0, path, h1 in paths:
        name = os.path.basename(path)
        if sizes[path] < median * 0.5:
            problems.append((name, f"file molto piu' piccolo degli altri "
                                   f"({sizes[path] // 2**20} MB contro {median // 2**20} MB): export parziale?"))
        start = day + datetime.timedelta(hours=h0)
        end = start + datetime.timedelta(hours=(h1 - h0) % 24 or 24)
        try:
            issues = check_tracker_file(path, start, end)
        except Exception as e:
            issues = [f"errore di lettura: {e}"]
        problems += [(name, i) for i in issues]
    return problems


# Archivio dati giornalieri (fallback: la cartella dello script)
DEFAULT_ROOT = r"\\s01\get\2025.01 Mazara 01 A2A\03 - REPORT\Report\01 Daily Reports"
TRACKER_ROOT = r"\\s01\get\2025.01 Mazara 01 A2A\03 - REPORT\Report\04 Tracker report\01_Original_files"
ALL_DAYS = "Tutto il mese"


def tracker_month_path(base, year=None, mon=None):
    """Cartella mese del tracker (base\\YYYY\\MM); l'ultima disponibile se anno/mese mancano."""
    if not os.path.isdir(base):
        return None
    if year and mon:
        p = os.path.join(base, year, mon)
        return p if os.path.isdir(p) else None
    months = [(y, m)
              for y in os.listdir(base)
              if re.fullmatch(r"20\d{2}", y) and os.path.isdir(os.path.join(base, y))
              for m in os.listdir(os.path.join(base, y))
              if re.fullmatch(r"\d{2}", m) and os.path.isdir(os.path.join(base, y, m))]
    return os.path.join(base, *max(months)) if months else None


class App:
    BG, FG, ACC = "#1e1e28", "#e6e6ef", "#4a9eff"
    FIELD, PANEL = "#2a2a38", "#33334a"

    def __init__(self, root):
        self.root = root
        root.title("Verifica Completezza Dati v2 - Mazara 01")
        root.geometry("1000x640")
        root.configure(bg=self.BG)
        try:
            base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
            root.iconbitmap(os.path.join(base, "assets", "logo.ico"))
        except Exception:
            pass
        self.q = queue.Queue()

        st = ttk.Style()
        st.theme_use("clam")
        st.configure(".", background=self.BG, foreground=self.FG,
                     fieldbackground=self.FIELD, bordercolor=self.PANEL, lightcolor=self.PANEL,
                     darkcolor=self.PANEL, focuscolor=self.ACC)
        st.configure("TLabel", background=self.BG, foreground=self.FG)
        st.configure("TFrame", background=self.BG)
        st.configure("TEntry", foreground=self.FG, fieldbackground=self.FIELD,
                     insertcolor=self.FG, selectbackground=self.ACC, selectforeground="#ffffff")
        st.configure("TButton", background=self.PANEL, foreground=self.FG, borderwidth=0, padding=5)
        st.map("TButton",
               background=[("active", self.ACC), ("disabled", self.FIELD)],
               foreground=[("active", "#ffffff"), ("disabled", "#7a7a8c")])
        st.configure("TCombobox", foreground=self.FG, fieldbackground=self.FIELD,
                     background=self.PANEL, arrowcolor=self.FG, selectbackground=self.FIELD,
                     selectforeground=self.FG)
        st.map("TCombobox",
               fieldbackground=[("readonly", self.FIELD)],
               foreground=[("readonly", self.FG), ("disabled", "#7a7a8c")],
               selectbackground=[("readonly", self.FIELD)],
               selectforeground=[("readonly", self.FG)])
        # La tendina del Combobox e' un Listbox Tk classico: si colora solo via option_add
        root.option_add("*TCombobox*Listbox.background", self.FIELD)
        root.option_add("*TCombobox*Listbox.foreground", self.FG)
        root.option_add("*TCombobox*Listbox.selectBackground", self.ACC)
        root.option_add("*TCombobox*Listbox.selectForeground", "#ffffff")
        st.configure("Horizontal.TProgressbar", background=self.ACC, troughcolor=self.FIELD,
                     bordercolor=self.FIELD, lightcolor=self.ACC, darkcolor=self.ACC)
        st.configure("Treeview", background=self.FIELD, fieldbackground=self.FIELD,
                     foreground=self.FG, rowheight=22, borderwidth=0)
        st.configure("Treeview.Heading", background=self.PANEL, foreground=self.FG,
                     relief="flat", padding=4)
        st.map("Treeview.Heading", background=[("active", self.ACC)], foreground=[("active", "#ffffff")])
        st.map("Treeview", background=[("selected", self.ACC)], foreground=[("selected", "#ffffff")])

        top = ttk.Frame(root, padding=10)
        top.pack(fill="x")
        ttk.Label(top, text="Daily Reports:").grid(row=0, column=0, sticky="w")
        self.root_var = tk.StringVar(
            value=DEFAULT_ROOT if os.path.isdir(DEFAULT_ROOT)
            else os.path.dirname(os.path.abspath(__file__)))
        ttk.Entry(top, textvariable=self.root_var, width=70).grid(row=0, column=1, padx=6, sticky="we")
        ttk.Button(top, text="Sfoglia...",
                   command=lambda: self.browse(self.root_var)).grid(row=0, column=2)

        ttk.Label(top, text="Tracker:").grid(row=1, column=0, sticky="w", pady=(8, 0))
        self.tracker_var = tk.StringVar(value=TRACKER_ROOT)
        ttk.Entry(top, textvariable=self.tracker_var, width=70).grid(row=1, column=1, padx=6,
                                                                    pady=(8, 0), sticky="we")
        ttk.Button(top, text="Sfoglia...",
                   command=lambda: self.browse(self.tracker_var)).grid(row=1, column=2, pady=(8, 0))

        sel = ttk.Frame(top)
        sel.grid(row=2, column=1, padx=6, pady=(8, 0), sticky="w")
        ttk.Label(top, text="Mese:").grid(row=2, column=0, sticky="w", pady=(8, 0))
        self.month_cb = ttk.Combobox(sel, state="readonly", width=14)
        self.month_cb.pack(side="left")
        ttk.Label(sel, text="Giorno:").pack(side="left", padx=(12, 6))
        self.day_cb = ttk.Combobox(sel, state="readonly", width=16)
        self.day_cb.pack(side="left")
        self.btn = ttk.Button(top, text="Analizza", command=self.start)
        self.btn.grid(row=2, column=2, pady=(8, 0))
        top.columnconfigure(1, weight=1)

        self.pb = ttk.Progressbar(root, mode="determinate")
        self.pb.pack(fill="x", padx=10)
        self.status = ttk.Label(root, text="Seleziona un mese e avvia l'analisi.", padding=(10, 6))
        self.status.pack(fill="x")

        cols = ("giorno", "fonte", "file", "problema")
        self.tree = ttk.Treeview(root, columns=cols, show="headings")
        for c, w in zip(cols, (60, 80, 220, 600)):
            self.tree.heading(c, text=c.capitalize())
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=10, pady=(0, 6))
        self.tree.tag_configure("ok", foreground="#5fd18a")
        self.tree.tag_configure("bad", foreground="#ff8a7a")

        bot = ttk.Frame(root, padding=(10, 0, 10, 10))
        bot.pack(fill="x")
        ttk.Button(bot, text="Esporta CSV", command=self.export).pack(side="right")

        self.months = {}
        self.root_var.trace_add("write", lambda *_: self.refresh_months())
        self.tracker_var.trace_add("write", lambda *_: self.refresh_days())
        self.month_cb.bind("<<ComboboxSelected>>", lambda *_: self.on_month())
        self.refresh_months()
        self.tracker_var.set(tracker_month_path(TRACKER_ROOT) or TRACKER_ROOT)

    def browse(self, var):
        d = filedialog.askdirectory(initialdir=var.get())
        if d:
            var.set(d)

    def refresh_months(self):
        """Mesi disponibili nei Daily Reports (cartelle 'YYYY MM')."""
        base, self.months = self.root_var.get(), {}
        if os.path.isdir(base):
            for n in sorted(os.listdir(base)):
                if re.fullmatch(r"\d{4}[ _-]\d{2}", n) and os.path.isdir(os.path.join(base, n)):
                    self.months[n] = os.path.join(base, n)
        names = list(self.months)
        self.month_cb["values"] = names
        if names and self.month_cb.get() not in names:
            self.month_cb.set(names[-1])
        elif not names:
            self.month_cb.set("")
        self.on_month()

    def on_month(self):
        """Il mese del tracker segue quello selezionato, se esiste sotto la radice."""
        month = self.month_cb.get()
        if month:
            year, mon = re.split(r"[-_ ]", month)
            same = tracker_month_path(TRACKER_ROOT, year, mon)
            if same and same != self.tracker_var.get():
                self.tracker_var.set(same)      # il trace ricarica i giorni
                return
        self.refresh_days()

    def day_list(self):
        """[(giorno, cartella SCADA o None, cartella tracker o None, data attesa)].

        Un giorno compare se e' presente in almeno uno dei due archivi.
        """
        month = self.month_cb.get()
        if not month:
            return []
        year, mon = re.split(r"[-_ ]", month)
        scada, tracker = {}, {}
        month_path = self.months.get(month)
        if month_path and os.path.isdir(month_path):
            scada = {n: os.path.join(month_path, n) for n in os.listdir(month_path)
                     if re.fullmatch(r"\d{2}", n) and os.path.isdir(os.path.join(month_path, n))}
        tpath = self.tracker_var.get()
        if os.path.isdir(tpath):
            # Solo 'YYYY-MM-DD' esatto: fuori restano '... test' e 'vecchie ...'.
            tracker = {n[-2:]: os.path.join(tpath, n) for n in os.listdir(tpath)
                       if re.fullmatch(rf"{year}-{mon}-\d{{2}}", n)
                       and os.path.isdir(os.path.join(tpath, n))}
        return [(d, scada.get(d), tracker.get(d), f"{year}-{mon}-{d}")
                for d in sorted(set(scada) | set(tracker))]

    def refresh_days(self):
        self.days = self.day_list()
        keep = self.day_cb.get()
        labels = [d[0] for d in self.days]
        self.day_cb["values"] = [ALL_DAYS] + labels
        self.day_cb.set(keep if keep in labels else ALL_DAYS)

    def start(self):
        if not self.month_cb.get():
            messagebox.showwarning("Mese mancante", "Nessuna cartella mese trovata nei Daily Reports.")
            return
        self.days = self.day_list()
        days = self.days
        if self.day_cb.get() != ALL_DAYS:
            days = [d for d in days if d[0] == self.day_cb.get()]
        if not days:
            messagebox.showwarning("Nessun giorno", "Nessuna cartella giorno trovata per il mese selezionato.")
            return
        self.tree.delete(*self.tree.get_children())
        self.btn.config(state="disabled")
        threading.Thread(target=self.scan, args=(days,), daemon=True).start()
        self.root.after(100, self.pump)

    def scan(self, days):
        self.q.put(("total", len(days)))
        for i, (label, scada_path, tracker_path, expected) in enumerate(days, 1):
            problems = []
            for source, path, check in (("SCADA", scada_path, check_day),
                                        ("Tracker", tracker_path, check_tracker_day)):
                self.q.put(("status", f"Analisi {source} {expected} ({i}/{len(days)})..."))
                if not path:
                    problems.append((source, "-", "cartella del giorno assente"))
                    continue
                try:
                    found = check(path, expected)
                except Exception as e:
                    found = [("-", f"errore: {e}")]
                problems += [(source, f, p) for f, p in found]
            self.q.put(("day", (label, problems)))
            self.q.put(("progress", i))
        self.q.put(("done", None))

    def pump(self):
        try:
            while True:
                kind, payload = self.q.get_nowait()
                if kind == "total":
                    self.pb.config(maximum=max(payload, 1), value=0)
                    self.bad_days = 0
                elif kind == "progress":
                    self.pb.config(value=payload)
                elif kind == "status":
                    self.status.config(text=payload)
                elif kind == "day":
                    day, problems = payload
                    if problems:
                        self.bad_days += 1
                        for s, f, p in problems:
                            self.tree.insert("", "end", values=(day, s, f, p), tags=("bad",))
                    else:
                        self.tree.insert("", "end", values=(day, "SCADA + Tracker", "-",
                                                            "OK - 24h complete"), tags=("ok",))
                elif kind == "done":
                    self.btn.config(state="normal")
                    self.status.config(text=f"Analisi completata. Giorni con dati mancanti: {self.bad_days}.")
                    return
        except queue.Empty:
            pass
        self.root.after(100, self.pump)

    def export(self):
        rows = [self.tree.item(i, "values") for i in self.tree.get_children()]
        if not rows:
            return
        path = filedialog.asksaveasfilename(defaultextension=".csv",
                                            initialfile=f"completezza_{self.month_cb.get().replace(' ', '_')}.csv")
        if not path:
            return
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(["Giorno", "Fonte", "File", "Problema"])
            w.writerows(rows)
        messagebox.showinfo("Esportato", f"Report salvato in:\n{path}")


def selftest():
    assert len(EXPECTED_TIMES) == 96 and EXPECTED_TIMES[-1] == "23:45:00"
    assert time_key("00:15:00.000") == "00:15:00"
    assert time_key(datetime.time(9, 30)) == "09:30:00"
    assert time_key("n/d") is None
    assert date_key(datetime.datetime(2026, 5, 1)) == "2026-05-01"
    assert summarize_gaps(set(EXPECTED_TIMES)) is None
    gap = summarize_gaps(set(EXPECTED_TIMES) - {"12:00:00", "12:15:00"}, "INV-01")
    assert gap.startswith("INV-01: 2/96 intervalli mancanti") and "12:00" in gap
    assert "+3)" in summarize_gaps(set(EXPECTED_TIMES[:85]))  # 11 mancanti -> 8 mostrati + 3
    assert filename_date("Regolazione_della_potenza_attiva_2026_05_01.xlsx") == "2026-05-01"
    assert filename_date("SATAC_Meter_15Min.xlsx") is None
    assert date_problem({"2026-05-01"}, "2026-05-01") is None
    assert date_problem(set(), "2026-05-01") is None
    assert date_problem({"2026-04-30"}, "2026-05-01").startswith("FILE DEL GIORNO SBAGLIATO")
    assert "altre date" in date_problem({"2026-05-01", "2026-05-02"}, "2026-05-01")

    # --- tracker ---
    assert tracker_ts("sig; Potenza;1.0;kW;SPONT;15/07/2026 09:00:00.000") == datetime.datetime(2026, 7, 15, 9, 0)
    assert tracker_ts("sig;;;;;n/d") is None
    assert TRACKER_NAME.match("15_07_2026_23_00.csv").groups() == ("15", "07", "2026", "23", "00")
    assert TRACKER_NAME.match("12_05_2026_00_01 (Mancano Dati).TXT").groups()[3:] == ("00", "01")
    full = [(h, (h + 1) % 24) for h in range(24)]
    assert missing_hours(full) == []
    assert missing_hours([s for s in full if s[0] != 5]) == [5]
    # ora legale di marzo: il file 01_03 copre 01 e 02
    dst = [s for s in full if s[0] not in (1, 2)] + [(1, 3)]
    assert missing_hours(dst) == []
    assert missing_hours([(23, 0)]) == list(range(23))
    print("selftest OK")


if __name__ == "__main__":
    if "--selftest" in sys.argv:
        selftest()
    else:
        r = tk.Tk()
        App(r)
        r.mainloop()
